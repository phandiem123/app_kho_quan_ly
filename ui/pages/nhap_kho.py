from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLineEdit, QComboBox, QMessageBox, QMenu, QGridLayout,
    QFileDialog,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QCursor, QAction
from database.receipts import Receipt, ReceiptLine, get_all as receipt_get_all, get_lines
from database.receipts import delete as receipt_delete

FONT = "Segoe UI"

_TABS = ["Nhập Hàng Mới", "Nhập Từ Đơn Vị", "Nhập Hàng Dùng Chung"]
_SUBTYPES = ["new", "from_unit", "shared_return"]

_COLS = [
    "STT", "Số Phiếu", "Số M.Hàng", "Kho Nhập",
    "Đơn Vị Giao", "Người giao", "Ngày Nhập", "Nội dung", "Vận Chuyển", "",
]

_TABLE_STYLE = """
    QTableWidget { border: none; background: white; outline: 0; gridline-color: transparent; }
    QHeaderView::section {
        background: white; color: #111; font-size: 12px;
        border: none; border-bottom: 1px solid #efefef;
        padding: 8px 12px; font-weight: normal;
    }
    QTableWidget::item {
        padding: 0 12px; border: none;
        color: #111; font-size: 13px;
    }
    QTableWidget::item:selected { background: #f0f4ff; color: #111; }
"""

_BTN_GRAY = """
    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
        background: #555; color: white; font-size: 12px; font-weight: 600; }
    QPushButton:hover { background: #444; }
"""
_BTN_DARK = """
    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
        background: #111; color: white; font-size: 12px; font-weight: 600; }
    QPushButton:hover { background: #333; }
"""


# ── Detail panel ──────────────────────────────────────────────────────────────
class DetailPanel(QWidget):
    def __init__(self, receipt: Receipt, lines: list[ReceiptLine], parent=None):
        super().__init__(parent)
        self.setAutoFillBackground(True)
        self.setStyleSheet("""
            DetailPanel { background: #f0f0f0; }
            QLabel { background: transparent; border: none; }
        """)

        root = QVBoxLayout(self)
        root.setContentsMargins(28, 14, 28, 14)
        root.setSpacing(10)

        self._receipt = receipt
        self._lines = lines

        top = QHBoxLayout()
        title = QLabel("Chi Tiết Phiếu nhập")
        title.setFont(QFont(FONT, 13, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        top.addWidget(title)
        top.addStretch()

        btn_export = QPushButton("Xuất Chi Tiết")
        btn_export.setFixedHeight(30)
        btn_export.setFont(QFont(FONT, 11))
        btn_export.setStyleSheet(_BTN_GRAY)
        btn_export.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_export.clicked.connect(self._export_excel)
        top.addSpacing(6)
        top.addWidget(btn_export)

        btn_print = QPushButton("Xuất Phiếu Nhập")
        btn_print.setFixedHeight(30)
        btn_print.setFont(QFont(FONT, 11))
        btn_print.setStyleSheet(_BTN_DARK)
        btn_print.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_print.clicked.connect(self._export_word)
        top.addSpacing(6)
        top.addWidget(btn_print)
        root.addLayout(top)

        grid = QGridLayout()
        grid.setSpacing(4)
        grid.setHorizontalSpacing(24)

        def pair(label, value):
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            hl = QHBoxLayout(w)
            hl.setContentsMargins(0, 0, 0, 0)
            hl.setSpacing(4)
            l = QLabel(label + ":")
            l.setFont(QFont(FONT, 11))
            l.setStyleSheet("color: #111;")
            l.setFixedWidth(90)
            v = QLabel(value or "—")
            v.setFont(QFont(FONT, 11, QFont.Weight.Medium))
            v.setStyleSheet("color: #111;")
            v.setWordWrap(True)
            hl.addWidget(l)
            hl.addWidget(v, 1)
            return w

        source = receipt.from_warehouse_name or receipt.supplier
        source_label = {
            "from_unit":    "Đơn Vị Giao",
            "unit_return":  "Đơn Vị Trả",
            "event_return": "Nguồn Trả",
        }.get(receipt.subtype, "Đơn Vị Giao")

        grid.addWidget(pair("Số Phiếu", receipt.reference_number), 0, 0)
        grid.addWidget(pair("Kho Nhập", receipt.to_warehouse_name), 0, 1)
        grid.addWidget(pair(source_label, source), 0, 2)
        grid.addWidget(pair("Người Giao", receipt.created_by), 1, 0)
        grid.addWidget(pair("Ngày Nhập", receipt.transaction_date), 1, 1)
        if receipt.subtype == "new":
            grid.addWidget(pair("Vận Chuyển", receipt.transporter), 1, 2)
        grid.addWidget(pair("Nội dung", receipt.notes), 2, 0, 1, 3)
        root.addLayout(grid)

        sub_lbl = QLabel("Danh Sách Mặt hàng")
        sub_lbl.setFont(QFont(FONT, 11, QFont.Weight.Bold))
        sub_lbl.setStyleSheet("color: #111;")
        root.addWidget(sub_lbl)

        show_price   = (receipt.subtype == "new")
        show_quality = (receipt.subtype == "from_unit")
        sub_cols = ["STT", "Tên Hàng", "ĐVT", "Số lượng"]
        if show_quality:
            sub_cols.append("Mức HH")
        if show_price:
            sub_cols += ["Đơn Giá", "Thành Tiền"]
        sub_cols.append("Ghi Chú")

        self._sub = QTableWidget(0, len(sub_cols))
        self._sub.setHorizontalHeaderLabels(sub_cols)
        self._sub.verticalHeader().setVisible(False)
        self._sub.setShowGrid(False)
        self._sub.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sub.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._sub.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self._sub.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._sub.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._sub.setStyleSheet("""
            QTableWidget { border: none; background: #f8f8f8; outline: 0;
                border-radius: 6px; }
            QHeaderView::section {
                background: #f0f0f0; color: #777; font-size: 10px;
                border: none; padding: 4px 8px;
            }
            QTableWidget::item { padding: 4px 8px; color: #111; font-size: 12px; border: none; }
        """)
        sh = self._sub.horizontalHeader()
        modes = [
            (QHeaderView.ResizeMode.Fixed, 44),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 72),
            (QHeaderView.ResizeMode.Fixed, 84),
        ]
        if show_quality:
            modes.append((QHeaderView.ResizeMode.Fixed, 72))
        if show_price:
            modes += [
                (QHeaderView.ResizeMode.Fixed, 100),
                (QHeaderView.ResizeMode.Fixed, 110),
            ]
        modes.append((QHeaderView.ResizeMode.Stretch, None))
        for i, (mode, w) in enumerate(modes):
            sh.setSectionResizeMode(i, mode)
            if w:
                self._sub.setColumnWidth(i, w)
        root.addWidget(self._sub)

        if show_price and lines:
            total = sum(l.quantity * l.unit_price for l in lines)
            if total:
                tw = QWidget()
                tw.setStyleSheet("background: transparent;")
                th = QHBoxLayout(tw)
                th.setContentsMargins(0, 4, 0, 0)
                th.addStretch()
                tl = QLabel(f"Tổng tiền: {total:,.0f} đ")
                tl.setFont(QFont(FONT, 12, QFont.Weight.Bold))
                tl.setStyleSheet("color: #111;")
                th.addWidget(tl)
                root.addWidget(tw)

        self._load_lines(lines, show_price, show_quality)

    def _load_lines(self, lines: list[ReceiptLine], show_price: bool, show_quality: bool = False):
        self._sub.setRowCount(0)
        center_cols = {0, 3}
        col_offset = 4
        if show_quality:
            center_cols.add(col_offset)
            col_offset += 1
        if show_price:
            center_cols.add(col_offset)
            center_cols.add(col_offset + 1)

        for i, line in enumerate(lines):
            r = self._sub.rowCount()
            self._sub.insertRow(r)
            self._sub.setRowHeight(r, 38)
            cells = [str(i + 1), line.item_name,
                     line.unit_of_measure, str(line.quantity)]
            if show_quality:
                cells.append(line.quality_level)
            if show_price:
                total = line.quantity * line.unit_price
                cells += [
                    f"{line.unit_price:,.0f}" if line.unit_price else "—",
                    f"{total:,.0f}" if total else "—",
                ]
            cells.append(line.notes)
            for c, val in enumerate(cells):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 11))
                if c in center_cols:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._sub.setItem(r, c, cell)
        self._sub.setFixedHeight(36 + max(1, len(lines)) * 38)

    def _export_excel(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
        from datetime import date as _date

        r = self._receipt
        lines = self._lines
        show_price = (r.subtype == "new")
        show_quality = (r.subtype == "from_unit")

        default_name = f"PhieuNhap_{r.reference_number or 'export'}_{_date.today().isoformat()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu file Excel", default_name,
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chi Tiết Phiếu Nhập"

        # ── header info block ──────────────────────────────────────────────
        hdr_font = XFont(name="Calibri", bold=True, size=11)
        val_font = XFont(name="Calibri", size=11)

        def write_pair(row, label, value):
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = hdr_font
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = val_font

        write_pair(1, "Số Phiếu:", r.reference_number)
        write_pair(2, "Kho Nhập:", r.to_warehouse_name)
        source_lbl = {"from_unit": "Đơn Vị Giao", "shared_return": "Đơn Vị Trả"}.get(r.subtype, "Đơn Vị Giao")
        write_pair(3, f"{source_lbl}:", r.from_warehouse_name or r.supplier)
        write_pair(4, "Người Giao:", r.created_by)
        write_pair(5, "Ngày Nhập:", r.transaction_date)
        if r.subtype == "new":
            write_pair(6, "Vận Chuyển:", r.transporter)
        write_pair(7, "Nội dung:", r.notes)

        # ── column headers ─────────────────────────────────────────────────
        tbl_row = 9
        cols = ["STT", "Tên Hàng", "ĐVT", "Số Lượng"]
        if show_quality:
            cols.append("Mức HH")
        if show_price:
            cols += ["Đơn Giá", "Thành Tiền"]
        cols.append("Ghi Chú")

        fill_hdr = PatternFill("solid", fgColor="333333")
        th_font = XFont(name="Calibri", bold=True, size=11, color="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)

        for ci, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=tbl_row, column=ci, value=col_name)
            cell.font = th_font
            cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── data rows ──────────────────────────────────────────────────────
        for i, line in enumerate(lines):
            dr = tbl_row + 1 + i
            row_vals = [i + 1, line.item_name,
                        line.unit_of_measure, line.quantity]
            if show_quality:
                row_vals.append(line.quality_level)
            if show_price:
                row_vals += [line.unit_price, line.quantity * line.unit_price]
            row_vals.append(line.notes)

            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=dr, column=ci, value=val)
                cell.font = val_font
                cell.border = border
                if ci in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center")
                if show_price and ci in (len(cols) - 1, len(cols)):
                    cell.number_format = '#,##0'

        # ── total row ──────────────────────────────────────────────────────
        if show_price and lines:
            total = sum(l.quantity * l.unit_price for l in lines)
            total_row = tbl_row + 1 + len(lines) + 1
            tc = ws.cell(row=total_row, column=len(cols) - 1, value="Tổng tiền:")
            tc.font = XFont(name="Calibri", bold=True, size=11)
            tc.alignment = Alignment(horizontal="right")
            tv = ws.cell(row=total_row, column=len(cols), value=total)
            tv.font = XFont(name="Calibri", bold=True, size=11)
            tv.number_format = '#,##0'

        # ── column widths ──────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        col_letters = [chr(ord("F") + k) for k in range(len(cols) - 5)]
        for ltr in col_letters:
            ws.column_dimensions[ltr].width = 16

        wb.save(path)
        QMessageBox.information(self, "Xuất Excel", f"Đã lưu file:\n{path}")

    def _export_word(self):
        from ui.word_export import export_nhap_kho
        export_nhap_kho(self, self._receipt, self._lines)


# ── Main Page ─────────────────────────────────────────────────────────────────
class NhapKhoPage(QWidget):

    def __init__(self):
        super().__init__()
        self.setStyleSheet("NhapKhoPage { background: #fafafa; }")
        self._cache: list[Receipt] = []
        self._expanded_data_row: int | None = None
        self._active_subtype = "new"
        self._receipts_order: list[Receipt] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(0)

        # ── Top row: tabs + search + year ─────────────────────────────────
        top = QHBoxLayout()
        top.setSpacing(10)

        # Pill tab bar
        tab_container = QWidget()
        tab_container.setFixedHeight(36)
        tab_container.setStyleSheet("background: #f0f0f0; border-radius: 9px;")
        tab_h = QHBoxLayout(tab_container)
        tab_h.setContentsMargins(3, 3, 3, 3)
        tab_h.setSpacing(2)
        self._tab_btns: list[QPushButton] = []
        for i, txt in enumerate(_TABS):
            btn = QPushButton(txt)
            btn.setFont(QFont(FONT, 11))
            btn.setFixedHeight(30)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self._tab_btns.append(btn)
            tab_h.addWidget(btn)
            btn.clicked.connect(lambda _, i=i: self._on_tab(i))
        self._apply_tab_style(0)
        top.addWidget(tab_container)
        top.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kiếm phiếu...")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        top.addWidget(self._search)

        self._year_combo = QComboBox()
        self._year_combo.setFixedHeight(34)
        self._year_combo.setFont(QFont(FONT, 12))
        self._year_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white;
                color: #111; min-width: 120px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 28px;
                border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
                background: white;
            }
            QComboBox::down-arrow {
                width: 10px; height: 10px;
                image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #e0e0e0; border-radius: 6px;
                background: white; color: #111; selection-background-color: #f0f0f0;
            }
        """)
        from datetime import date as _date
        cur_year = _date.today().year
        self._year_combo.addItem("Tất cả", None)
        for y in range(cur_year, cur_year - 5, -1):
            self._year_combo.addItem(f"Năm {y}", y)
        self._year_combo.setCurrentIndex(1)
        self._year_combo.currentIndexChanged.connect(self._reload)
        top.addWidget(self._year_combo)

        root.addLayout(top)
        root.addSpacing(16)

        # ── Table card ────────────────────────────────────────────────────
        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }")
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(0, 0, 0, 0)
        card_v.setSpacing(0)

        self._table = QTableWidget(0, len(_COLS))
        self._table.setHorizontalHeaderLabels(_COLS)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setMinimumSectionSize(0)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)

        h = self._table.horizontalHeader()
        for i, (mode, w) in enumerate([
            (QHeaderView.ResizeMode.Fixed, 52),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 88),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 100),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 52),
        ]):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)

        self._table.cellClicked.connect(self._on_cell_clicked)
        card_v.addWidget(self._table)
        root.addWidget(card, 1)
        root.addSpacing(16)

        # ── Bottom ─────────────────────────────────────────────────────────
        bot = QHBoxLayout()
        bot.addStretch()
        self._btn_new = QPushButton("+ Tạo Phiếu")
        self._btn_new.setFixedHeight(38)
        self._btn_new.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        self._btn_new.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_new.setStyleSheet(_BTN_DARK)
        self._btn_new.clicked.connect(self._on_add)
        bot.addWidget(self._btn_new)
        root.addLayout(bot)

        # Ẩn cột Vận Chuyển cho tab 2 và 3 (index 8)
        self._update_col_visibility()
        self.refresh()

    # ── Tab ───────────────────────────────────────────────────────────────────

    def _on_tab(self, idx: int):
        self._active_subtype = _SUBTYPES[idx]
        self._apply_tab_style(idx)
        self._search.clear()
        self._update_col_visibility()
        # Update column header for source warehouse
        label = {
            "from_unit":    "Đơn Vị Giao",
            "unit_return":  "Đơn Vị Trả",
            "event_return": "Nguồn Trả",
        }.get(self._active_subtype, "Đơn Vị Giao")
        self._table.setHorizontalHeaderItem(4, _hdr_item(label))
        self._reload()

    def _apply_tab_style(self, active_idx: int):
        for i, btn in enumerate(self._tab_btns):
            if i == active_idx:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: #111; color: white; font-size: 11px; font-weight: 600; }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: transparent; color: #111; font-size: 11px; }
                    QPushButton:hover { background: #e3e3e3; color: #111; }
                """)

    def _update_col_visibility(self):
        self._table.setColumnHidden(8, self._active_subtype != "new")

    # ── Data ──────────────────────────────────────────────────────────────────

    def refresh(self):
        self._reload()

    def _reload(self):
        year = self._year_combo.currentData()
        self._cache = receipt_get_all(year=year, subtype=self._active_subtype)
        self._apply_filter()

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        data = [
            r for r in self._cache
            if not q or any(q in s for s in [
                r.reference_number.lower(),
                r.to_warehouse_name.lower(),
                (r.from_warehouse_name or "").lower(),
                (r.supplier or "").lower(),
                (r.created_by or "").lower(),
                (r.transporter or "").lower(),
                (r.notes or "").lower(),
                r.transaction_date.lower(),
            ])
        ]
        self._load_table(data)

    def _load_table(self, receipts: list[Receipt]):
        self._expanded_data_row = None
        self._table.clearSpans()
        self._table.setRowCount(len(receipts) * 2)
        n = len(_COLS)
        for i, r in enumerate(receipts):
            dr = i * 2
            det = i * 2 + 1
            self._table.setRowHeight(dr, 52)
            self._table.setRowHeight(det, 0)
            self._table.setSpan(det, 0, 1, n)

            source = r.from_warehouse_name or r.supplier
            cells = [
                str(i + 1), r.reference_number, str(r.line_count),
                r.to_warehouse_name, source,
                r.created_by, r.transaction_date,
                r.notes, r.transporter,
            ]
            for c, val in enumerate(cells):
                item = QTableWidgetItem(val)
                item.setFont(QFont(FONT, 12))
                if c == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._table.setItem(dr, c, item)

            dots = QPushButton("···")
            dots.setFlat(True)
            dots.setFont(QFont(FONT, 14))
            dots.setFixedSize(40, 40)
            dots.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            dots.setStyleSheet("""
                QPushButton { color: #111; border: none; background: transparent;
                    letter-spacing: 2px; border-radius: 6px; }
                QPushButton:hover { background: #f0f0f0; color: #555; }
            """)
            rx = r
            dots.clicked.connect(lambda _, rx=rx: self._show_menu(rx))
            self._table.setCellWidget(dr, n - 1, dots)

        self._receipts_order = receipts

    # ── Interaction ───────────────────────────────────────────────────────────

    def _on_cell_clicked(self, row: int, col: int):
        if row % 2 == 1:
            return
        data_idx = row // 2
        det_row = row + 1

        if self._expanded_data_row == row:
            self._table.setRowHeight(det_row, 0)
            self._table.removeCellWidget(det_row, 0)
            self._expanded_data_row = None
            return

        if self._expanded_data_row is not None:
            prev = self._expanded_data_row + 1
            self._table.setRowHeight(prev, 0)
            self._table.removeCellWidget(prev, 0)

        receipt = self._receipts_order[data_idx]
        lines = get_lines(receipt.id)
        panel = DetailPanel(receipt, lines)
        h = 200 + max(1, len(lines)) * 38 + 40
        self._table.setCellWidget(det_row, 0, panel)
        self._table.setRowHeight(det_row, h)
        self._expanded_data_row = row

    def _show_menu(self, receipt: Receipt):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #e5e5e5;
                border-radius: 8px; padding: 4px; }
            QMenu::item { padding: 8px 20px; font-size: 13px;
                border-radius: 6px; color: #111; }
            QMenu::item:selected { background: #f5f5f5; }
            QMenu::separator { height: 1px; background: #efefef; margin: 2px 8px; }
        """)
        act_edit = QAction("Sửa", self)
        act_edit.triggered.connect(lambda: self._edit(receipt))
        act_del = QAction("Xóa", self)
        act_del.triggered.connect(lambda: self._delete(receipt))
        menu.addAction(act_edit)
        menu.addSeparator()
        menu.addAction(act_del)
        menu.exec(self._table.cursor().pos())

    def _on_add(self):
        from ui.dialogs.nhap_kho_form import NhapKhoFormDialog
        dlg = NhapKhoFormDialog(self, subtype=self._active_subtype)
        if dlg.exec():
            self.refresh()

    def _edit(self, receipt: Receipt):
        from ui.dialogs.nhap_kho_form import NhapKhoFormDialog
        dlg = NhapKhoFormDialog(self, receipt=receipt)
        if dlg.exec():
            self.refresh()

    def _delete(self, receipt: Receipt):
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Xóa Phiếu <b>{receipt.reference_number}</b>?<br>"
            "Tồn kho liên quan sẽ được hoàn lại.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            receipt_delete(receipt.id)
            self.refresh()


def _hdr_item(text: str) -> QTableWidgetItem:
    item = QTableWidgetItem(text)
    item.setFont(QFont(FONT, 12))
    return item
