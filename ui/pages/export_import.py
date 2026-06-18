"""Trang Export / Import dữ liệu."""

from __future__ import annotations
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QFileDialog, QMessageBox, QSizePolicy, QInputDialog,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
import database

FONT = "Segoe UI"

CARD_STYLE = """
    QFrame {
        background: white;
        border: 1px solid #e8e8e8;
        border-radius: 10px;
    }
"""
BTN_PRIMARY = """
    QPushButton {
        background: #111;
        color: white;
        border-radius: 6px;
        padding: 6px 16px;
        border: none;
    }
    QPushButton:hover { background: #333; }
"""
BTN_SECONDARY = """
    QPushButton {
        background: white;
        border: 1px solid #ccc;
        border-radius: 6px;
        padding: 6px 14px;
        color: #444;
    }
    QPushButton:hover { background: #f5f5f5; }
"""
BTN_GRAY = """
    QPushButton {
        background: #666;
        color: white;
        border-radius: 6px;
        padding: 6px 14px;
        border: none;
    }
    QPushButton:hover { background: #555; }
"""


class _Card(QFrame):
    def __init__(self, title: str, description: str,
                 primary_label: str, primary_cb,
                 secondary_label: str | None = None, secondary_cb=None,
                 tertiary_label: str | None = None, tertiary_cb=None):
        super().__init__()
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(CARD_STYLE)

        v = QVBoxLayout(self)
        v.setContentsMargins(20, 18, 20, 18)
        v.setSpacing(8)

        t = QLabel(title)
        t.setFont(QFont(FONT, 13, QFont.Weight.Bold))
        t.setStyleSheet("color: #111; border: none;")
        v.addWidget(t)

        d = QLabel(description)
        d.setFont(QFont(FONT, 11))
        d.setStyleSheet("color: #666; border: none;")
        d.setWordWrap(True)
        d.setMinimumWidth(1)
        d.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)
        v.addWidget(d)

        row = QHBoxLayout()
        row.setSpacing(8)
        row.addStretch()

        if tertiary_label and tertiary_cb:
            tb = QPushButton(tertiary_label)
            tb.setFont(QFont(FONT, 11))
            tb.setStyleSheet(BTN_GRAY)
            tb.clicked.connect(tertiary_cb)
            row.addWidget(tb)

        if secondary_label and secondary_cb:
            sb = QPushButton(secondary_label)
            sb.setFont(QFont(FONT, 11))
            sb.setStyleSheet(BTN_SECONDARY)
            sb.clicked.connect(secondary_cb)
            row.addWidget(sb)

        pb = QPushButton(primary_label)
        pb.setFont(QFont(FONT, 11))
        pb.setStyleSheet(BTN_PRIMARY)
        pb.clicked.connect(primary_cb)
        row.addWidget(pb)

        v.addLayout(row)


class ExportImportPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(32, 28, 32, 28)
        v.setSpacing(20)

        title = QLabel("Export / Import Dữ Liệu")
        title.setFont(QFont(FONT, 20, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        v.addWidget(title)

        # ── Xuất dữ liệu ──────────────────────────────────────────────────────
        v.addWidget(_section_header("XUẤT DỮ LIỆU"))

        export_row2 = QHBoxLayout()
        export_row2.setSpacing(16)
        export_row2.addWidget(_Card(
            "Tổng Hàng Tại Kho",
            "Xuất tổng hợp tồn kho (H1-H4, giá trị) tại các kho tổng, nhóm theo mặt hàng.",
            "Xuất Excel", self._export_inv_kho,
        ))
        export_row2.addWidget(_Card(
            "Tổng Hàng Tại Đơn Vị",
            "Xuất tổng hợp tồn kho (H1-H4, thâm niên) tại tất cả đơn vị, nhóm theo mặt hàng.",
            "Xuất Excel", self._export_inv_donvi,
        ))
        export_row2.addWidget(_Card(
            "Hàng Dùng Chung",
            "Xuất danh sách hàng dùng chung hiện tồn tại kho (H1-H4, đơn giá, giá trị).",
            "Xuất Excel", self._export_shared,
        ))
        v.addLayout(export_row2)

        export_row3 = QHBoxLayout()
        export_row3.setSpacing(16)
        export_row3.addWidget(_Card(
            "H4 Tại Kho",
            "Xuất danh sách hàng H4 đang tồn tại các kho tổng (không kèm giá).",
            "Xuất Excel", self._export_h4_kho,
        ))
        export_row3.addWidget(_Card(
            "H4 Tại Đơn Vị (Chọn)",
            "Chọn một đơn vị, xuất danh sách hàng H4 đang tồn tại đó.",
            "Xuất Excel", self._export_h4_donvi_sel,
        ))
        export_row3.addWidget(_Card(
            "Tổng H4 Tất Cả Đơn Vị",
            "Xuất tổng hợp hàng H4 tại tất cả đơn vị (kèm thâm niên, không kèm giá).",
            "Xuất Excel", self._export_h4_all_donvi,
        ))
        v.addLayout(export_row3)

        export_row4 = QHBoxLayout()
        export_row4.setSpacing(16)
        export_row4.addWidget(_Card(
            "Báo Cáo Năm – Kho Được Chọn",
            "Chọn kho và năm, xuất báo cáo tồn kho theo năm (đầu năm, nhập, xuất, cuối năm).",
            "Xuất Excel", self._export_annual_kho_sel,
        ))
        export_row4.addWidget(_Card(
            "Báo Cáo Năm – ĐV Được Chọn",
            "Chọn đơn vị và năm, xuất báo cáo tồn kho theo năm tại đơn vị đó.",
            "Xuất Excel", self._export_annual_dv_sel,
        ))
        export_row4.addStretch()
        v.addLayout(export_row4)

        export_row5 = QHBoxLayout()
        export_row5.setSpacing(16)
        export_row5.addWidget(_Card(
            "Báo Cáo Năm – Tất Cả Kho",
            "Xuất báo cáo năm của toàn bộ kho tổng (mọi năm đã lưu snapshot).",
            "Xuất Excel", self._export_annual_all_kho,
        ))
        export_row5.addWidget(_Card(
            "Báo Cáo Năm – Tất Cả Đơn Vị",
            "Xuất báo cáo năm của toàn bộ đơn vị (mọi năm đã lưu snapshot).",
            "Xuất Excel", self._export_annual_all_dv,
        ))
        export_row5.addStretch()
        v.addLayout(export_row5)

        # ── Nhập dữ liệu ──────────────────────────────────────────────────────
        v.addWidget(_section_header("NHẬP DỮ LIỆU"))

        import_row = QHBoxLayout()
        import_row.setSpacing(16)
        import_row.addWidget(_Card(
            "Danh Mục Mặt Hàng",
            "Nhập danh sách mặt hàng từ file Excel. Mã đã tồn tại sẽ được cập nhật.",
            "Nhập từ Excel", self._import_item_types,
            secondary_label="Tải Mẫu", secondary_cb=self._template_item_types,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_item_types,
        ))
        import_row.addWidget(_Card(
            "Danh Sách Kho / Đơn Vị",
            "Nhập danh sách kho/đơn vị từ file Excel. Mã đã tồn tại sẽ được cập nhật.",
            "Nhập từ Excel", self._import_warehouses,
            secondary_label="Tải Mẫu", secondary_cb=self._template_warehouses,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_warehouses,
        ))
        import_row.addStretch()
        v.addLayout(import_row)

        import_row2 = QHBoxLayout()
        import_row2.setSpacing(16)
        import_row2.addWidget(_Card(
            "Phiếu Nhập Kho",
            "Nhập hàng loạt phiếu nhập kho từ Excel. Mỗi dòng là một mặt hàng trong phiếu.",
            "Nhập từ Excel", self._import_receipts,
            secondary_label="Tải Mẫu", secondary_cb=self._template_receipts,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_receipts,
        ))
        import_row2.addWidget(_Card(
            "Phiếu Xuất Kho",
            "Nhập hàng loạt phiếu xuất kho từ Excel. Mỗi dòng là một mặt hàng trong phiếu.",
            "Nhập từ Excel", self._import_issues,
            secondary_label="Tải Mẫu", secondary_cb=self._template_issues,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_issues,
        ))
        import_row2.addStretch()
        v.addLayout(import_row2)

        import_row3 = QHBoxLayout()
        import_row3.setSpacing(16)
        import_row3.addWidget(_Card(
            "Phiếu Luân Chuyển",
            "Nhập hàng loạt phiếu luân chuyển từ Excel. Mỗi dòng là một mặt hàng trong phiếu.",
            "Nhập từ Excel", self._import_transfers,
            secondary_label="Tải Mẫu", secondary_cb=self._template_transfers,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_transfers,
        ))
        import_row3.addWidget(_Card(
            "Tồn Kho / Đơn Vị",
            "Nhập số lượng tồn kho ban đầu từ Excel cho từng kho/đơn vị, mặt hàng, mức HH.",
            "Nhập từ Excel", self._import_inventory,
            secondary_label="Tải Mẫu", secondary_cb=self._template_inventory,
            tertiary_label="Xuất Excel", tertiary_cb=self._export_inventory,
        ))
        import_row3.addStretch()
        v.addLayout(import_row3)

        v.addStretch()

    # ── Export ────────────────────────────────────────────────────────────────

    def _export_inventory(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Tồn Kho", "ton_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name, w.type AS wh_type,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       i.quality_level, i.quantity,
                       i.received_at_unit_date, i.manufacture_date, i.lot_number,
                       i.is_shared, i.notes
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE i.quantity > 0
                ORDER BY w.type, w.name, t.name, i.quality_level
            """).fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tồn Kho"

            headers = [
                "Mã Kho", "Tên Kho", "Loại Kho",
                "Mã Hàng", "Tên Hàng", "ĐVT",
                "Mức HH", "Số Lượng",
                "Ngày Nhập ĐV", "Ngày SX", "Số Lô",
                "Dùng Chung", "Ghi Chú",
            ]
            _write_header(ws, headers)

            for r in rows:
                ws.append([
                    r["wh_code"], r["wh_name"],
                    "Kho Tổng" if r["wh_type"] == "TONG" else "Đơn Vị",
                    r["item_code"], r["item_name"], r["unit_of_measure"],
                    r["quality_level"], r["quantity"],
                    r["received_at_unit_date"] or "",
                    r["manufacture_date"] or "",
                    r["lot_number"] or "",
                    "Có" if r["is_shared"] else "",
                    r["notes"] or "",
                ])

            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_item_types(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Danh Mục Mặt Hàng", "danh_muc_hang.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl

            conn = database.get_conn()
            rows = conn.execute(
                "SELECT code, name, unit_of_measure, total_lifespan_months,"
                "       unit_price, notes, is_active"
                " FROM item_types ORDER BY name"
            ).fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mặt Hàng"

            _write_header(ws, [
                "Mã Hàng", "Tên Hàng", "Đơn Vị Tính",
                "Niên Hạn (Năm)", "Đơn Giá", "Ghi Chú", "Trạng Thái",
            ])
            for r in rows:
                ws.append([
                    r["code"], r["name"], r["unit_of_measure"],
                    round((r["total_lifespan_months"] or 0) / 12.0, 1),
                    float(r["unit_price"] or 0),
                    r["notes"] or "",
                    "Hoạt động" if r["is_active"] else "Đã ẩn",
                ])

            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} mặt hàng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_warehouses(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Danh Sách Kho", "danh_sach_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl

            conn = database.get_conn()
            rows = conn.execute(
                "SELECT code, name, type, address, notes, is_active"
                " FROM warehouses ORDER BY type, name"
            ).fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kho"

            _write_header(ws, [
                "Mã Kho", "Tên Kho", "Loại (TONG/DON_VI)",
                "Địa Chỉ", "Ghi Chú", "Trạng Thái",
            ])
            for r in rows:
                ws.append([
                    r["code"], r["name"], r["type"],
                    r["address"] or "", r["notes"] or "",
                    "Hoạt động" if r["is_active"] else "Đã ẩn",
                ])

            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} kho/đơn vị.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Xuất tồn kho tổng hợp ─────────────────────────────────────────────────

    def _export_inv_kho(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Tổng Hàng Tại Kho", "tong_hang_tai_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       ROUND(t.total_lifespan_months / 12.0, 1) AS lifespan_years,
                       t.unit_price,
                       SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                       SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                       SUM(i.quantity) AS total
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE w.type = 'TONG' AND i.is_shared = 0 AND i.quantity > 0
                GROUP BY w.id, t.id
                ORDER BY w.name, t.name
            """).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng Hàng Tại Kho"
            _write_header(ws, [
                "Mã Kho", "Tên Kho", "Mã Hàng", "Tên Hàng", "ĐVT",
                "Niên Hạn (Năm)", "Đơn Giá", "H1", "H2", "H3", "H4", "Tổng", "Giá Trị H1",
            ])
            for r in rows:
                price = float(r["unit_price"] or 0)
                ws.append([
                    r["wh_code"], r["wh_name"], r["item_code"], r["item_name"],
                    r["unit_of_measure"], r["lifespan_years"], price,
                    r["h1"], r["h2"], r["h3"], r["h4"], r["total"],
                    r["h1"] * price,
                ])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_inv_donvi(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Tổng Hàng Tại Đơn Vị", "tong_hang_tai_don_vi.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       t.total_lifespan_months,
                       MAX(CAST((julianday('now') - julianday(i.received_at_unit_date)) / 30.44 AS INTEGER)) AS months_at_unit,
                       SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                       SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                       SUM(i.quantity) AS total
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE w.type = 'DON_VI' AND i.quantity > 0
                GROUP BY w.id, t.id
                ORDER BY w.name, t.name
            """).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng Hàng Tại ĐV"
            _write_header(ws, [
                "Mã ĐV", "Tên ĐV", "Mã Hàng", "Tên Hàng", "ĐVT",
                "Tại ĐV (tháng)", "Còn Lại (tháng)", "H1", "H2", "H3", "H4", "Tổng",
            ])
            for r in rows:
                mm = r["months_at_unit"] or 0
                remaining = max(0, (r["total_lifespan_months"] or 0) - mm)
                ws.append([
                    r["wh_code"], r["wh_name"], r["item_code"], r["item_name"],
                    r["unit_of_measure"], mm, remaining,
                    r["h1"], r["h2"], r["h3"], r["h4"], r["total"],
                ])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_shared(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Hàng Dùng Chung", "hang_dung_chung.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure, t.unit_price,
                       SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                       SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                       SUM(i.quantity) AS total
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE i.is_shared = 1 AND i.quantity > 0
                GROUP BY w.id, t.id
                ORDER BY w.name, t.name
            """).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hàng Dùng Chung"
            _write_header(ws, [
                "Mã Kho", "Tên Kho", "Mã Hàng", "Tên Hàng", "ĐVT", "Đơn Giá",
                "H1", "H2", "H3", "H4", "Tổng", "Giá Trị H1",
            ])
            for r in rows:
                price = float(r["unit_price"] or 0)
                ws.append([
                    r["wh_code"], r["wh_name"], r["item_code"], r["item_name"],
                    r["unit_of_measure"], price,
                    r["h1"], r["h2"], r["h3"], r["h4"], r["total"],
                    r["h1"] * price,
                ])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Xuất H4 ───────────────────────────────────────────────────────────────

    def _export_h4_kho(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất H4 Tại Kho", "h4_tai_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       SUM(i.quantity) AS quantity
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE w.type = 'TONG' AND i.quality_level = 'H4'
                  AND i.is_shared = 0 AND i.quantity > 0
                GROUP BY w.id, t.id
                ORDER BY w.name, t.name
            """).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "H4 Tại Kho"
            _write_header(ws, ["Mã Kho", "Tên Kho", "Mã Hàng", "Tên Hàng", "ĐVT", "Số Lượng"])
            for r in rows:
                ws.append([r["wh_code"], r["wh_name"], r["item_code"], r["item_name"],
                           r["unit_of_measure"], r["quantity"]])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_h4_donvi_sel(self):
        if not _check_openpyxl(self):
            return
        conn = database.get_conn()
        units = conn.execute(
            "SELECT id, code, name FROM warehouses WHERE type='DON_VI' AND is_active=1 ORDER BY name"
        ).fetchall()
        if not units:
            QMessageBox.information(self, "Thông báo", "Không có đơn vị nào.")
            return
        items = [f"{r['code']} – {r['name']}" for r in units]
        chosen, ok = QInputDialog.getItem(self, "Chọn Đơn Vị", "Đơn vị:", items, 0, False)
        if not ok:
            return
        wh_id = units[items.index(chosen)]["id"]
        wh_label = chosen.replace(" – ", "_").replace(" ", "")
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất H4 Tại Đơn Vị", f"h4_{wh_label}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            rows = conn.execute("""
                SELECT t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       MAX(CAST((julianday('now') - julianday(i.received_at_unit_date)) / 30.44 AS INTEGER)) AS months_at_unit,
                       SUM(i.quantity) AS quantity
                FROM inventory i
                JOIN item_types t ON t.id = i.item_type_id
                WHERE i.warehouse_id = ? AND i.quality_level = 'H4' AND i.quantity > 0
                GROUP BY t.id
                ORDER BY t.name
            """, (wh_id,)).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "H4"
            _write_header(ws, ["Mã Hàng", "Tên Hàng", "ĐVT", "Tại ĐV (tháng)", "Số Lượng"])
            for r in rows:
                ws.append([r["item_code"], r["item_name"], r["unit_of_measure"],
                           r["months_at_unit"] or 0, r["quantity"]])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_h4_all_donvi(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Tổng H4 Tất Cả ĐV", "h4_tat_ca_don_vi.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       MAX(CAST((julianday('now') - julianday(i.received_at_unit_date)) / 30.44 AS INTEGER)) AS months_at_unit,
                       SUM(i.quantity) AS quantity
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE w.type = 'DON_VI' AND i.quality_level = 'H4' AND i.quantity > 0
                GROUP BY w.id, t.id
                ORDER BY w.name, t.name
            """).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng H4 Tất Cả ĐV"
            _write_header(ws, ["Mã ĐV", "Tên ĐV", "Mã Hàng", "Tên Hàng", "ĐVT",
                               "Tại ĐV (tháng)", "Số Lượng"])
            for r in rows:
                ws.append([r["wh_code"], r["wh_name"], r["item_code"], r["item_name"],
                           r["unit_of_measure"], r["months_at_unit"] or 0, r["quantity"]])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Xuất Báo Cáo Năm ──────────────────────────────────────────────────────

    def _pick_wh_and_year(self, wh_type: str):
        """Hiện 2 dialog liên tiếp để chọn kho/ĐV và năm. Trả về (wh_id, year) hoặc None."""
        conn = database.get_conn()
        label = "Đơn Vị" if wh_type == "DON_VI" else "Kho"
        units = conn.execute(
            "SELECT id, code, name FROM warehouses WHERE type=? AND is_active=1 ORDER BY name",
            (wh_type,)
        ).fetchall()
        if not units:
            QMessageBox.information(self, "Thông báo", f"Không có {label} nào.")
            return None
        items = [f"{r['code']} – {r['name']}" for r in units]
        chosen, ok = QInputDialog.getItem(self, f"Chọn {label}", f"{label}:", items, 0, False)
        if not ok:
            return None
        wh_id = units[items.index(chosen)]["id"]

        years_raw = conn.execute(
            "SELECT DISTINCT year FROM annual_snapshots WHERE warehouse_id=? ORDER BY year DESC",
            (wh_id,)
        ).fetchall()
        if not years_raw:
            QMessageBox.information(self, "Thông báo", "Chưa có dữ liệu báo cáo năm cho kho/ĐV này.")
            return None
        year_items = [str(r["year"]) for r in years_raw]
        yr_str, ok2 = QInputDialog.getItem(self, "Chọn Năm", "Năm báo cáo:", year_items, 0, False)
        if not ok2:
            return None
        return wh_id, int(yr_str), chosen

    def _write_annual_rows(self, ws, rows, include_price: bool):
        headers = ["Mã Hàng", "Tên Hàng", "ĐVT", "Niên Hạn (Năm)"]
        if include_price:
            headers.append("Đơn Giá")
        headers += ["H1 Cuối Năm", "H2 Cuối Năm", "H3 Cuối Năm", "H4 Cuối Năm",
                    "Tổng Cuối Năm", "Nhập Năm", "Xuất Năm", "Nâng Mức", "TXL"]
        _write_header(ws, headers)
        for r in rows:
            total = r["h1_qty"] + r["h2_qty"] + r["h3_qty"] + r["h4_qty"]
            row_data = [
                r["item_code"], r["item_name"], r["unit_of_measure"],
                round(r["total_lifespan_months"] / 12.0, 1),
            ]
            if include_price:
                row_data.append(float(r["unit_price"] or 0))
            row_data += [
                r["h1_qty"], r["h2_qty"], r["h3_qty"], r["h4_qty"], total,
                r["imported_qty"], r["exported_qty"], r["upgraded_qty"], r["disposed_qty"],
            ]
            ws.append(row_data)

    def _export_annual_kho_sel(self):
        if not _check_openpyxl(self):
            return
        result = self._pick_wh_and_year("TONG")
        if not result:
            return
        wh_id, year, label = result
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Báo Cáo Năm Kho", f"bc_nam_kho_{year}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       t.total_lifespan_months, t.unit_price,
                       s.h1_qty, s.h2_qty, s.h3_qty, s.h4_qty,
                       s.imported_qty, s.exported_qty, s.upgraded_qty, s.disposed_qty
                FROM annual_snapshots s
                JOIN item_types t ON t.id = s.item_type_id
                WHERE s.warehouse_id = ? AND s.year = ?
                ORDER BY t.name
            """, (wh_id, year)).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"BC Năm {year}"
            ws.append([f"Báo cáo năm {year} – {label}"])
            ws.append([])
            self._write_annual_rows(ws, rows, include_price=True)
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} mặt hàng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_annual_dv_sel(self):
        if not _check_openpyxl(self):
            return
        result = self._pick_wh_and_year("DON_VI")
        if not result:
            return
        wh_id, year, label = result
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Báo Cáo Năm ĐV", f"bc_nam_dv_{year}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       t.total_lifespan_months, t.unit_price,
                       s.h1_qty, s.h2_qty, s.h3_qty, s.h4_qty,
                       s.imported_qty, s.exported_qty, s.upgraded_qty, s.disposed_qty
                FROM annual_snapshots s
                JOIN item_types t ON t.id = s.item_type_id
                WHERE s.warehouse_id = ? AND s.year = ?
                ORDER BY t.name
            """, (wh_id, year)).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"BC Năm {year}"
            ws.append([f"Báo cáo năm {year} – {label}"])
            ws.append([])
            self._write_annual_rows(ws, rows, include_price=False)
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} mặt hàng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_annual_all(self, wh_type: str, dialog_title: str,
                           default_name: str, sheet_name: str, include_price: bool):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(self, dialog_title, default_name, "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            rows = conn.execute("""
                SELECT s.year, w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name, t.unit_of_measure,
                       t.total_lifespan_months, t.unit_price,
                       s.h1_qty, s.h2_qty, s.h3_qty, s.h4_qty,
                       s.imported_qty, s.exported_qty, s.upgraded_qty, s.disposed_qty
                FROM annual_snapshots s
                JOIN warehouses w ON w.id = s.warehouse_id
                JOIN item_types t  ON t.id = s.item_type_id
                WHERE w.type = ?
                ORDER BY s.year DESC, w.name, t.name
            """, (wh_type,)).fetchall()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            headers = ["Năm", "Mã Kho/ĐV", "Tên Kho/ĐV", "Mã Hàng", "Tên Hàng", "ĐVT",
                       "Niên Hạn (Năm)"]
            if include_price:
                headers.append("Đơn Giá")
            headers += ["H1 Cuối Năm", "H2 Cuối Năm", "H3 Cuối Năm", "H4 Cuối Năm",
                        "Tổng Cuối Năm", "Nhập Năm", "Xuất Năm", "Nâng Mức", "TXL"]
            _write_header(ws, headers)
            for r in rows:
                total = r["h1_qty"] + r["h2_qty"] + r["h3_qty"] + r["h4_qty"]
                row_data = [
                    r["year"], r["wh_code"], r["wh_name"],
                    r["item_code"], r["item_name"], r["unit_of_measure"],
                    round(r["total_lifespan_months"] / 12.0, 1),
                ]
                if include_price:
                    row_data.append(float(r["unit_price"] or 0))
                row_data += [
                    r["h1_qty"], r["h2_qty"], r["h3_qty"], r["h4_qty"], total,
                    r["imported_qty"], r["exported_qty"], r["upgraded_qty"], r["disposed_qty"],
                ]
                ws.append(row_data)
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_annual_all_kho(self):
        self._export_annual_all(
            "TONG", "Xuất BC Năm Tất Cả Kho", "bc_nam_tat_ca_kho.xlsx",
            "BC Năm Kho", include_price=True,
        )

    def _export_annual_all_dv(self):
        self._export_annual_all(
            "DON_VI", "Xuất BC Năm Tất Cả ĐV", "bc_nam_tat_ca_don_vi.xlsx",
            "BC Năm ĐV", include_price=False,
        )

    def _export_receipts(self):
        self._export_transactions("NHAP_KHO", "Xuất Phiếu Nhập Kho", "phieu_nhap_kho.xlsx", "Nhập Kho")

    def _export_issues(self):
        self._export_transactions("XUAT_KHO", "Xuất Phiếu Xuất Kho", "phieu_xuat_kho.xlsx", "Xuất Kho")

    def _export_transfers(self):
        self._export_transactions(
            ("LUAN_CHUYEN_KHO", "LUAN_CHUYEN_DV"),
            "Xuất Phiếu Luân Chuyển", "phieu_luan_chuyen.xlsx", "Luân Chuyển",
        )

    def _export_transactions(self, tx_type, dialog_title: str, default_name: str, sheet_name: str):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(self, dialog_title, default_name, "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl

            conn = database.get_conn()
            if isinstance(tx_type, tuple):
                placeholders = ",".join("?" * len(tx_type))
                sql = f"""
                    SELECT tx.reference_number, tx.transaction_date, tx.created_by, tx.notes AS tx_notes,
                           wf.code AS from_code, wf.name AS from_name,
                           wt.code AS to_code,   wt.name AS to_name,
                           t.code AS item_code,  t.name AS item_name, t.unit_of_measure,
                           tl.quality_level_from, tl.quality_level_to, tl.quantity,
                           tl.lot_number, tl.notes AS line_notes
                    FROM transactions tx
                    JOIN transaction_lines tl ON tl.transaction_id = tx.id
                    JOIN item_types t  ON t.id  = tl.item_type_id
                    LEFT JOIN warehouses wf ON wf.id = tx.from_warehouse_id
                    LEFT JOIN warehouses wt ON wt.id = tx.to_warehouse_id
                    WHERE tx.type IN ({placeholders})
                    ORDER BY tx.transaction_date DESC, tx.id, tl.id
                """
                rows = conn.execute(sql, tx_type).fetchall()
            else:
                rows = conn.execute("""
                    SELECT tx.reference_number, tx.transaction_date, tx.created_by, tx.notes AS tx_notes,
                           wf.code AS from_code, wf.name AS from_name,
                           wt.code AS to_code,   wt.name AS to_name,
                           t.code AS item_code,  t.name AS item_name, t.unit_of_measure,
                           tl.quality_level_from, tl.quality_level_to, tl.quantity,
                           tl.lot_number, tl.notes AS line_notes
                    FROM transactions tx
                    JOIN transaction_lines tl ON tl.transaction_id = tx.id
                    JOIN item_types t  ON t.id  = tl.item_type_id
                    LEFT JOIN warehouses wf ON wf.id = tx.from_warehouse_id
                    LEFT JOIN warehouses wt ON wt.id = tx.to_warehouse_id
                    WHERE tx.type = ?
                    ORDER BY tx.transaction_date DESC, tx.id, tl.id
                """, (tx_type,)).fetchall()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            _write_header(ws, [
                "Số Phiếu", "Ngày", "Người Lập", "Ghi Chú Phiếu",
                "Mã Kho Từ", "Tên Kho Từ", "Mã Kho Đến", "Tên Kho Đến",
                "Mã Hàng", "Tên Hàng", "ĐVT",
                "Mức HH Từ", "Mức HH Đến", "Số Lượng",
                "Số Lô", "Ghi Chú Dòng",
            ])
            for r in rows:
                ws.append([
                    r["reference_number"] or "",
                    r["transaction_date"] or "",
                    r["created_by"] or "",
                    r["tx_notes"] or "",
                    r["from_code"] or "", r["from_name"] or "",
                    r["to_code"] or "",   r["to_name"] or "",
                    r["item_code"], r["item_name"], r["unit_of_measure"],
                    r["quality_level_from"] or "",
                    r["quality_level_to"],
                    r["quantity"],
                    r["lot_number"] or "",
                    r["line_notes"] or "",
                ])

            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Import ────────────────────────────────────────────────────────────────

    def _import_item_types(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn File Excel", "", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))

            if not all_rows:
                QMessageBox.warning(self, "Lỗi", "File Excel trống.")
                return

            # Detect format from header row:
            # Export format: [Mã Hàng, Tên Hàng, ĐVT, Niên Hạn, Đơn Giá, Ghi Chú, Trạng Thái]
            # Template format: [Tên Hàng *, ĐVT *, Niên Hạn *, Đơn Giá, Ghi Chú]
            header0 = str(all_rows[0][0]).strip().lower() if all_rows[0] and all_rows[0][0] else ""
            has_code_col = "mã hàng" in header0 or "ma hang" in header0

            rows = all_rows[1:]  # skip header

            conn = database.get_conn()
            inserted = updated = skipped = 0
            errors = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                try:
                    if has_code_col:
                        # Export format: [code, name, uom, lifespan, price, notes, status]
                        code  = str(row[0]).strip() if row[0] else ""
                        name  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        uom   = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        life  = int(round(float(row[3]) * 12)) if len(row) > 3 and row[3] is not None else 0
                        price = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0
                        notes = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                        status = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                        is_active = 0 if status == "Đã ẩn" else 1
                    else:
                        # Template format: [name, uom, lifespan, price, notes]
                        code  = ""
                        name  = str(row[0]).strip() if row[0] else ""
                        uom   = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        life  = int(round(float(row[2]) * 12)) if len(row) > 2 and row[2] is not None else 0
                        price = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                        notes = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                        is_active = 1

                    if not name or not uom or not life:
                        errors.append(f"Dòng {i}: thiếu dữ liệu bắt buộc")
                        skipped += 1
                        continue

                    # Lookup by code (reliable) if available, else by name
                    if code:
                        existing = conn.execute(
                            "SELECT id FROM item_types WHERE code=?", (code,)
                        ).fetchone()
                    else:
                        existing = conn.execute(
                            "SELECT id FROM item_types WHERE name=?", (name,)
                        ).fetchone()

                    if existing:
                        conn.execute(
                            "UPDATE item_types SET name=?, unit_of_measure=?,"
                            " total_lifespan_months=?, unit_price=?, notes=?, is_active=?"
                            " WHERE id=?",
                            (name, uom, life, price, notes, is_active, existing["id"]),
                        )
                        updated += 1
                    else:
                        if not code:
                            n = 1
                            while True:
                                code = f"HH{n:04d}"
                                if not conn.execute("SELECT 1 FROM item_types WHERE code=?", (code,)).fetchone():
                                    break
                                n += 1
                        conn.execute(
                            "INSERT INTO item_types (code, name, unit_of_measure,"
                            " total_lifespan_months, unit_price, notes)"
                            " VALUES (?,?,?,?,?,?)",
                            (code, name, uom, life, price, notes),
                        )
                        inserted += 1
                except Exception as row_err:
                    errors.append(f"Dòng {i}: {row_err}")
                    skipped += 1

            conn.commit()

            msg = f"Thêm mới: {inserted}   Cập nhật: {updated}   Bỏ qua: {skipped}"
            if errors:
                msg += "\n\nChi tiết lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_warehouses(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn File Excel", "", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))

            if not all_rows:
                QMessageBox.warning(self, "Lỗi", "File Excel trống.")
                return

            # Detect format from header row:
            # Export format: [Mã Kho, Tên Kho, Loại, Địa Chỉ, Ghi Chú, Trạng Thái]
            # Template format: [Tên Kho *, Loại *, Địa Chỉ, Ghi Chú]
            header0 = str(all_rows[0][0]).strip().lower() if all_rows[0] and all_rows[0][0] else ""
            has_code_col = "mã kho" in header0 or "ma kho" in header0

            rows = all_rows[1:]  # skip header

            conn = database.get_conn()
            inserted = updated = skipped = 0
            errors = []
            VALID_TYPES = {"TONG", "DON_VI"}

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                try:
                    if has_code_col:
                        # Export format: [code, name, type, address, notes, status]
                        code    = str(row[0]).strip() if row[0] else ""
                        name    = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        wh_type = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ""
                        address = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        notes   = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                        status  = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                        is_active = 0 if status == "Đã ẩn" else 1
                    else:
                        # Template format: [name, type, address, notes]
                        code    = ""
                        name    = str(row[0]).strip() if row[0] else ""
                        wh_type = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ""
                        address = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        notes   = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        is_active = 1

                    if not name or wh_type not in VALID_TYPES:
                        errors.append(
                            f"Dòng {i}: thiếu hoặc sai dữ liệu "
                            f"(Loại phải là TONG hoặc DON_VI)"
                        )
                        skipped += 1
                        continue

                    # Lookup by code (reliable) if available, else by name
                    if code:
                        existing = conn.execute(
                            "SELECT id FROM warehouses WHERE code=?", (code,)
                        ).fetchone()
                    else:
                        existing = conn.execute(
                            "SELECT id FROM warehouses WHERE name=?", (name,)
                        ).fetchone()

                    if existing:
                        conn.execute(
                            "UPDATE warehouses SET name=?, type=?, address=?, notes=?, is_active=?"
                            " WHERE id=?",
                            (name, wh_type, address, notes, is_active, existing["id"]),
                        )
                        updated += 1
                    else:
                        if not code:
                            n = 1
                            while True:
                                code = f"KHO{n:04d}"
                                if not conn.execute("SELECT 1 FROM warehouses WHERE code=?", (code,)).fetchone():
                                    break
                                n += 1
                        conn.execute(
                            "INSERT INTO warehouses (code, name, type, address, notes)"
                            " VALUES (?,?,?,?,?)",
                            (code, name, wh_type, address, notes),
                        )
                        inserted += 1
                except Exception as row_err:
                    errors.append(f"Dòng {i}: {row_err}")
                    skipped += 1

            conn.commit()

            msg = f"Thêm mới: {inserted}   Cập nhật: {updated}   Bỏ qua: {skipped}"
            if errors:
                msg += "\n\nChi tiết lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Download templates ────────────────────────────────────────────────────

    def _template_receipts(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_phieu_nhap_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Phiếu Nhập Kho"
            _write_header(ws, [
                "Số Phiếu *", "Ngày * (yyyy-mm-dd)", "Người Lập", "Ghi Chú",
                "Tên Kho Nhận *", "Tên Hàng *", "Số Lượng *", "Ghi Chú Dòng",
            ])
            ws.append(["NK001", "2024-01-15", "Nguyễn Văn A", "", "Kho Tổng D6", "Súng AK47", 10, ""])
            ws.append(["NK001", "2024-01-15", "Nguyễn Văn A", "", "Kho Tổng D6", "Súng B40",  5,  ""])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _template_issues(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_phieu_xuat_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Phiếu Xuất Kho"
            _write_header(ws, [
                "Số Phiếu *", "Ngày * (yyyy-mm-dd)", "Người Lập", "Ghi Chú",
                "Tên Kho Từ *", "Tên Đơn Vị Nhận", "Tên Hàng *", "Số Lượng *", "Ghi Chú Dòng",
            ])
            ws.append(["XK001", "2024-01-20", "Trần Thị B", "", "Kho Tổng D6", "Đơn Vị 01", "Súng AK47", 3, ""])
            ws.append(["XK001", "2024-01-20", "Trần Thị B", "", "Kho Tổng D6", "Đơn Vị 01", "Súng B40",  2, ""])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _template_transfers(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_phieu_luan_chuyen.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Phiếu Luân Chuyển"
            _write_header(ws, [
                "Số Phiếu *", "Ngày * (yyyy-mm-dd)", "Loại * (TONG/DON_VI)",
                "Người Lập", "Ghi Chú",
                "Tên Kho Từ *", "Tên Kho Đến *",
                "Tên Hàng *", "Mức HH (DON_VI)", "Số Lượng *", "Ghi Chú Dòng",
            ])
            ws.append(["LC001", "2024-02-01", "TONG", "Lê Văn C", "", "Kho Tổng D6", "Kho Tổng D7", "Súng AK47", "", 5, ""])
            ws.append(["LC002", "2024-02-05", "DON_VI", "Lê Văn C", "", "Đơn Vị 01", "Đơn Vị 02", "Súng AK47", "H2", 2, ""])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_receipts(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from collections import OrderedDict
            from database.receipts import Receipt, ReceiptLine, insert as _insert

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            conn = database.get_conn()
            wh_map = {r["name"]: r["id"] for r in conn.execute(
                "SELECT name, id FROM warehouses WHERE is_active=1"
            ).fetchall()}
            item_map = {r["name"]: dict(r) for r in conn.execute(
                "SELECT code, id, name, unit_of_measure FROM item_types WHERE is_active=1"
            ).fetchall()}

            groups: OrderedDict = OrderedDict()
            errors: list[str] = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                so_phieu  = str(row[0]).strip()
                ngay      = str(row[1]).strip() if row[1] else ""
                nguoi_lap = str(row[2]).strip() if row[2] else ""
                ghi_chu   = str(row[3]).strip() if row[3] else ""
                ten_kho   = str(row[4]).strip() if row[4] else ""
                ten_hang  = str(row[5]).strip() if row[5] else ""
                so_luong  = int(row[6]) if row[6] else 0
                ghi_chu_dong = str(row[7]).strip() if len(row) > 7 and row[7] else ""

                if not so_phieu or not ngay or not ten_kho or not ten_hang or not so_luong:
                    errors.append(f"Dòng {i}: thiếu dữ liệu bắt buộc")
                    continue
                if ten_kho not in wh_map:
                    errors.append(f"Dòng {i}: không tìm thấy kho '{ten_kho}'")
                    continue
                if ten_hang not in item_map:
                    errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                    continue

                key = (so_phieu, ngay, nguoi_lap, ghi_chu, ten_kho)
                if key not in groups:
                    groups[key] = dict(so_phieu=so_phieu, ngay=ngay, nguoi_lap=nguoi_lap,
                                       ghi_chu=ghi_chu, wh_id=wh_map[ten_kho], lines=[])
                it = item_map[ten_hang]
                groups[key]["lines"].append(ReceiptLine(
                    item_type_id=it["id"], item_code=it["code"],
                    item_name=it["name"], unit_of_measure=it["unit_of_measure"],
                    quantity=so_luong, notes=ghi_chu_dong,
                ))

            inserted = 0
            for g in groups.values():
                _insert(Receipt(
                    id=None, reference_number=g["so_phieu"],
                    to_warehouse_id=g["wh_id"], to_warehouse_name="",
                    transaction_date=g["ngay"], created_by=g["nguoi_lap"],
                    notes=g["ghi_chu"], lines=g["lines"],
                ))
                inserted += 1

            msg = f"Đã tạo {inserted} phiếu nhập kho."
            if errors:
                msg += f"\n\nBỏ qua {len(errors)} dòng lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_issues(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from collections import OrderedDict
            from database.xuat_kho import Issue, IssueLine, insert as _insert

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            conn = database.get_conn()
            wh_map = {r["name"]: r["id"] for r in conn.execute(
                "SELECT name, id FROM warehouses WHERE is_active=1"
            ).fetchall()}
            item_map = {r["name"]: dict(r) for r in conn.execute(
                "SELECT code, id, name, unit_of_measure FROM item_types WHERE is_active=1"
            ).fetchall()}

            groups: OrderedDict = OrderedDict()
            errors: list[str] = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                so_phieu    = str(row[0]).strip()
                ngay        = str(row[1]).strip() if row[1] else ""
                nguoi_lap   = str(row[2]).strip() if row[2] else ""
                ghi_chu     = str(row[3]).strip() if row[3] else ""
                ten_kho_tu  = str(row[4]).strip() if row[4] else ""
                ten_kho_den = str(row[5]).strip() if row[5] else ""
                ten_hang    = str(row[6]).strip() if row[6] else ""
                so_luong    = int(row[7]) if row[7] else 0
                ghi_chu_dong = str(row[8]).strip() if len(row) > 8 and row[8] else ""

                if not so_phieu or not ngay or not ten_kho_tu or not ten_hang or not so_luong:
                    errors.append(f"Dòng {i}: thiếu dữ liệu bắt buộc")
                    continue
                if ten_kho_tu not in wh_map:
                    errors.append(f"Dòng {i}: không tìm thấy kho '{ten_kho_tu}'")
                    continue
                if ten_hang not in item_map:
                    errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                    continue
                to_id = wh_map.get(ten_kho_den) if ten_kho_den else None

                key = (so_phieu, ngay, nguoi_lap, ghi_chu, ten_kho_tu, ten_kho_den)
                if key not in groups:
                    groups[key] = dict(
                        so_phieu=so_phieu, ngay=ngay, nguoi_lap=nguoi_lap,
                        ghi_chu=ghi_chu, from_id=wh_map[ten_kho_tu], to_id=to_id,
                        lines=[],
                    )
                it = item_map[ten_hang]
                groups[key]["lines"].append(IssueLine(
                    item_type_id=it["id"], item_code=it["code"],
                    item_name=it["name"], unit_of_measure=it["unit_of_measure"],
                    quantity=so_luong, notes=ghi_chu_dong,
                ))

            inserted = 0
            for g in groups.values():
                _insert(Issue(
                    id=None, reference_number=g["so_phieu"],
                    from_warehouse_id=g["from_id"], from_warehouse_name="",
                    to_warehouse_id=g["to_id"], to_warehouse_name="",
                    transaction_date=g["ngay"], created_by=g["nguoi_lap"],
                    notes=g["ghi_chu"], lines=g["lines"],
                ))
                inserted += 1

            msg = f"Đã tạo {inserted} phiếu xuất kho."
            if errors:
                msg += f"\n\nBỏ qua {len(errors)} dòng lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_transfers(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from collections import OrderedDict
            from database.luan_chuyen import Transfer, TransferLine, insert as _insert

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            conn = database.get_conn()
            wh_map = {r["name"]: r["id"] for r in conn.execute(
                "SELECT name, id FROM warehouses WHERE is_active=1"
            ).fetchall()}
            item_map = {r["name"]: dict(r) for r in conn.execute(
                "SELECT code, id, name, unit_of_measure FROM item_types WHERE is_active=1"
            ).fetchall()}

            groups: OrderedDict = OrderedDict()
            errors: list[str] = []
            VALID_TYPES = {"TONG", "DON_VI"}

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                so_phieu  = str(row[0]).strip()
                ngay      = str(row[1]).strip() if row[1] else ""
                loai      = str(row[2]).strip().upper() if row[2] else ""
                nguoi_lap = str(row[3]).strip() if row[3] else ""
                ghi_chu   = str(row[4]).strip() if row[4] else ""
                ten_tu    = str(row[5]).strip() if row[5] else ""
                ten_den   = str(row[6]).strip() if row[6] else ""
                ten_hang  = str(row[7]).strip() if row[7] else ""
                muc_hh    = str(row[8]).strip().upper() if row[8] else "H1"
                so_luong  = int(row[9]) if row[9] else 0
                ghi_chu_dong = str(row[10]).strip() if len(row) > 10 and row[10] else ""

                if not so_phieu or not ngay or loai not in VALID_TYPES \
                        or not ten_tu or not ten_den or not ten_hang or not so_luong:
                    errors.append(f"Dòng {i}: thiếu hoặc sai dữ liệu bắt buộc")
                    continue
                if ten_tu not in wh_map:
                    errors.append(f"Dòng {i}: không tìm thấy kho '{ten_tu}'")
                    continue
                if ten_den not in wh_map:
                    errors.append(f"Dòng {i}: không tìm thấy kho '{ten_den}'")
                    continue
                if ten_hang not in item_map:
                    errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                    continue
                if muc_hh not in ("H1", "H2", "H3", "H4"):
                    muc_hh = "H1"

                tx_type = "LUAN_CHUYEN_KHO" if loai == "TONG" else "LUAN_CHUYEN_DV"
                key = (so_phieu, ngay, tx_type, nguoi_lap, ghi_chu, ten_tu, ten_den)
                if key not in groups:
                    groups[key] = dict(
                        so_phieu=so_phieu, ngay=ngay, tx_type=tx_type,
                        nguoi_lap=nguoi_lap, ghi_chu=ghi_chu,
                        from_id=wh_map[ten_tu], to_id=wh_map[ten_den],
                        lines=[],
                    )
                it = item_map[ten_hang]
                groups[key]["lines"].append(TransferLine(
                    item_type_id=it["id"], item_code=it["code"],
                    item_name=it["name"], unit_of_measure=it["unit_of_measure"],
                    quantity=so_luong, quality_level=muc_hh, notes=ghi_chu_dong,
                ))

            inserted = 0
            for g in groups.values():
                _insert(Transfer(
                    id=None, reference_number=g["so_phieu"],
                    from_warehouse_id=g["from_id"], from_warehouse_name="",
                    to_warehouse_id=g["to_id"], to_warehouse_name="",
                    transaction_date=g["ngay"], tx_type=g["tx_type"],
                    created_by=g["nguoi_lap"], notes=g["ghi_chu"],
                    lines=g["lines"],
                ))
                inserted += 1

            msg = f"Đã tạo {inserted} phiếu luân chuyển."
            if errors:
                msg += f"\n\nBỏ qua {len(errors)} dòng lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _template_inventory(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_ton_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tồn Kho"
            _write_header(ws, [
                "Tên Kho / ĐV *", "Tên Hàng *", "Mức HH * (H1/H2/H3/H4)",
                "Số Lượng *", "Ghi Chú",
            ])
            ws.append(["Kho Tổng D6", "Súng AK47", "H1", 50, ""])
            ws.append(["Đơn Vị 01",   "Súng AK47", "H2", 10, "Hàng cũ"])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_inventory(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            conn = database.get_conn()
            wh_map = {r["name"]: r["id"] for r in conn.execute(
                "SELECT name, id FROM warehouses WHERE is_active=1"
            ).fetchall()}
            item_map = {r["name"]: dict(r) for r in conn.execute(
                "SELECT id, name FROM item_types WHERE is_active=1"
            ).fetchall()}
            VALID_QL = {"H1", "H2", "H3", "H4"}

            inserted = updated = skipped = 0
            errors: list[str] = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                try:
                    ten_kho  = str(row[0]).strip() if row[0] else ""
                    ten_hang = str(row[1]).strip() if row[1] else ""
                    ql       = str(row[2]).strip().upper() if row[2] else ""
                    so_luong = int(row[3]) if row[3] else 0
                    ghi_chu  = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                    if not ten_kho or not ten_hang or ql not in VALID_QL or so_luong <= 0:
                        errors.append(f"Dòng {i}: thiếu hoặc sai dữ liệu bắt buộc")
                        skipped += 1
                        continue
                    if ten_kho not in wh_map:
                        errors.append(f"Dòng {i}: không tìm thấy kho '{ten_kho}'")
                        skipped += 1
                        continue
                    if ten_hang not in item_map:
                        errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                        skipped += 1
                        continue

                    wh_id      = wh_map[ten_kho]
                    item_id    = item_map[ten_hang]["id"]

                    existing = conn.execute("""
                        SELECT id FROM inventory
                        WHERE warehouse_id=? AND item_type_id=?
                          AND quality_level=? AND is_shared=0
                          AND received_at_unit_date IS NULL
                        LIMIT 1
                    """, (wh_id, item_id, ql)).fetchone()

                    if existing:
                        conn.execute(
                            "UPDATE inventory SET quantity=quantity+?, notes=? WHERE id=?",
                            (so_luong, ghi_chu, existing["id"]),
                        )
                        updated += 1
                    else:
                        conn.execute("""
                            INSERT INTO inventory
                                (warehouse_id, item_type_id, quality_level, quantity, notes)
                            VALUES (?, ?, ?, ?, ?)
                        """, (wh_id, item_id, ql, so_luong, ghi_chu))
                        inserted += 1
                except Exception as row_err:
                    errors.append(f"Dòng {i}: {row_err}")
                    skipped += 1

            conn.commit()
            msg = f"Thêm mới: {inserted}   Cộng dồn: {updated}   Bỏ qua: {skipped}"
            if errors:
                msg += f"\n\nChi tiết lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _template_item_types(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_mat_hang.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mặt Hàng"
            _write_header(ws, [
                "Tên Hàng *", "Đơn Vị Tính *",
                "Niên Hạn (Năm) *", "Đơn Giá", "Ghi Chú",
            ])
            ws.append(["Súng AK47", "cái", 10, 15000000, ""])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _template_warehouses(self):
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kho"
            _write_header(ws, [
                "Tên Kho *", "Loại (TONG/DON_VI) *",
                "Địa Chỉ", "Ghi Chú",
            ])
            ws.append(["Kho Tổng D6", "TONG", "123 Đường ABC", ""])
            ws.append(["Đơn Vị 01", "DON_VI", "456 Đường XYZ", ""])
            _auto_width(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _section_header(text: str) -> QLabel:
    lbl = QLabel(text)
    lbl.setFont(QFont(FONT, 10, QFont.Weight.Bold))
    lbl.setStyleSheet("color: #999;")
    return lbl


def _check_openpyxl(parent: QWidget) -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        QMessageBox.warning(
            parent, "Thiếu thư viện",
            "Vui lòng cài đặt:\n\npip install openpyxl"
        )
        return False


def _write_header(ws, headers: list[str]):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(fill_type="solid", fgColor="111111")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
