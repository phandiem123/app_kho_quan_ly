"""Trang Thống Kê Tồn Kho – dùng chung cho Kho, Đơn Vị, Hàng Dùng Chung."""

from __future__ import annotations
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLineEdit, QComboBox, QCompleter,
    QDialog, QMenu,
)
from PyQt6.QtCore import Qt, pyqtSignal, QSize, QRect
from PyQt6.QtGui import QFont, QCursor, QColor, QPen, QAction, QPainterPath
import datetime
import database

FONT = "Segoe UI"

_TABLE_STYLE = """
    QTableWidget { border: none; background: white; outline: 0; gridline-color: transparent; }
    QHeaderView::section {
        background: white; color: #111; font-size: 12px;
        border: none; border-bottom: 1px solid #efefef;
        padding: 8px 12px; font-weight: normal;
    }
    QTableWidget::item {
        padding: 0 12px; border: none;
        color: #111; font-size: 13px;
    }
    QTableWidget::item:selected { background: #f5f7fa; color: #111; }
"""

_BADGE = {
    "H1": ("H1", "#1a7a4a", "#e6f4ed"),
    "H2": ("H2", "#1565c0", "#e3f0fb"),
    "H3": ("H3", "#e65100", "#fff3e0"),
    "H4": ("H4", "#c0392b", "#fdecea"),
}

_EXPORT_BTN_STYLE = """
    QPushButton { border: 1px solid #d0d0d0; border-radius: 8px; padding: 0 16px;
        font-size: 12px; font-weight: 600; background: #f5f5f5; color: #444; }
    QPushButton:hover { background: #e8e8e8; }
"""


def _check_openpyxl(parent: QWidget) -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.warning(
            parent, "Thiếu thư viện",
            "Vui lòng cài đặt:\n\npip install openpyxl"
        )
        return False


def _write_xlsx_header(ws, headers: list[str]):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="111111")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _auto_width_xlsx(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _find_col(header: list[str], *keywords) -> int | None:
    """Return first column index whose header contains any of the keywords (case-insensitive, NFC)."""
    import unicodedata
    def _n(s): return unicodedata.normalize("NFC", str(s or "").strip().lower())
    norm_kws = [_n(k) for k in keywords]
    for i, h in enumerate(header):
        nh = _n(h)
        if any(kw in nh for kw in norm_kws):
            return i
    return None


def _az_key(name: str) -> str:
    """Sort key: bỏ dấu tiếng Việt, chữ thường — cho phép sort A→Z theo ký tự cơ bản."""
    import unicodedata
    nfd = unicodedata.normalize("NFD", str(name or "").lower())
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn")


def _norm_name(s) -> str:
    """Chuẩn hoá tên để so khớp: NFC, strip, thu gọn dấu cách thừa."""
    import unicodedata, re
    s = unicodedata.normalize("NFC", str(s or "").strip())
    return re.sub(r'\s+', ' ', s)


class _SmartHeader(QHeaderView):
    """Header với sort arrows, eye-icon ẩn/hiện, và drag reorder."""
    sortRequested = pyqtSignal(int, bool)   # (logical_col, ascending)
    visToggled    = pyqtSignal(int, bool)   # (logical_col, is_now_visible)

    _H = 40

    def __init__(self, skip: set | frozenset = frozenset({0}), parent=None):
        super().__init__(Qt.Orientation.Horizontal, parent)
        self._skip = frozenset(skip)
        self._scol = -1
        self._sasc = True
        self.setSectionsMovable(True)
        self.setSectionsClickable(True)
        self.setHighlightSections(False)

    def sizeHint(self):
        s = super().sizeHint()
        return QSize(s.width(), self._H)

    def paintSection(self, painter, rect, li):
        if not rect.isValid():
            return
        painter.save()
        painter.setClipRect(rect)
        painter.setRenderHint(painter.RenderHint.Antialiasing)

        sorted_col = li == self._scol
        painter.fillRect(rect, QColor("#f0f5ff") if sorted_col else QColor("white"))
        painter.setPen(QPen(QColor("#efefef")))
        painter.drawLine(rect.left(), rect.bottom(), rect.right(), rect.bottom())

        label = self.model().headerData(li, Qt.Orientation.Horizontal) or ""
        is_skip = li in self._skip
        cy = float(rect.top() + rect.height() / 2)
        right_pad = 8

        if not is_skip:
            # ── Sort arrows: ↑↓ always visible ──────────────────────────
            aw = 9
            ax = float(rect.right() - aw - 8)

            up_clr = QColor("#2255cc") if (sorted_col and self._sasc)     else QColor("#ddd")
            dn_clr = QColor("#2255cc") if (sorted_col and not self._sasc) else QColor("#ddd")

            up = QPainterPath()
            up.moveTo(ax + aw / 2, cy - 7)
            up.lineTo(ax,          cy - 1)
            up.lineTo(ax + aw,     cy - 1)
            up.closeSubpath()
            painter.fillPath(up, up_clr)

            dn = QPainterPath()
            dn.moveTo(ax + aw / 2, cy + 7)
            dn.lineTo(ax,          cy + 1)
            dn.lineTo(ax + aw,     cy + 1)
            dn.closeSubpath()
            painter.fillPath(dn, dn_clr)

            right_pad += aw + 10

        lrect = QRect(rect.left() + 10, rect.top(),
                      rect.width() - right_pad - 8, rect.height())
        painter.setFont(QFont(FONT, 11))
        painter.setPen(QColor("#1a3a8a" if sorted_col else "#111"))
        painter.drawText(lrect, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, label)
        painter.restore()

    def mousePressEvent(self, event):
        if event.button() != Qt.MouseButton.LeftButton:
            super().mousePressEvent(event)
            return
        li = self.logicalIndexAt(event.pos())
        if li < 0 or li in self._skip:
            super().mousePressEvent(event)
            return
        if self._scol == li:
            self._sasc = not self._sasc
        else:
            self._scol = li
            self._sasc = True
        self.sortRequested.emit(self._scol, self._sasc)
        self.viewport().update()

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            li = self.logicalIndexAt(event.pos())
            if li >= 0 and li not in self._skip:
                self.setSectionHidden(li, True)
                self.visToggled.emit(li, False)
                self.viewport().update()
                return
        super().mouseDoubleClickEvent(event)

    def reset_sort(self):
        self._scol = -1
        self._sasc = True
        self.viewport().update()

    def set_skip(self, skip):
        self._skip = frozenset(skip)
        self.viewport().update()


class _ColVisBtn(QPushButton):
    """Nút 'Hiển thị thông tin' mở dropdown với checkbox ẩn/hiện từng cột."""

    def __init__(self, parent=None):
        super().__init__("Hiển thị thông tin  ▾", parent)
        self.setFixedHeight(34)
        self.setFont(QFont(FONT, 11))
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.setStyleSheet("""
            QPushButton { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 14px; background: white; color: #111; }
            QPushButton:hover { background: #f5f5f5; }
        """)
        self._menu = QMenu(self)
        self._menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #e0e0e0;
                    border-radius: 8px; padding: 4px; min-width: 160px; }
            QMenu::item { padding: 7px 28px 7px 10px; border-radius: 4px;
                          color: #111; font-size: 12px; }
            QMenu::item:selected { background: #f5f5f5; }
            QMenu::indicator { width: 14px; height: 14px; margin-left: 4px; }
            QMenu::indicator:checked { image: none; background: #3366cc;
                border-radius: 3px; }
            QMenu::indicator:unchecked { background: white;
                border: 1px solid #ccc; border-radius: 3px; }
        """)
        self._table: QTableWidget | None = None
        self._hdr: _SmartHeader | None = None
        self._actions: dict[int, QAction] = {}
        self.clicked.connect(self._show)

    def bind(self, table: QTableWidget, hdr, col_names: list[str], skip: set):
        self._table = table
        self._hdr = hdr
        self._menu.clear()
        self._actions.clear()
        for i, name in enumerate(col_names):
            if i in skip:
                continue
            act = QAction(name, self._menu)
            act.setCheckable(True)
            act.setChecked(True)
            act.toggled.connect(lambda chk, c=i: self._toggle(c, chk))
            self._actions[i] = act
            self._menu.addAction(act)

    def _show(self):
        if self._table:
            for li, act in self._actions.items():
                act.blockSignals(True)
                act.setChecked(not self._table.isColumnHidden(li))
                act.blockSignals(False)
        self._menu.exec(self.mapToGlobal(self.rect().bottomLeft()))

    def _toggle(self, li: int, checked: bool):
        if self._table:
            self._table.setColumnHidden(li, not checked)
            if self._hdr:
                self._hdr.viewport().update()

    def on_vis_toggled(self, li: int, visible: bool):
        if li in self._actions:
            a = self._actions[li]
            a.blockSignals(True)
            a.setChecked(visible)
            a.blockSignals(False)


def _apply_inv_qty(wh_id: int, item_type_id: int, quality_level: str, new_qty: int):
    """Overwrite the total inventory quantity for (wh, item, ql).
    If multiple rows exist, keeps the first and zeros the rest.
    """
    conn = database.get_conn()
    rows = conn.execute("""
        SELECT id FROM inventory
        WHERE warehouse_id=? AND item_type_id=? AND quality_level=? AND is_shared=0
        ORDER BY received_at_unit_date NULLS FIRST, id
    """, (wh_id, item_type_id, quality_level)).fetchall()
    if not rows:
        if new_qty > 0:
            conn.execute("""
                INSERT INTO inventory (warehouse_id, item_type_id, quality_level, quantity, is_shared)
                VALUES (?, ?, ?, ?, 0)
            """, (wh_id, item_type_id, quality_level, new_qty))
    else:
        conn.execute("UPDATE inventory SET quantity=? WHERE id=?", (new_qty, rows[0]["id"]))
        for r in rows[1:]:
            conn.execute("UPDATE inventory SET quantity=0 WHERE id=?", (r["id"],))
    conn.commit()


# ── Edit-quantity dialog ───────────────────────────────────────────────────────

class _EditQtyDialog(QDialog):
    _QL_STYLE = {
        "H1": ("#1a7a4a", "#e6f4ed"),
        "H2": ("#1565c0", "#e3f0fb"),
        "H3": ("#e65100", "#fff3e0"),
        "H4": ("#c0392b", "#fdecea"),
    }

    def __init__(self, item_name: str, ql: str, current: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sửa số lượng")
        self.setModal(True)
        self.setFixedWidth(380)
        self._result: int | None = None
        self._build(item_name, ql, current)

    def _build(self, item_name: str, ql: str, current: int):
        self.setStyleSheet("QDialog { background: white; }")
        root = QVBoxLayout(self)
        root.setContentsMargins(28, 24, 28, 22)
        root.setSpacing(0)

        # ── Header ──────────────────────────────────────────────────────────
        title = QLabel("Sửa số lượng tồn kho")
        title.setFont(QFont(FONT, 14, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        root.addWidget(title)
        root.addSpacing(5)

        sub = QLabel(item_name)
        sub.setFont(QFont(FONT, 11))
        sub.setStyleSheet("color: #888;")
        sub.setWordWrap(True)
        root.addWidget(sub)

        root.addSpacing(18)
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("QFrame { color: #f0f0f0; }")
        root.addWidget(sep)
        root.addSpacing(20)

        # ── QL badge + input ─────────────────────────────────────────────────
        fg, bg = self._QL_STYLE.get(ql, ("#111", "#f0f0f0"))
        row = QHBoxLayout()
        row.setSpacing(14)

        badge = QLabel(ql)
        badge.setFixedSize(52, 52)
        badge.setFont(QFont(FONT, 15, QFont.Weight.Bold))
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge.setStyleSheet(f"background: {bg}; color: {fg}; border-radius: 11px;")
        row.addWidget(badge)

        inp_col = QVBoxLayout()
        inp_col.setSpacing(5)
        inp_lbl = QLabel("Số lượng mới")
        inp_lbl.setFont(QFont(FONT, 10))
        inp_lbl.setStyleSheet("color: #aaa;")
        inp_col.addWidget(inp_lbl)

        self._inp = QLineEdit(str(current))
        self._inp.setFont(QFont(FONT, 18, QFont.Weight.Bold))
        self._inp.setFixedHeight(46)
        self._inp.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._inp.setStyleSheet("""
            QLineEdit {
                border: 2px solid #e0e0e0; border-radius: 10px;
                background: #fafafa; color: #111; padding: 0 12px;
            }
            QLineEdit:focus { border-color: #111; background: white; }
        """)
        self._inp.selectAll()
        self._inp.returnPressed.connect(self._on_save)
        inp_col.addWidget(self._inp)

        row.addLayout(inp_col, 1)
        root.addLayout(row)
        root.addSpacing(24)

        # ── Buttons ──────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        cancel = QPushButton("Hủy")
        cancel.setFixedHeight(40)
        cancel.setFont(QFont(FONT, 12))
        cancel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        cancel.setStyleSheet("""
            QPushButton {
                border: 1px solid #e0e0e0; border-radius: 9px;
                background: white; color: #555;
            }
            QPushButton:hover { background: #f5f5f5; }
        """)
        cancel.clicked.connect(self.reject)

        save = QPushButton("Lưu")
        save.setFixedHeight(40)
        save.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        save.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        save.setStyleSheet("""
            QPushButton {
                border: none; border-radius: 9px;
                background: #111; color: white;
            }
            QPushButton:hover { background: #333; }
            QPushButton:pressed { background: #000; }
        """)
        save.clicked.connect(self._on_save)

        btn_row.addWidget(cancel)
        btn_row.addWidget(save)
        root.addLayout(btn_row)

    def _on_save(self):
        text = self._inp.text().strip()
        if not text.isdigit():
            self._inp.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #e53935; border-radius: 10px;
                    background: #fff5f5; color: #111; padding: 0 12px;
                }
            """)
            return
        self._result = int(text)
        self.accept()

    def result_value(self) -> int | None:
        return self._result


# ── Summary cards ─────────────────────────────────────────────────────────────

class _SummaryCard(QFrame):
    def __init__(self, label: str, value: str, sub: str = ""):
        super().__init__()
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet("""
            QFrame { background: white; border: 1px solid #e8e8e8; border-radius: 10px; }
            QLabel { border: none; }
        """)
        v = QVBoxLayout(self)
        v.setContentsMargins(20, 14, 20, 14)
        v.setSpacing(4)

        lbl = QLabel(label)
        lbl.setFont(QFont(FONT, 11))
        lbl.setStyleSheet("color: #111;")
        v.addWidget(lbl)

        val = QLabel(value)
        val.setFont(QFont(FONT, 22, QFont.Weight.Bold))
        val.setStyleSheet("color: #111;")
        v.addWidget(val)

        if sub:
            s = QLabel(sub)
            s.setFont(QFont(FONT, 10))
            s.setStyleSheet("color: #aaa;")
            v.addWidget(s)

        self._val = val
        self._sub_lbl = QLabel(sub) if sub else None

    def set_value(self, value: str, sub: str = ""):
        self._val.setText(value)
        if self._sub_lbl:
            self._sub_lbl.setText(sub)


# ── Main page ─────────────────────────────────────────────────────────────────

class ThongKePage(QWidget):
    """
    mode = 'kho'        → Thống kê tại Kho Tổng (H1, H4)
    mode = 'don_vi'     → Thống kê tại Đơn Vị (H1–H4)
    mode = 'shared'     → Thống kê Hàng Dùng Chung
    """

    def __init__(self, mode: str):
        super().__init__()
        self._mode = mode
        self.setStyleSheet("ThongKePage { background: #fafafa; }")
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(16)

        # ── Title + filter row ─────────────────────────────────────────────
        title_row = QHBoxLayout()
        title = QLabel(self._page_title())
        title.setFont(QFont(FONT, 18, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        title_row.addWidget(title)
        title_row.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kho / mặt hàng...")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        title_row.addWidget(self._search)

        self._wh_combo = QComboBox()
        self._wh_combo.setFixedHeight(34)
        self._wh_combo.setFont(QFont(FONT, 12))
        self._wh_combo.setMinimumWidth(180)
        self._wh_combo.setStyleSheet("""
            QComboBox { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white; color: #111; }
            QComboBox::drop-down { subcontrol-origin: padding;
                subcontrol-position: center right; width: 28px;
                border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px; border-bottom-right-radius: 8px;
                background: white; }
            QComboBox::down-arrow { width: 10px; height: 10px; image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent; }
            QComboBox QAbstractItemView { border: 1px solid #e0e0e0;
                border-radius: 6px; background: white; color: #111;
                selection-background-color: #f0f0f0; }
        """)
        self._wh_combo.addItem("Tất cả kho", None)
        self._wh_combo.currentIndexChanged.connect(self._apply_filter)
        title_row.addWidget(self._wh_combo)
        root.addLayout(title_row)

        # ── Summary cards ──────────────────────────────────────────────────
        card_row = QHBoxLayout()
        card_row.setSpacing(16)
        self._cards: dict[str, _SummaryCard] = {}
        for key, label in self._card_defs():
            card = _SummaryCard(label, "—")
            self._cards[key] = card
            card_row.addWidget(card)
        card_row.addStretch()
        root.addLayout(card_row)

        # ── Table card ─────────────────────────────────────────────────────
        tbl_card = QFrame()
        tbl_card.setStyleSheet(
            "QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }"
        )
        tbl_v = QVBoxLayout(tbl_card)
        tbl_v.setContentsMargins(0, 0, 0, 0)

        cols = self._table_columns()
        self._table = QTableWidget(0, len(cols))
        self._table.setHorizontalHeaderLabels(cols)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._setup_col_widths()
        tbl_v.addWidget(self._table)
        root.addWidget(tbl_card, 1)

        self.refresh()

    # ── Config per mode ───────────────────────────────────────────────────────

    def _page_title(self) -> str:
        return {
            "kho": "Tồn Kho – Tất Cả Kho",
            "don_vi": "Tồn Kho – Đơn Vị",
            "shared": "Tồn Kho – Hàng Dùng Chung",
        }[self._mode]

    def _card_defs(self) -> list[tuple[str, str]]:
        if self._mode == "kho":
            return [("total", "Tổng tồn"), ("h1", "H1"), ("h2", "H2"),
                    ("h3", "H3"), ("h4", "H4 (Chờ TXL)")]
        if self._mode == "don_vi":
            return [("total", "Tổng tồn"), ("h1", "H1"), ("h2", "H2"),
                    ("h3", "H3"), ("h4", "H4")]
        return [("total", "Tổng tồn"), ("borrowed", "Đang mượn"), ("available", "Sẵn sàng")]

    def _table_columns(self) -> list[str]:
        if self._mode == "kho":
            return ["STT", "Loại", "Kho / Đơn Vị", "Tên Hàng",
                    "ĐVT", "H1", "H2", "H3", "H4", "Tổng"]
        if self._mode == "don_vi":
            return ["STT", "Đơn Vị", "Tên Hàng", "ĐVT",
                    "H1", "H2", "H3", "H4", "Tổng", "Ngày Nhận ĐV"]
        return ["STT", "Kho", "Tên Hàng", "ĐVT",
                "Tổng", "Đang Mượn", "Sẵn Sàng"]

    def _setup_col_widths(self):
        h = self._table.horizontalHeader()
        if self._mode == "kho":
            specs = [
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 56),
                (QHeaderView.ResizeMode.Fixed, 64),
                (QHeaderView.ResizeMode.Fixed, 64),
                (QHeaderView.ResizeMode.Fixed, 64),
                (QHeaderView.ResizeMode.Fixed, 64),
                (QHeaderView.ResizeMode.Fixed, 72),
            ]
        elif self._mode == "don_vi":
            specs = [
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 110),
            ]
        else:
            specs = [
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 60),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 88),
                (QHeaderView.ResizeMode.Fixed, 88),
            ]
        for i, (mode, w) in enumerate(specs):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)

    # ── Data ──────────────────────────────────────────────────────────────────

    def refresh(self):
        self._reload_warehouses()
        self._reload_data()

    def _reload_warehouses(self):
        current = self._wh_combo.currentData()
        self._wh_combo.blockSignals(True)
        self._wh_combo.clear()
        self._wh_combo.addItem("Tất cả", None)
        conn = database.get_conn()
        if self._mode == "kho":
            rows = conn.execute(
                "SELECT id, name, code, type FROM warehouses WHERE is_active=1 ORDER BY type, name"
            ).fetchall()
            for r in rows:
                self._wh_combo.addItem(r["name"], r["id"])
        else:
            wh_type = {"don_vi": "DON_VI", "shared": "TONG"}[self._mode]
            rows = conn.execute(
                "SELECT id, name, code FROM warehouses WHERE type=? AND is_active=1 ORDER BY name",
                (wh_type,)
            ).fetchall()
            for r in rows:
                self._wh_combo.addItem(r["name"], r["id"])
        for i in range(self._wh_combo.count()):
            if self._wh_combo.itemData(i) == current:
                self._wh_combo.setCurrentIndex(i)
                break
        self._wh_combo.blockSignals(False)

    def _reload_data(self):
        conn = database.get_conn()
        wh_id = self._wh_combo.currentData()
        params: list = []

        if self._mode == "kho":
            base = """
                SELECT w.type AS wh_type, w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name,
                       t.unit_of_measure,
                       SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                       SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                       SUM(i.quantity) AS total
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE i.is_shared = 0
            """
            if wh_id:
                base += " AND w.id = ?"
                params.append(wh_id)
            base += " GROUP BY w.id, t.id HAVING total > 0 ORDER BY w.type, w.name, t.name"
            rows = conn.execute(base, params).fetchall()
            self._raw = rows

            self._cards["total"].set_value(str(sum(r["total"] for r in rows)))
            self._cards["h1"].set_value(str(sum(r["h1"] for r in rows)))
            self._cards["h2"].set_value(str(sum(r["h2"] for r in rows)))
            self._cards["h3"].set_value(str(sum(r["h3"] for r in rows)))
            self._cards["h4"].set_value(str(sum(r["h4"] for r in rows)))

        elif self._mode == "don_vi":
            base = """
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name,
                       t.unit_of_measure,
                       SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                       SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                       SUM(i.quantity) AS total,
                       MIN(i.received_at_unit_date) AS first_received
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE w.type = 'DON_VI' AND i.is_shared = 0
            """
            if wh_id:
                base += " AND w.id = ?"
                params.append(wh_id)
            base += " GROUP BY w.id, t.id HAVING total > 0 ORDER BY w.name, t.name"
            rows = conn.execute(base, params).fetchall()
            self._raw = rows

            total = sum(r["total"] for r in rows)
            self._cards["total"].set_value(str(total))
            self._cards["h1"].set_value(str(sum(r["h1"] for r in rows)))
            self._cards["h2"].set_value(str(sum(r["h2"] for r in rows)))
            self._cards["h3"].set_value(str(sum(r["h3"] for r in rows)))
            self._cards["h4"].set_value(str(sum(r["h4"] for r in rows)))

        else:  # shared
            base = """
                SELECT w.code AS wh_code, w.name AS wh_name,
                       t.code AS item_code, t.name AS item_name,
                       t.unit_of_measure,
                       SUM(i.quantity) AS total_qty,
                       COALESCE((
                           SELECT SUM(sb.quantity)
                           FROM shared_borrows sb
                           JOIN inventory si ON si.id = sb.inventory_id
                           WHERE si.item_type_id = t.id
                             AND si.warehouse_id = w.id
                             AND sb.status = 'DANG_MUON'
                       ), 0) AS borrowed
                FROM inventory i
                JOIN warehouses w ON w.id = i.warehouse_id
                JOIN item_types t  ON t.id = i.item_type_id
                WHERE i.is_shared = 1
            """
            if wh_id:
                base += " AND w.id = ?"
                params.append(wh_id)
            base += " GROUP BY w.id, t.id HAVING total_qty > 0 ORDER BY w.name, t.name"
            rows = conn.execute(base, params).fetchall()
            self._raw = rows

            total = sum(r["total_qty"] for r in rows)
            borrowed = sum(r["borrowed"] for r in rows)
            self._cards["total"].set_value(str(total))
            self._cards["borrowed"].set_value(str(borrowed))
            self._cards["available"].set_value(str(total - borrowed))

        self._apply_filter()

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        rows = self._raw if hasattr(self, "_raw") else []
        if q:
            rows = [r for r in rows
                    if q in r["item_name"].lower()
                    or q in r["wh_name"].lower()
                    or q in r["wh_code"].lower()]
        self._load_table(rows)

    def _load_table(self, rows):
        self._table.setRowCount(0)
        for i, r in enumerate(rows):
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setRowHeight(ri, 52)

            if self._mode == "kho":
                type_label = "Kho" if r["wh_type"] == "TONG" else "Đơn Vị"
                cells = [
                    (str(i + 1), True),
                    (type_label, True),
                    (r["wh_name"], False),
                    (r["item_name"], False),
                    (r["unit_of_measure"], True),
                    (str(r["h1"]) if r["h1"] else "—", True),
                    (str(r["h2"]) if r["h2"] else "—", True),
                    (str(r["h3"]) if r["h3"] else "—", True),
                    (str(r["h4"]) if r["h4"] else "—", True),
                    (str(r["total"]), True),
                ]
            elif self._mode == "don_vi":
                cells = [
                    (str(i + 1), True),
                    (r["wh_name"], False),
                    (r["item_name"], False),
                    (r["unit_of_measure"], True),
                    (str(r["h1"]) if r["h1"] else "—", True),
                    (str(r["h2"]) if r["h2"] else "—", True),
                    (str(r["h3"]) if r["h3"] else "—", True),
                    (str(r["h4"]) if r["h4"] else "—", True),
                    (str(r["total"]), True),
                    (r["first_received"] or "—", True),
                ]
            else:
                avail = r["total_qty"] - r["borrowed"]
                cells = [
                    (str(i + 1), True),
                    (r["wh_name"], False),
                    (r["item_name"], False),
                    (r["unit_of_measure"], True),
                    (str(r["total_qty"]), True),
                    (str(r["borrowed"]) if r["borrowed"] else "—", True),
                    (str(avail) if avail else "—", True),
                ]

            for c, (val, center) in enumerate(cells):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 12))
                if center:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # highlight H4 column
                if ((self._mode == "kho" and c == 8) or
                        (self._mode == "don_vi" and c == 7)):
                    if val not in ("—", "0"):
                        cell.setForeground(Qt.GlobalColor.red)
                self._table.setItem(ri, c, cell)


# ── H4 detail panel ──────────────────────────────────────────────────────────

class _H4DetailPanel(QWidget):
    def __init__(self, row: dict, parent=None):
        super().__init__(parent)
        self.setAutoFillBackground(True)
        self.setStyleSheet("""
            _H4DetailPanel { background: #f5f5f5; }
            QLabel { background: transparent; border: none; }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(28, 14, 28, 14)
        root.setSpacing(8)

        title = QLabel("Chi Tiết Phiếu Chuyển H4")
        title.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        root.addWidget(title)

        def pair(label, value):
            w = QWidget(); w.setStyleSheet("background: transparent;")
            hl = QHBoxLayout(w); hl.setContentsMargins(0, 0, 0, 0); hl.setSpacing(6)
            lbl = QLabel(label + ":"); lbl.setFont(QFont(FONT, 11))
            lbl.setStyleSheet("color: #888;"); lbl.setFixedWidth(80)
            val = QLabel(value or "—"); val.setFont(QFont(FONT, 11, QFont.Weight.Medium))
            val.setStyleSheet("color: #111;")
            hl.addWidget(lbl); hl.addWidget(val, 1)
            return w

        info = QHBoxLayout(); info.setSpacing(32)
        info.addWidget(pair("Số Phiếu", row["reference_number"]))
        info.addWidget(pair("Ngày",     row["transaction_date"]))
        info.addWidget(pair("Kho",      row["wh_name"]))
        info.addStretch()
        root.addLayout(info)

        sub_lbl = QLabel("Danh Sách Mặt Hàng")
        sub_lbl.setFont(QFont(FONT, 10, QFont.Weight.Bold))
        sub_lbl.setStyleSheet("color: #555;")
        root.addWidget(sub_lbl)

        conn = database.get_conn()
        lines = conn.execute("""
            SELECT t.name AS item_name, t.unit_of_measure,
                   tl.quality_level_from, tl.quantity
            FROM transaction_lines tl
            JOIN item_types t ON t.id = tl.item_type_id
            WHERE tl.transaction_id = ?
            ORDER BY t.name
        """, (row["id"],)).fetchall()

        sub = QTableWidget(0, 4)
        sub.setHorizontalHeaderLabels(["STT", "Tên Hàng", "ĐVT", "Số Lượng"])
        sub.verticalHeader().setVisible(False)
        sub.setShowGrid(False)
        sub.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        sub.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        sub.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        sub.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sub.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sub.setStyleSheet("""
            QTableWidget { border: none; background: #efefef; outline: 0; border-radius: 6px; }
            QHeaderView::section { background: #e8e8e8; color: #777; font-size: 10px;
                border: none; padding: 4px 8px; }
            QTableWidget::item { padding: 4px 8px; color: #111; font-size: 12px; border: none; }
        """)
        sh = sub.horizontalHeader()
        F, S = QHeaderView.ResizeMode.Fixed, QHeaderView.ResizeMode.Stretch
        for i, (mode, w) in enumerate([(F, 44), (S, None), (F, 72), (F, 84)]):
            sh.setSectionResizeMode(i, mode)
            if w:
                sub.setColumnWidth(i, w)

        for i, ln in enumerate(lines):
            ri = sub.rowCount(); sub.insertRow(ri); sub.setRowHeight(ri, 34)
            for c, (val, center) in enumerate([
                (str(i + 1),          True),
                (ln["item_name"],     False),
                (ln["unit_of_measure"] or "—", True),
                (str(ln["quantity"]), True),
            ]):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 11))
                if center:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                sub.setItem(ri, c, cell)

        root.addWidget(sub)
        self._line_count = len(lines)

    def preferred_height(self) -> int:
        return 160 + max(1, self._line_count) * 34


# ── Thin wrappers ─────────────────────────────────────────────────────────────

class ThongKeSharedPage(QWidget):
    """Hàng dùng chung – tồn kho (H1–H3), đang cho mượn, chờ thanh xử lý (H4)."""

    _COLS_TON   = ["STT", "Tên Hàng", "ĐVT", "H3", "H4"]
    _COLS_MUON  = ["STT", "Số Phiếu", "Ngày Mượn", "Tên Hàng",
                   "ĐVT", "Số Lượng", "Đơn Vị / Người Mượn"]
    _COLS_H4    = ["STT", "Số Phiếu", "Ngày", "Kho", "Số M.Hàng", ""]

    _TABS = [
        ("ton",      "Tồn Kho"),
        ("muon",     "Đang Cho Mượn"),
        ("phieu_h4", "Phiếu Chuyển H4"),
    ]

    _SORT_KEYS: dict[str, dict] = {
        "ton": {
            1: lambda r: (r["item_name"] or "").lower(),
            2: lambda r: (r["unit_of_measure"] or "").lower(),
            3: lambda r: r["h3"],
            4: lambda r: r["h4"],
        },
        "muon": {
            1: lambda r: (r["reference_number"] or "").lower(),
            2: lambda r: r["transaction_date"] or "",
            3: lambda r: (r["item_name"] or "").lower(),
            4: lambda r: (r["unit_of_measure"] or "").lower(),
            5: lambda r: r["quantity"],
            6: lambda r: (r["borrower"] or "").lower(),
        },
        "phieu_h4": {
            1: lambda r: (r["reference_number"] or "").lower(),
            2: lambda r: r["transaction_date"] or "",
            3: lambda r: (r["wh_name"] or "").lower(),
            4: lambda r: r["line_count"],
        },
    }

    def __init__(self):
        super().__init__()
        self.setStyleSheet("ThongKeSharedPage { background: #fafafa; }")
        self._active_tab     = "ton"
        self._active_wh_id:   int | None = None
        self._active_wh_type: str | None = None  # "TONG" | "DON_VI" | None
        self._raw_ton:    list = []
        self._raw_muon:   list = []
        self._raw_h4:     list = []
        self._h4_rows_order: list = []
        self._expanded_h4_row: int | None = None
        self._sort: dict[str, tuple[int, bool]] = {
            "ton": (-1, True), "muon": (-1, True), "phieu_h4": (-1, True),
        }
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 52, 24)
        root.setSpacing(0)

        top = QHBoxLayout()
        title = QLabel("Hàng Dùng Chung")
        title.setFont(QFont(FONT, 18, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        top.addWidget(title)
        top.addStretch()

        _btn_style = """
            QPushButton { border: 1px solid #d0d0d0; border-radius: 8px; padding: 0 16px;
                font-size: 12px; font-weight: 600; background: #f5f5f5; color: #444; }
            QPushButton:hover { background: #e8e8e8; }
        """
        btn_nhan_sk = QPushButton("Nhập Hàng Dùng Chung")
        btn_nhan_sk.setFixedHeight(34)
        btn_nhan_sk.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        btn_nhan_sk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_nhan_sk.setStyleSheet(_btn_style)
        btn_nhan_sk.clicked.connect(self._on_nhan_tu_su_kien)
        top.addWidget(btn_nhan_sk)
        top.addSpacing(8)

        btn_xuat = QPushButton("+ Xuất Cho Mượn")
        btn_xuat.setFixedHeight(34)
        btn_xuat.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        btn_xuat.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_xuat.setStyleSheet(_btn_style)
        btn_xuat.clicked.connect(self._on_xuat_cho_muon)
        top.addWidget(btn_xuat)
        top.addSpacing(8)

        btn_h4 = QPushButton("Chuyển Sang H4")
        btn_h4.setFixedHeight(34)
        btn_h4.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        btn_h4.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_h4.setStyleSheet(_btn_style)
        btn_h4.clicked.connect(self._on_chuyen_h4)
        top.addWidget(btn_h4)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm mặt hàng...")
        self._search.setFixedSize(220, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)

        self._muon_search = QLineEdit()
        self._muon_search.setPlaceholderText("Tìm số phiếu / đơn vị / ngày mượn...")
        self._muon_search.setFixedSize(300, 34)
        self._muon_search.setFont(QFont(FONT, 12))
        self._muon_search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._muon_search.textChanged.connect(self._apply_filter)
        self._muon_search.setVisible(False)

        _combo_style = """
            QComboBox { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white; color: #111; }
            QComboBox::drop-down { subcontrol-origin: padding;
                subcontrol-position: center right; width: 28px;
                border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px; border-bottom-right-radius: 8px;
                background: white; }
            QComboBox::down-arrow { width: 10px; height: 10px; image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent; }
            QComboBox QAbstractItemView { border: 1px solid #e0e0e0;
                border-radius: 6px; background: white; color: #111;
                selection-background-color: #f0f0f0; }
        """

        self._wh_combo = QComboBox()
        self._wh_combo.setFixedHeight(34)
        self._wh_combo.setMinimumWidth(220)
        self._wh_combo.setFont(QFont(FONT, 12))
        self._wh_combo.setStyleSheet(_combo_style)
        self._wh_combo.currentIndexChanged.connect(self._on_wh_changed)

        root.addLayout(top)
        root.addSpacing(20)

        # summary cards
        card_h = QHBoxLayout()
        card_h.setSpacing(14)
        self._cards: dict[str, _SummaryCard] = {}
        for key, label in [
            ("muon",     "Đang Cho Mượn"),
            ("san_sang", "Sẵn Sàng"),
            ("txl",      "Chờ TXL (H4)"),
        ]:
            c = _SummaryCard(label, "—")
            self._cards[key] = c
            card_h.addWidget(c)
        card_h.addStretch()
        root.addLayout(card_h)
        root.addSpacing(16)

        # tab bar
        tab_frame = QFrame()
        tab_frame.setFixedHeight(44)
        tab_frame.setStyleSheet("QFrame { border: none; background: transparent; }")
        tab_h = QHBoxLayout(tab_frame)
        tab_h.setContentsMargins(0, 4, 0, 4)
        tab_h.setSpacing(6)
        self._tab_btns: dict[str, QPushButton] = {}
        self._vis_btn = _ColVisBtn()
        for key, label in self._TABS:
            btn = QPushButton(label)
            btn.setFont(QFont(FONT, 11))
            btn.setFixedHeight(32)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self._tab_btns[key] = btn
            btn.clicked.connect(lambda _, k=key: self._on_tab(k))
            tab_h.addWidget(btn)
        tab_h.addStretch()
        tab_h.addWidget(self._muon_search)
        tab_h.addSpacing(8)
        tab_h.addWidget(self._wh_combo)
        tab_h.addSpacing(8)
        tab_h.addWidget(self._search)
        tab_h.addSpacing(8)
        tab_h.addWidget(self._vis_btn)
        self._apply_tab_style()
        root.addWidget(tab_frame)
        root.addSpacing(8)

        # table card
        tbl_card = QFrame()
        tbl_card.setStyleSheet(
            "QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }"
        )
        tbl_v = QVBoxLayout(tbl_card)
        tbl_v.setContentsMargins(0, 0, 0, 0)

        self._smart_hdr = _SmartHeader({0})
        self._table = QTableWidget(0, len(self._COLS_TON))
        self._table.setHorizontalHeader(self._smart_hdr)
        self._table.setHorizontalHeaderLabels(self._COLS_TON)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._smart_hdr.sortRequested.connect(self._on_sort)
        self._smart_hdr.visToggled.connect(self._vis_btn.on_vis_toggled)
        self._table.cellClicked.connect(self._on_h4_cell_clicked)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_table_context_menu)
        tbl_v.addWidget(self._table)
        root.addWidget(tbl_card, 1)

        self.refresh()

    def _apply_tab_style(self):
        for key, btn in self._tab_btns.items():
            if key == self._active_tab:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: #111; color: white; font-size: 11px; font-weight: 600; }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton { border: 1px solid #e0e0e0; border-radius: 7px; padding: 0 14px;
                        background: white; color: #111; font-size: 11px; }
                    QPushButton:hover { background: #f5f5f5; color: #111; }
                """)

    # ── Interactions ──────────────────────────────────────────────────────────

    def _on_tab(self, key: str):
        self._sort[self._active_tab] = (self._smart_hdr._scol, self._smart_hdr._sasc)
        self._expanded_h4_row = None
        self._active_tab = key
        sc, sa = self._sort.get(key, (-1, True))
        self._smart_hdr._scol = sc
        self._smart_hdr._sasc = sa
        self._apply_tab_style()
        is_muon = (key == "muon")
        self._wh_combo.setVisible(True)
        self._search.setVisible(not is_muon)
        self._muon_search.setVisible(is_muon)
        if key == "ton":
            self._search.setPlaceholderText("Tìm mặt hàng...")
        elif key == "phieu_h4":
            self._search.setPlaceholderText("Tìm số phiếu / kho...")
        self._apply_filter()

    def _on_wh_changed(self, _):
        data = self._wh_combo.currentData()
        if data is None:  # separator — không làm gì
            return
        self._active_wh_id, self._active_wh_type = data
        self._reload_data()

    def _on_nhap_moi(self):
        from ui.dialogs.nhap_kho_form import NhapKhoFormDialog
        dlg = NhapKhoFormDialog(self, subtype="shared_new")
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh()

    def _on_nhan_tu_su_kien(self):
        from ui.dialogs.nhap_kho_form import NhapKhoFormDialog
        dlg = NhapKhoFormDialog(self, subtype="shared_return")
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh()

    def _on_chuyen_h4(self):
        from ui.dialogs.chuyen_h4_form import ChuyenH4FormDialog
        wh_id = self._active_wh_id if self._active_wh_type == "TONG" else None
        dlg = ChuyenH4FormDialog(self, wh_id=wh_id)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh()

    def _on_xuat_cho_muon(self):
        from ui.dialogs.xuat_kho_form import XuatKhoFormDialog
        dlg = XuatKhoFormDialog(self, subtype="shared_loan")
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh()

    # ── Data ──────────────────────────────────────────────────────────────────

    def refresh(self):
        self._reload_warehouses()
        self._reload_data()

    def _reload_warehouses(self):
        current = self._wh_combo.currentData()
        self._wh_combo.blockSignals(True)
        self._wh_combo.clear()
        conn = database.get_conn()
        tong_rows = conn.execute(
            "SELECT id, name FROM warehouses WHERE type='TONG' AND is_active=1 ORDER BY name"
        ).fetchall()
        dv_rows = conn.execute(
            "SELECT id, name FROM warehouses WHERE type='DON_VI' AND is_active=1 ORDER BY name"
        ).fetchall()
        self._wh_combo.addItem("Tất cả kho", (None, "TONG"))
        for r in tong_rows:
            self._wh_combo.addItem(r["name"], (r["id"], "TONG"))
        self._wh_combo.insertSeparator(self._wh_combo.count())
        self._wh_combo.addItem("Tất cả đơn vị", (None, "DON_VI"))
        for r in dv_rows:
            self._wh_combo.addItem(r["name"], (r["id"], "DON_VI"))
        restored = False
        if current is not None:
            for i in range(self._wh_combo.count()):
                if self._wh_combo.itemData(i) == current:
                    self._wh_combo.setCurrentIndex(i)
                    restored = True
                    break
        if not restored:
            self._wh_combo.setCurrentIndex(0)
        self._wh_combo.blockSignals(False)
        data = self._wh_combo.currentData()
        if isinstance(data, tuple):
            self._active_wh_id, self._active_wh_type = data
        else:
            self._active_wh_id, self._active_wh_type = None, "TONG"

    def _reload_data(self):
        conn = database.get_conn()
        wh_id   = self._active_wh_id
        wh_type = self._active_wh_type

        # Tab 1: Tồn kho H3/H4
        # TONG+id → lọc theo kho; TONG+None → tất cả kho
        # DON_VI+id → hàng DC đơn vị đó đang giữ; DON_VI+None → hàng DC tất cả đơn vị
        if wh_type == "DON_VI":
            if wh_id:
                inv_f, p_ton = "AND i.warehouse_id = ?", [wh_id]
            else:
                inv_f, p_ton = "AND wh.type = 'DON_VI'", []
            self._raw_ton = conn.execute(f"""
                SELECT t.id AS item_type_id, t.name AS item_name, t.unit_of_measure,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4
                FROM inventory i
                JOIN item_types t ON t.id = i.item_type_id
                JOIN warehouses wh ON wh.id = i.warehouse_id
                WHERE i.is_shared=1 AND i.quality_level IN ('H3','H4') {inv_f}
                GROUP BY t.id HAVING (h3 + h4) > 0 ORDER BY t.name
            """, p_ton).fetchall()
        else:  # TONG
            inv_f = "AND i.warehouse_id = ?" if wh_id else ""
            p_ton = [wh_id] if wh_id else []
            self._raw_ton = conn.execute(f"""
                SELECT t.id AS item_type_id, t.name AS item_name, t.unit_of_measure,
                       SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                       SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4
                FROM inventory i
                JOIN item_types t ON t.id = i.item_type_id
                WHERE i.is_shared=1 AND i.quality_level IN ('H3','H4') {inv_f}
                GROUP BY t.id HAVING (h3 + h4) > 0 ORDER BY t.name
            """, p_ton).fetchall()

        # Tab 2: Đang cho mượn
        # TONG+id → lọc kho xuất; DON_VI+id → lọc đơn vị mượn; else → tất cả
        if wh_type == "TONG" and wh_id:
            muon_f, p_muon = "AND tx.from_warehouse_id = ?", [wh_id]
        elif wh_type == "DON_VI" and wh_id:
            muon_f, p_muon = "AND tx.to_warehouse_id = ?", [wh_id]
        else:
            muon_f, p_muon = "", []
        self._raw_muon = conn.execute(f"""
            SELECT tx.transaction_date, tx.reference_number,
                   t.name AS item_name, t.unit_of_measure,
                   tl.quantity,
                   COALESCE(wto.name, tx.supplier, '') AS borrower
            FROM transactions tx
            JOIN transaction_lines tl ON tl.transaction_id = tx.id
            JOIN item_types t ON t.id = tl.item_type_id
            LEFT JOIN warehouses wto ON wto.id = tx.to_warehouse_id
            WHERE tx.type='MUON' {muon_f}
            ORDER BY tx.transaction_date DESC, tx.id DESC, tl.id
        """, p_muon).fetchall()

        # Tab 3: Phiếu chuyển H4 (chỉ lọc khi chọn kho TONG cụ thể)
        if wh_type == "TONG" and wh_id:
            h4_f, p_h4 = "AND tx.to_warehouse_id = ?", [wh_id]
        else:
            h4_f, p_h4 = "", []
        self._raw_h4 = conn.execute(f"""
            SELECT tx.id, tx.reference_number, tx.transaction_date,
                   w.name AS wh_name,
                   COUNT(tl.id) AS line_count
            FROM transactions tx
            JOIN warehouses w ON w.id = tx.to_warehouse_id
            JOIN transaction_lines tl ON tl.transaction_id = tx.id
            WHERE tx.type='CHUYEN_H4' {h4_f}
            GROUP BY tx.id
            ORDER BY tx.transaction_date DESC, tx.id DESC
        """, p_h4).fetchall()

        h3_total = sum(r["h3"] for r in self._raw_ton)
        h4_total = sum(r["h4"] for r in self._raw_ton)
        muon     = sum(r["quantity"] for r in self._raw_muon)
        self._cards["muon"].set_value(str(muon))
        self._cards["san_sang"].set_value(str(h3_total))
        self._cards["txl"].set_value(str(h4_total))
        self._apply_filter()

    def _apply_filter(self):
        if self._active_tab == "ton":
            q = self._search.text().strip().lower()
            rows = [r for r in self._raw_ton
                    if not q or q in r["item_name"].lower()]
            self._load_ton(rows)
        elif self._active_tab == "muon":
            q = self._muon_search.text().strip().lower()
            rows = [r for r in self._raw_muon
                    if not q
                    or q in (r["reference_number"] or "").lower()
                    or q in (r["borrower"] or "").lower()
                    or q in (r["transaction_date"] or "")]
            self._load_muon(rows)
        else:
            q = self._search.text().strip().lower()
            rows = [r for r in self._raw_h4
                    if not q
                    or q in (r["reference_number"] or "").lower()
                    or q in (r["wh_name"] or "").lower()
                    or q in (r["transaction_date"] or "")]
            self._load_phieu_h4(rows)

    # ── Table helpers ─────────────────────────────────────────────────────────

    def _setup_cols(self, cols: list[str], specs: list[tuple]):
        self._table.setColumnCount(len(cols))
        self._table.setHorizontalHeaderLabels(cols)
        for i in range(len(cols)):
            self._table.setColumnHidden(i, False)
        h = self._table.horizontalHeader()
        for i, (mode, w) in enumerate(specs):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)
        skip = {0}
        self._smart_hdr.set_skip(skip)
        self._vis_btn.bind(self._table, self._smart_hdr, cols, skip)
        self._smart_hdr.viewport().update()

    def _mk(self, val: str, center=False, clr=None) -> QTableWidgetItem:
        cell = QTableWidgetItem(val)
        cell.setFont(QFont(FONT, 12))
        if center:
            cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if clr == "red":
            cell.setForeground(Qt.GlobalColor.red)
        return cell

    def _sort_rows(self, rows):
        sc, sa = self._sort.get(self._active_tab, (-1, True))
        if sc <= 0:
            return rows
        fn = self._SORT_KEYS.get(self._active_tab, {}).get(sc)
        if fn is None:
            return rows
        return sorted(rows, key=fn, reverse=not sa)

    def _on_sort(self, logical: int, ascending: bool):
        self._sort[self._active_tab] = (logical, ascending)
        self._apply_filter()

    def _load_ton(self, rows):
        rows = self._sort_rows(rows)
        F, S = QHeaderView.ResizeMode.Fixed, QHeaderView.ResizeMode.Stretch
        self._setup_cols(self._COLS_TON, [
            (F, 52), (S, None), (F, 60), (F, 80), (F, 80),
        ])
        self._table.setRowCount(0)
        for i, r in enumerate(rows):
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setRowHeight(ri, 48)
            for c, args in enumerate([
                (str(i + 1),                              True),
                (r["item_name"],                          False),
                (r["unit_of_measure"],                    True),
                (str(r["h3"]) if r["h3"] else "—",       True),
                (str(r["h4"]) if r["h4"] else "—",       True, "red" if r["h4"] else None),
            ]):
                self._table.setItem(ri, c, self._mk(*args))

    def _load_muon(self, rows):
        rows = self._sort_rows(rows)
        F, S = QHeaderView.ResizeMode.Fixed, QHeaderView.ResizeMode.Stretch
        self._setup_cols(self._COLS_MUON, [
            (F, 52), (F, 110), (F, 110), (S, None),
            (F, 60), (F, 88), (S, None),
        ])
        self._table.setRowCount(0)
        for i, r in enumerate(rows):
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setRowHeight(ri, 48)
            for c, args in enumerate([
                (str(i + 1),                         True),
                (r["reference_number"] or "—",       True),
                (r["transaction_date"],              True),
                (r["item_name"],                     False),
                (r["unit_of_measure"],               True),
                (str(r["quantity"]),                 True),
                (r["borrower"] or "—",               False),
            ]):
                self._table.setItem(ri, c, self._mk(*args))

    def _load_phieu_h4(self, rows):
        rows = self._sort_rows(rows)
        self._h4_rows_order = list(rows)
        self._expanded_h4_row = None
        F, S = QHeaderView.ResizeMode.Fixed, QHeaderView.ResizeMode.Stretch
        self._setup_cols(self._COLS_H4, [
            (F, 52), (F, 130), (F, 110), (S, None), (F, 90), (F, 44),
        ])
        self._table.clearSpans()
        self._table.setRowCount(len(rows) * 2)
        n = len(self._COLS_H4)
        for i, r in enumerate(rows):
            dr  = i * 2
            det = i * 2 + 1
            self._table.setRowHeight(dr, 48)
            self._table.setRowHeight(det, 0)
            self._table.setSpan(det, 0, 1, n)

            stt_cell = self._mk(str(i + 1), True)
            self._table.setItem(dr, 0, stt_cell)
            for c, args in enumerate([
                (r["reference_number"] or "—",  True),
                (r["transaction_date"],          True),
                (r["wh_name"] or "—",           False),
                (str(r["line_count"]),           True),
            ], start=1):
                self._table.setItem(dr, c, self._mk(*args))

            tx_id = r["id"]
            btn = QPushButton("•••")
            btn.setFlat(True)
            btn.setFont(QFont(FONT, 14))
            btn.setFixedSize(40, 40)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setStyleSheet("""
                QPushButton { color: #111; border: none; background: transparent;
                    letter-spacing: 2px; border-radius: 6px; }
                QPushButton:hover { background: #f0f0f0; color: #555; }
            """)
            btn.clicked.connect(lambda _, tid=tx_id, b=btn: self._show_phieu_h4_menu(tid, b))
            self._table.setCellWidget(dr, n - 1, btn)

    def _on_h4_cell_clicked(self, row: int, col: int):
        if self._active_tab != "phieu_h4":
            return
        if col == len(self._COLS_H4) - 1:  # "•••" column — let button handle it
            return
        if row % 2 == 1:  # detail row
            return
        det_row = row + 1

        if self._expanded_h4_row == row:
            self._table.setRowHeight(det_row, 0)
            self._table.removeCellWidget(det_row, 0)
            self._expanded_h4_row = None
            return

        if self._expanded_h4_row is not None:
            prev_det = self._expanded_h4_row + 1
            self._table.setRowHeight(prev_det, 0)
            self._table.removeCellWidget(prev_det, 0)

        data_idx = row // 2
        if data_idx >= len(self._h4_rows_order):
            return
        r = self._h4_rows_order[data_idx]
        panel = _H4DetailPanel(r)
        h = panel.preferred_height()
        self._table.setCellWidget(det_row, 0, panel)
        self._table.setRowHeight(det_row, h)
        self._expanded_h4_row = row

    def _on_table_context_menu(self, pos):
        if self._active_tab != "phieu_h4":
            return
        row = self._table.rowAt(pos.y())
        if row < 0 or row % 2 == 1:
            return
        data_idx = row // 2
        if data_idx >= len(self._h4_rows_order):
            return
        tx_id = self._h4_rows_order[data_idx]["id"]
        self._show_phieu_h4_menu(tx_id, None)

    def _show_phieu_h4_menu(self, tx_id: int, anchor):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #e0e0e0;
                border-radius: 8px; padding: 4px; }
            QMenu::item { padding: 8px 20px; border-radius: 5px;
                font-size: 12px; color: #111; }
            QMenu::item:selected { background: #f5f5f5; }
        """)
        act_edit = QAction("Sửa", self)
        act_del  = QAction("Xóa", self)
        menu.addAction(act_edit)
        menu.addSeparator()
        menu.addAction(act_del)
        if anchor:
            pos = anchor.mapToGlobal(anchor.rect().bottomLeft())
        else:
            pos = QCursor.pos()
        act = menu.exec(pos)
        if act == act_edit:
            self._edit_phieu_h4(tx_id)
        elif act == act_del:
            self._confirm_delete_phieu_h4(tx_id)

    def _edit_phieu_h4(self, tx_id: int):
        from ui.dialogs.chuyen_h4_form import ChuyenH4FormDialog
        dlg = ChuyenH4FormDialog(self, tx_id=tx_id)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh()

    def _confirm_delete_phieu_h4(self, tx_id: int):
        from PyQt6.QtWidgets import QMessageBox
        conn = database.get_conn()
        tx = conn.execute(
            "SELECT reference_number FROM transactions WHERE id=?", (tx_id,)
        ).fetchone()
        ref = tx["reference_number"] or f"#{tx_id}" if tx else f"#{tx_id}"
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Xóa phiếu chuyển H4 '{ref}'?\nTồn kho sẽ được hoàn lại.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        tx_full = conn.execute(
            "SELECT to_warehouse_id FROM transactions WHERE id=?", (tx_id,)
        ).fetchone()
        if not tx_full:
            return
        wh_id = tx_full["to_warehouse_id"]
        for ol in conn.execute(
            "SELECT item_type_id, quality_level_from, quantity"
            " FROM transaction_lines WHERE transaction_id=?", (tx_id,)
        ).fetchall():
            conn.execute("""
                UPDATE inventory SET quantity=quantity+?
                WHERE warehouse_id=? AND item_type_id=? AND quality_level=? AND is_shared=1
            """, (ol["quantity"], wh_id, ol["item_type_id"], ol["quality_level_from"]))
            conn.execute("""
                UPDATE inventory SET quantity=MAX(0, quantity-?)
                WHERE warehouse_id=? AND item_type_id=? AND quality_level='H4' AND is_shared=1
            """, (ol["quantity"], wh_id, ol["item_type_id"]))
        conn.execute("DELETE FROM transaction_lines WHERE transaction_id=?", (tx_id,))
        conn.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
        conn.commit()
        self.refresh()


# ── Per-warehouse tab page ────────────────────────────────────────────────────

class ThongKeKhoPage(QWidget):
    """Tồn kho theo từng kho/đơn vị — mỗi kho là một tab riêng."""

    _COLS = ["", "STT", "Tên Hàng", "ĐVT", "Niên Hạn (Năm)", "Đơn Giá",
             "H1", "H2", "H3", "H4", "Tổng"]
    _COLS_SEARCH = ["", "STT", "Kho", "Tên Hàng", "ĐVT", "Niên Hạn (Năm)", "Đơn Giá",
                    "H1", "H2", "H3", "H4", "Tổng"]

    _SORT_KEYS_NORMAL: dict = {
        2: lambda r: (r["item_name"] or "").lower(),
        3: lambda r: (r["unit_of_measure"] or "").lower(),
        4: lambda r: r["total_lifespan_months"] or 0,
        5: lambda r: r["don_gia"] or 0,
        6: lambda r: r["h1"],
        7: lambda r: r["h2"],
        8: lambda r: r["h3"],
        9: lambda r: r["h4"],
        10: lambda r: r["total"],
    }
    _SORT_KEYS_SEARCH: dict = {
        2: lambda r: (r["wh_name"] or "").lower(),
        3: lambda r: (r["item_name"] or "").lower(),
        4: lambda r: (r["unit_of_measure"] or "").lower(),
        5: lambda r: r["total_lifespan_months"] or 0,
        6: lambda r: r["don_gia"] or 0,
        7: lambda r: r["h1"],
        8: lambda r: r["h2"],
        9: lambda r: r["h3"],
        10: lambda r: r["h4"],
        11: lambda r: r["total"],
    }

    def __init__(self):
        super().__init__()
        self.setStyleSheet("ThongKeKhoPage { background: #fafafa; }")
        self._wh_list: list = []
        self._active_wh_id: int | None = None
        self._raw: list = []
        self._year_map: dict = {}
        self._search_raw: list = []
        self._search_year_map: dict = {}
        self._was_searching: bool = False
        self._tab_btns: dict[int, QPushButton] = {}
        self._sort_col = -1
        self._sort_asc = True
        self._sort_s_col = -1
        self._sort_s_asc = True
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(0)

        top = QHBoxLayout()
        title = QLabel("Tồn Kho – Tất Cả Kho")
        title.setFont(QFont(FONT, 18, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        top.addWidget(title)
        top.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kiếm toàn kho")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        top.addWidget(self._search)
        root.addLayout(top)
        root.addSpacing(16)

        # ── Tab bar — hidden while searching ──────────────────────────────
        self._tab_row = QWidget()
        self._tab_row.setStyleSheet("background: transparent;")
        tab_row_v = QVBoxLayout(self._tab_row)
        tab_row_v.setContentsMargins(0, 0, 0, 16)
        tab_row_v.setSpacing(0)

        tab_frame = QFrame()
        tab_frame.setFixedHeight(44)
        tab_frame.setStyleSheet("QFrame { border: none; background: transparent; }")
        tab_h = QHBoxLayout(tab_frame)
        tab_h.setContentsMargins(0, 4, 0, 4)
        tab_h.setSpacing(6)
        self._tab_layout = tab_h
        self._tab_layout.addStretch()
        tab_row_v.addWidget(tab_frame)
        root.addWidget(self._tab_row)

        # ── Summary cards ──────────────────────────────────────────────────
        card_h = QHBoxLayout()
        card_h.setSpacing(12)
        self._cards: dict[str, _SummaryCard] = {}
        for key, label in [("total", "Tổng"), ("h1", "H1"), ("h2", "H2"),
                            ("h3", "H3"), ("h4", "H4 (Chờ TXL)")]:
            c = _SummaryCard(label, "—")
            self._cards[key] = c
            card_h.addWidget(c)
        card_h.addStretch()

        _search_style = """
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """
        self._vis_btn = _ColVisBtn()
        card_h.addWidget(self._vis_btn)

        self._local_search = QLineEdit()
        self._local_search.setPlaceholderText("Tìm trong kho này...")
        self._local_search.setFixedSize(220, 34)
        self._local_search.setFont(QFont(FONT, 12))
        self._local_search.setStyleSheet(_search_style)
        self._local_search.textChanged.connect(self._apply_local_filter)
        card_h.addWidget(self._local_search)

        root.addLayout(card_h)
        root.addSpacing(10)

        export_h = QHBoxLayout()
        self._btn_del_sel = QPushButton("Xoá đã chọn (0)")
        self._btn_del_sel.setFixedHeight(34)
        self._btn_del_sel.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        self._btn_del_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_del_sel.setStyleSheet(
            "QPushButton { background: #d32f2f; color: white; border-radius: 8px;"
            " padding: 0 16px; border: none; }"
            " QPushButton:hover { background: #b71c1c; }"
        )
        self._btn_del_sel.clicked.connect(self._on_delete_selected)
        self._btn_del_sel.setVisible(False)
        export_h.addWidget(self._btn_del_sel)
        export_h.addStretch()
        for label, slot in [
            ("Tải Mẫu",    self._download_template),
            ("Nhập Excel", self._import_excel),
            ("Xuất Excel", self._export_excel),
        ]:
            btn = QPushButton(label)
            btn.setFixedHeight(34)
            btn.setFont(QFont(FONT, 12, QFont.Weight.Bold))
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setStyleSheet(_EXPORT_BTN_STYLE)
            btn.clicked.connect(slot)
            export_h.addWidget(btn)
        root.addLayout(export_h)
        root.addSpacing(16)

        # ── Table ──────────────────────────────────────────────────────────
        tbl_card = QFrame()
        tbl_card.setStyleSheet(
            "QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }"
        )
        tbl_v = QVBoxLayout(tbl_card)
        tbl_v.setContentsMargins(0, 0, 0, 0)

        self._smart_hdr = _SmartHeader({0, 1})
        self._table = QTableWidget(0, len(self._COLS))
        self._table.setHorizontalHeader(self._smart_hdr)
        self._table.setHorizontalHeaderLabels(self._COLS)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.cellDoubleClicked.connect(self._on_cell_dblclick)
        self._smart_hdr.sortRequested.connect(self._on_sort)
        self._smart_hdr.visToggled.connect(self._vis_btn.on_vis_toggled)
        self._table.itemChanged.connect(
            lambda item: self._update_sel_count() if item.column() == 0 else None
        )
        self._smart_hdr.setSectionsClickable(True)
        self._smart_hdr.sectionClicked.connect(
            lambda col: self._toggle_all() if col == 0 else None
        )

        tbl_v.addWidget(self._table)
        root.addWidget(tbl_card, 1)

        self.refresh()

    # ── Tabs ──────────────────────────────────────────────────────────────────

    def _rebuild_tabs(self):
        for btn in list(self._tab_btns.values()):
            self._tab_layout.removeWidget(btn)
            btn.deleteLater()
        self._tab_btns.clear()

        for r in self._wh_list:
            btn = QPushButton(r["name"])
            btn.setFont(QFont(FONT, 11))
            btn.setFixedHeight(32)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self._tab_btns[r["id"]] = btn
            wid = r["id"]
            btn.clicked.connect(lambda _, w=wid: self._on_tab(w))
            self._tab_layout.insertWidget(self._tab_layout.count() - 1, btn)

        self._apply_tab_style()

    def _on_tab(self, wh_id: int):
        self._active_wh_id = wh_id
        self._apply_tab_style()
        self._search.clear()
        self._reload_data()

    def _apply_tab_style(self):
        for wid, btn in self._tab_btns.items():
            if wid == self._active_wh_id:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: #111; color: white; font-size: 11px; font-weight: 600; }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton { border: 1px solid #e0e0e0; border-radius: 7px; padding: 0 14px;
                        background: white; color: #111; font-size: 11px; }
                    QPushButton:hover { background: #f5f5f5; color: #111; }
                """)

    def _configure_table(self, search_mode: bool):
        cols = self._COLS_SEARCH if search_mode else self._COLS
        self._table.setColumnCount(len(cols))
        self._table.setHorizontalHeaderLabels(cols)
        for i in range(len(cols)):
            self._table.setColumnHidden(i, False)
        h = self._table.horizontalHeader()
        specs = (
            [
                (QHeaderView.ResizeMode.Fixed, 36),
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Fixed, 100),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 56),
                (QHeaderView.ResizeMode.Fixed, 110),
                (QHeaderView.ResizeMode.Fixed, 120),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 80),
            ] if search_mode else [
                (QHeaderView.ResizeMode.Fixed, 36),
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 56),
                (QHeaderView.ResizeMode.Fixed, 110),
                (QHeaderView.ResizeMode.Fixed, 120),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 80),
            ]
        )
        for i, (mode, w) in enumerate(specs):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)
        skip = {0, 1}
        self._smart_hdr.set_skip(skip)
        self._vis_btn.bind(self._table, self._smart_hdr, cols, skip)
        if search_mode:
            self._sort_s_col = -1
            self._sort_s_asc = True
        else:
            self._sort_col = -1
            self._sort_asc = True
        self._smart_hdr.reset_sort()

    # ── Data ──────────────────────────────────────────────────────────────────

    def switch_to(self, wh_id: int):
        self._active_wh_id = wh_id
        self.refresh()

    def refresh(self):
        self._search_raw = []
        self._was_searching = False
        self._tab_row.setVisible(True)
        self._configure_table(search_mode=False)
        conn = database.get_conn()
        rows = conn.execute(
            "SELECT id, code, name, type FROM warehouses WHERE is_active=1 AND type='TONG' ORDER BY name"
        ).fetchall()
        self._wh_list = rows
        known_ids = {r["id"] for r in rows}
        if self._active_wh_id not in known_ids:
            self._active_wh_id = rows[0]["id"] if rows else None
        self._rebuild_tabs()
        self._reload_data()

    def _reload_data(self):
        if self._active_wh_id is None:
            self._raw = []
            self._update_cards([])
            self._apply_filter()
            return
        conn = database.get_conn()
        rows = conn.execute("""
            SELECT t.id AS item_type_id,
                   t.code AS item_code, t.name AS item_name,
                   t.unit_of_measure,
                   t.total_lifespan_months,
                   SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                   SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                   SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                   SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                   SUM(i.quantity) AS total,
                   (SELECT tl.unit_price
                    FROM transaction_lines tl
                    JOIN transactions tx ON tx.id = tl.transaction_id
                    WHERE tl.item_type_id = t.id
                      AND tx.to_warehouse_id = ?
                      AND COALESCE(tl.unit_price, 0) > 0
                    ORDER BY tx.transaction_date DESC, tl.id DESC
                    LIMIT 1) AS don_gia
            FROM inventory i
            JOIN item_types t ON t.id = i.item_type_id
            WHERE i.warehouse_id = ? AND i.is_shared = 0 AND i.quantity > 0
            GROUP BY t.id
            ORDER BY t.name
        """, (self._active_wh_id, self._active_wh_id)).fetchall()
        self._raw = rows

        year = str(datetime.date.today().year)
        wh_id = self._active_wh_id
        yr_rows = conn.execute("""
            SELECT sub.item_type_id, sub.ql,
                   SUM(sub.nhap) AS nhap, SUM(sub.xuat) AS xuat
            FROM (
                SELECT tl.item_type_id, tl.quality_level_to AS ql,
                       tl.quantity AS nhap, 0 AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                WHERE tx.to_warehouse_id = ?
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_to IS NOT NULL
                UNION ALL
                SELECT tl.item_type_id, tl.quality_level_from AS ql,
                       0 AS nhap, tl.quantity AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                WHERE tx.from_warehouse_id = ?
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_from IS NOT NULL
            ) sub
            GROUP BY sub.item_type_id, sub.ql
        """, (wh_id, year, wh_id, year)).fetchall()
        self._year_map = {
            (r["item_type_id"], r["ql"]): (r["nhap"], r["xuat"])
            for r in yr_rows
        }

        self._update_cards(rows)
        self._apply_filter()

    def _load_search_data(self):
        conn = database.get_conn()
        year = str(datetime.date.today().year)
        self._search_raw = conn.execute("""
            SELECT w.id AS wh_id, w.code AS wh_code, w.name AS wh_name,
                   t.id AS item_type_id,
                   t.code AS item_code, t.name AS item_name,
                   t.unit_of_measure,
                   t.total_lifespan_months,
                   SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                   SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                   SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                   SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                   SUM(i.quantity) AS total,
                   (SELECT tl2.unit_price
                    FROM transaction_lines tl2
                    JOIN transactions tx2 ON tx2.id = tl2.transaction_id
                    WHERE tl2.item_type_id = t.id
                      AND tx2.to_warehouse_id = w.id
                      AND COALESCE(tl2.unit_price, 0) > 0
                    ORDER BY tx2.transaction_date DESC, tl2.id DESC
                    LIMIT 1) AS don_gia
            FROM inventory i
            JOIN warehouses w ON w.id = i.warehouse_id
            JOIN item_types t ON t.id = i.item_type_id
            WHERE i.is_shared = 0 AND i.quantity > 0
              AND w.type = 'TONG' AND w.is_active = 1
            GROUP BY w.id, t.id
            ORDER BY t.name, w.name
        """).fetchall()

        yr_rows = conn.execute("""
            SELECT sub.wh_id, sub.item_type_id, sub.ql,
                   SUM(sub.nhap) AS nhap, SUM(sub.xuat) AS xuat
            FROM (
                SELECT tx.to_warehouse_id AS wh_id,
                       tl.item_type_id, tl.quality_level_to AS ql,
                       tl.quantity AS nhap, 0 AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                JOIN warehouses w ON w.id = tx.to_warehouse_id
                WHERE w.type = 'TONG' AND w.is_active = 1
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_to IS NOT NULL
                UNION ALL
                SELECT tx.from_warehouse_id AS wh_id,
                       tl.item_type_id, tl.quality_level_from AS ql,
                       0 AS nhap, tl.quantity AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                JOIN warehouses w ON w.id = tx.from_warehouse_id
                WHERE w.type = 'TONG' AND w.is_active = 1
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_from IS NOT NULL
            ) sub
            GROUP BY sub.wh_id, sub.item_type_id, sub.ql
        """, (year, year)).fetchall()
        self._search_year_map = {
            (r["wh_id"], r["item_type_id"], r["ql"]): (r["nhap"], r["xuat"])
            for r in yr_rows
        }

    def _update_cards(self, rows):
        self._cards["total"].set_value(str(sum(r["total"] for r in rows)))
        for ql in ("h1", "h2", "h3", "h4"):
            self._cards[ql].set_value(str(sum(r[ql] for r in rows)))

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        is_searching = bool(q)

        if is_searching and not self._was_searching:
            self._load_search_data()
            self._configure_table(search_mode=True)
            self._tab_row.setVisible(False)
            self._local_search.setVisible(False)
            self._local_search.clear()
        elif not is_searching and self._was_searching:
            self._configure_table(search_mode=False)
            self._tab_row.setVisible(True)
            self._local_search.setVisible(True)

        self._was_searching = is_searching

        if is_searching:
            rows = [r for r in self._search_raw
                    if q in r["item_name"].lower()
                    or q in r["wh_name"].lower() or q in r["wh_code"].lower()]
            self._load_table(rows, search_mode=True)
        else:
            self._apply_local_filter()

    def _apply_local_filter(self):
        q = self._local_search.text().strip().lower()
        rows = self._raw if not q else [
            r for r in self._raw
            if q in r["item_name"].lower()
        ]
        self._load_table(rows, search_mode=False)

    def _load_table(self, rows, search_mode: bool = False):
        rows = self._sort_rows(rows, search_mode)
        self._table.blockSignals(True)
        self._table.setRowCount(0)
        for i, r in enumerate(rows):
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setRowHeight(ri, 48)

            # Checkbox col 0
            chk = QTableWidgetItem()
            chk.setFlags(
                Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsUserCheckable
            )
            chk.setCheckState(Qt.CheckState.Unchecked)
            chk.setData(Qt.ItemDataRole.UserRole, r["item_type_id"])
            self._table.setItem(ri, 0, chk)

            mm = r["total_lifespan_months"]
            nien_han = f"{mm // 12} năm" if mm else "—"
            nien_han_red = False

            dg = r["don_gia"]
            don_gia = f"{int(dg):,}".replace(",", ".") if dg else "—"

            tid = r["item_type_id"]

            def _tip(ql: str, current: int) -> str:
                if search_mode:
                    nhap, xuat = self._search_year_map.get((r["wh_id"], tid, ql), (0, 0))
                else:
                    nhap, xuat = self._year_map.get((tid, ql), (0, 0))
                so_du = current - nhap + xuat

                def _row(label, val, last=False):
                    sep = "" if last else "border-bottom: 1px solid #ebebeb;"
                    return (
                        f'<tr>'
                        f'<td style="padding: 9px 28px 9px 0; color: #111; {sep}">{label}</td>'
                        f'<td style="padding: 9px 0; color: #111; font-weight: 600; text-align: right; {sep}">{val}</td>'
                        f'</tr>'
                    )

                return (
                    '<table style="font-family: \'Segoe UI\'; font-size: 13px;'
                    ' min-width: 230px; border-spacing: 0;">'
                    '<tr><td colspan="2" style="padding: 0 0 10px 0;">'
                    '<b style="font-size: 14px; color: #111;">Thông tin Chi tiết</b>'
                    '</td></tr>'
                    + _row("Số dư đầu năm", so_du)
                    + _row("Tổng nhập", nhap)
                    + _row("Tổng xuất", xuat)
                    + _row("Hiện tại", current, last=True)
                    + '</table>'
                )

            if search_mode:
                cells = [
                    (str(i + 1),           True,  False, None),
                    (r["wh_name"],         False, False, None),
                    (r["item_name"],       False, False, None),
                    (r["unit_of_measure"], True,  False, None),
                    (nien_han,             True,  nien_han_red, None),
                    (don_gia,              True,  False, None),
                    (str(r["h1"]),         True,  False, _tip("H1", r["h1"])),
                    (str(r["h2"]),         True,  False, _tip("H2", r["h2"])),
                    (str(r["h3"]),         True,  False, _tip("H3", r["h3"])),
                    (str(r["h4"]),         True,  r["h4"] > 0, _tip("H4", r["h4"])),
                    (str(r["total"]),      True,  False, None),
                ]
            else:
                cells = [
                    (str(i + 1),           True,  False, None),
                    (r["item_name"],       False, False, None),
                    (r["unit_of_measure"], True,  False, None),
                    (nien_han,             True,  nien_han_red, None),
                    (don_gia,              True,  False, None),
                    (str(r["h1"]),         True,  False, _tip("H1", r["h1"])),
                    (str(r["h2"]),         True,  False, _tip("H2", r["h2"])),
                    (str(r["h3"]),         True,  False, _tip("H3", r["h3"])),
                    (str(r["h4"]),         True,  r["h4"] > 0, _tip("H4", r["h4"])),
                    (str(r["total"]),      True,  False, None),
                ]

            for c, (val, center, red, tip) in enumerate(cells, start=1):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 12))
                if center:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if red:
                    cell.setForeground(Qt.GlobalColor.red)
                if tip:
                    cell.setToolTip(tip)
                self._table.setItem(ri, c, cell)

            # Store edit key on H1-H4 cells so double-click knows what to update
            wh_id = r["wh_id"] if search_mode else self._active_wh_id
            h_start = 7 if search_mode else 6
            for qi, ql in enumerate(("H1", "H2", "H3", "H4")):
                item = self._table.item(ri, h_start + qi)
                if item:
                    item.setData(Qt.ItemDataRole.UserRole,
                                 (wh_id, r["item_type_id"], ql, r["item_name"]))

        self._table.blockSignals(False)
        self._update_sel_count()

    def _sort_rows(self, rows, search_mode: bool):
        col = self._sort_s_col if search_mode else self._sort_col
        asc = self._sort_s_asc if search_mode else self._sort_asc
        if col <= 0:
            return rows
        keys = self._SORT_KEYS_SEARCH if search_mode else self._SORT_KEYS_NORMAL
        fn = keys.get(col)
        if fn is None:
            return rows
        return sorted(rows, key=fn, reverse=not asc)

    # ── Multi-select ──────────────────────────────────────────────────────────

    def _get_checked_ids(self) -> list[int]:
        ids = []
        for row in range(self._table.rowCount()):
            item = self._table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                ids.append(item.data(Qt.ItemDataRole.UserRole))
        return ids

    def _update_sel_count(self):
        ids = self._get_checked_ids()
        n = len(ids)
        total = self._table.rowCount()
        all_checked = (total > 0 and n == total)
        hdr_label = "☑" if all_checked else "☐"
        self._table.horizontalHeaderItem(0).setText(hdr_label)
        if n > 0:
            self._btn_del_sel.setText(f"Xoá đã chọn ({n})")
            self._btn_del_sel.setVisible(True)
        else:
            self._btn_del_sel.setVisible(False)

    def _toggle_all(self):
        total = self._table.rowCount()
        if total == 0:
            return
        ids = self._get_checked_ids()
        new_state = Qt.CheckState.Unchecked if len(ids) == total else Qt.CheckState.Checked
        self._table.blockSignals(True)
        for row in range(total):
            item = self._table.item(row, 0)
            if item:
                item.setCheckState(new_state)
        self._table.blockSignals(False)
        self._update_sel_count()

    def _on_delete_selected(self):
        from PyQt6.QtWidgets import QMessageBox
        ids = self._get_checked_ids()
        if not ids:
            return
        reply = QMessageBox.question(
            self, "Xác nhận xoá",
            f"Xoá tồn kho của {len(ids)} mặt hàng đã chọn khỏi kho này?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        conn = database.get_conn()
        for item_type_id in ids:
            conn.execute(
                "DELETE FROM inventory WHERE warehouse_id=? AND item_type_id=? AND is_shared=0",
                (self._active_wh_id, item_type_id),
            )
        conn.commit()
        self.refresh()

    def _download_template(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_ton_kho_kho.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            items = sorted(
                conn.execute(
                    "SELECT name, unit_of_measure, total_lifespan_months, unit_price"
                    " FROM item_types WHERE is_active=1"
                ).fetchall(),
                key=lambda r: _az_key(r["name"]),
            )
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tồn Kho"
            _write_xlsx_header(ws, ["STT", "Tên hàng", "ĐVT", "Niên hạn (Năm)", "Đơn giá", "H1", "H2", "H3", "H4", "Tổng"])
            for i, r in enumerate(items, 1):
                mm = r["total_lifespan_months"]
                ws.append([i, r["name"], r["unit_of_measure"] or "",
                           mm // 12 if mm else "", r["unit_price"] or "",
                           "", "", "", "", ""])
            _auto_width_xlsx(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu ({len(items)} mặt hàng):\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        if self._active_wh_id is None:
            QMessageBox.information(self, "Nhập Excel", "Vui lòng chọn kho trước.")
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not all_rows:
                return

            hdr = [str(c or "").strip() for c in all_rows[0]]
            i_ten = _find_col(hdr, "tên hàng", "ten hang") or 1
            i_h1  = _find_col(hdr, "h1")
            i_h2  = _find_col(hdr, "h2")
            i_h3  = _find_col(hdr, "h3")
            i_h4  = _find_col(hdr, "h4")
            # Fallback nếu không có header H1-H4
            if i_h1 is None: i_h1 = 5
            if i_h2 is None: i_h2 = 6
            if i_h3 is None: i_h3 = 7
            if i_h4 is None: i_h4 = 8

            conn = database.get_conn()
            item_map = {_norm_name(r["name"]): r["id"] for r in conn.execute(
                "SELECT id, name FROM item_types WHERE is_active=1"
            ).fetchall()}
            inserted = updated = skipped = 0
            errors = []

            def _qty(row, idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() == "":
                    return 0
                try:
                    return int(float(str(v)))
                except (ValueError, TypeError):
                    return 0

            for i, row in enumerate(all_rows[1:], start=2):
                if not row or not row[i_ten]:
                    continue
                ten_hang = _norm_name(row[i_ten])
                if not ten_hang:
                    continue
                if ten_hang not in item_map:
                    errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                    skipped += 1
                    continue

                item_id = item_map[ten_hang]
                for ql, idx in [("H1", i_h1), ("H2", i_h2), ("H3", i_h3), ("H4", i_h4)]:
                    sl = _qty(row, idx)
                    try:
                        existing = conn.execute("""
                            SELECT id FROM inventory
                            WHERE warehouse_id=? AND item_type_id=?
                              AND quality_level=? AND is_shared=0
                              AND received_at_unit_date IS NULL
                            LIMIT 1
                        """, (self._active_wh_id, item_id, ql)).fetchone()
                        if sl <= 0:
                            if existing:
                                conn.execute(
                                    "UPDATE inventory SET quantity=0 WHERE id=?",
                                    (existing["id"],),
                                )
                                updated += 1
                        elif existing:
                            conn.execute(
                                "UPDATE inventory SET quantity=? WHERE id=?",
                                (sl, existing["id"]),
                            )
                            updated += 1
                        else:
                            conn.execute(
                                "INSERT INTO inventory (warehouse_id, item_type_id, quality_level, quantity)"
                                " VALUES (?, ?, ?, ?)",
                                (self._active_wh_id, item_id, ql, sl),
                            )
                            inserted += 1
                    except Exception as row_err:
                        errors.append(f"Dòng {i} {ql}: {row_err}")
                        skipped += 1

            conn.commit()
            msg = f"Thêm mới: {inserted}   Cập nhật: {updated}   Bỏ qua: {skipped}"
            if errors:
                msg += "\n\nChi tiết lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
            self.refresh()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        if self._active_wh_id is None or not self._raw:
            QMessageBox.information(self, "Xuất Excel", "Không có dữ liệu để xuất.")
            return
        wh_name = next((r["name"] for r in self._wh_list if r["id"] == self._active_wh_id), "Kho")
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Excel", f"ton_kho_{wh_name}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (wh_name or "Tồn Kho")[:31]

        rows = self._sort_rows(self._raw, search_mode=False)
        _write_xlsx_header(ws, self._COLS[1:])
        for i, r in enumerate(rows, start=1):
            mm = r["total_lifespan_months"]
            ws.append([
                i, r["item_name"], r["unit_of_measure"],
                mm // 12 if mm else "",
                r["don_gia"] or "",
                r["h1"], r["h2"], r["h3"], r["h4"], r["total"],
            ])
        _auto_width_xlsx(ws)
        wb.save(path)
        QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")

    def _on_sort(self, logical: int, ascending: bool):
        if self._was_searching:
            self._sort_s_col = logical
            self._sort_s_asc = ascending
            q = self._search.text().strip().lower()
            rows = [r for r in self._search_raw
                    if q in r["item_name"].lower()
                    or q in r["wh_name"].lower() or q in r["wh_code"].lower()]
            self._load_table(rows, search_mode=True)
        else:
            self._sort_col = logical
            self._sort_asc = ascending
            self._apply_local_filter()

    def _on_cell_dblclick(self, row: int, col: int):
        item = self._table.item(row, col)
        if item is None:
            return
        key = item.data(Qt.ItemDataRole.UserRole)
        if key is None:
            return
        wh_id, item_type_id, ql, item_name = key
        current = int(item.text()) if item.text().isdigit() else 0
        dlg = _EditQtyDialog(item_name, ql, current, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        new_val = dlg.result_value()
        if new_val is None or new_val == current:
            return
        _apply_inv_qty(wh_id, item_type_id, ql, new_val)
        self.refresh()


# ── Per-unit tab page ─────────────────────────────────────────────────────────

class ThongKeDonViPage(QWidget):
    """Tồn kho tại Đơn Vị — mỗi đơn vị là một tab, niên hạn tính từ ngày nhập đơn vị."""

    _COLS = ["", "STT", "Tên Hàng", "ĐVT", "H1", "H2", "H3", "H4", "Tổng"]
    _COLS_SEARCH = ["", "STT", "Đơn Vị", "Tên Hàng", "ĐVT",
                    "H1", "H2", "H3", "H4", "Tổng"]

    _SORT_KEYS_NORMAL: dict = {
        2: lambda r: (r["item_name"] or "").lower(),
        3: lambda r: (r["unit_of_measure"] or "").lower(),
        4: lambda r: r["h1"],
        5: lambda r: r["h2"],
        6: lambda r: r["h3"],
        7: lambda r: r["h4"],
        8: lambda r: r["total"],
    }
    _SORT_KEYS_SEARCH: dict = {
        2: lambda r: (r["wh_code"] or "").lower(),
        3: lambda r: (r["item_name"] or "").lower(),
        4: lambda r: (r["unit_of_measure"] or "").lower(),
        5: lambda r: r["h1"],
        6: lambda r: r["h2"],
        7: lambda r: r["h3"],
        8: lambda r: r["h4"],
        9: lambda r: r["total"],
    }

    def __init__(self):
        super().__init__()
        self.setStyleSheet("ThongKeDonViPage { background: #fafafa; }")
        self._wh_list: list = []
        self._active_wh_id: int | None = None
        self._raw: list = []
        self._year_map: dict = {}
        self._search_raw: list = []
        self._search_year_map: dict = {}
        self._was_searching: bool = False
        self._sort_col = -1
        self._sort_asc = True
        self._sort_s_col = -1
        self._sort_s_asc = True
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(0)

        top = QHBoxLayout()
        title = QLabel("Tồn Kho – Đơn Vị")
        title.setFont(QFont(FONT, 18, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        top.addWidget(title)
        top.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kiếm toàn đơn vị")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        top.addWidget(self._search)
        root.addLayout(top)
        root.addSpacing(16)

        # ── Searchable unit selector — hidden while cross-searching ──────────
        self._unit_row = QWidget()
        self._unit_row.setStyleSheet("background: transparent;")
        unit_row_h = QHBoxLayout(self._unit_row)
        unit_row_h.setContentsMargins(0, 0, 0, 16)
        unit_row_h.setSpacing(10)

        unit_lbl = QLabel("Đơn vị:")
        unit_lbl.setFont(QFont(FONT, 12))
        unit_lbl.setStyleSheet("color: #111;")
        unit_row_h.addWidget(unit_lbl)

        self._unit_combo = QComboBox()
        self._unit_combo.setEditable(True)
        self._unit_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self._unit_combo.setMinimumWidth(300)
        self._unit_combo.setFixedHeight(36)
        self._unit_combo.setFont(QFont(FONT, 12))
        self._unit_combo.setStyleSheet("""
            QComboBox { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 10px; background: white; color: #111; }
            QComboBox:focus { border-color: #bbb; }
            QComboBox::drop-down { border: none; width: 28px; }
            QComboBox QAbstractItemView { border: 1px solid #ddd; border-radius: 6px;
                background: white; color: #111;
                selection-background-color: #f0f0f0; selection-color: #111; }
        """)
        cpl = QCompleter(self._unit_combo.model(), self._unit_combo)
        cpl.setFilterMode(Qt.MatchFlag.MatchContains)
        cpl.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._unit_combo.setCompleter(cpl)
        self._unit_combo.currentIndexChanged.connect(self._on_unit_selected)
        unit_row_h.addWidget(self._unit_combo)
        unit_row_h.addStretch()
        root.addWidget(self._unit_row)

        # ── Summary cards ─────────────────────────────────────────────────────
        card_h = QHBoxLayout()
        card_h.setSpacing(12)
        self._cards: dict[str, _SummaryCard] = {}
        for key, label in [("total", "Tổng"), ("h1", "H1"), ("h2", "H2"),
                            ("h3", "H3"), ("h4", "H4 (Chờ TXL)")]:
            c = _SummaryCard(label, "—")
            self._cards[key] = c
            card_h.addWidget(c)
        card_h.addStretch()

        self._vis_btn = _ColVisBtn()
        card_h.addWidget(self._vis_btn)

        self._local_search = QLineEdit()
        self._local_search.setPlaceholderText("Tìm trong đơn vị này...")
        self._local_search.setFixedSize(220, 34)
        self._local_search.setFont(QFont(FONT, 12))
        self._local_search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._local_search.textChanged.connect(self._apply_local_filter)
        card_h.addWidget(self._local_search)

        root.addLayout(card_h)

        btn_row = QHBoxLayout()
        self._btn_del_sel = QPushButton("Xoá đã chọn (0)")
        self._btn_del_sel.setFixedHeight(34)
        self._btn_del_sel.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        self._btn_del_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_del_sel.setStyleSheet(
            "QPushButton { background: #d32f2f; color: white; border-radius: 8px;"
            " padding: 0 16px; border: none; }"
            " QPushButton:hover { background: #b71c1c; }"
        )
        self._btn_del_sel.clicked.connect(self._on_delete_selected)
        self._btn_del_sel.setVisible(False)
        btn_row.addWidget(self._btn_del_sel)
        btn_row.addStretch()
        for label, slot in [
            ("Tải Mẫu",    self._download_template),
            ("Nhập Excel", self._import_excel),
            ("Xuất Excel", self._export_excel),
        ]:
            btn = QPushButton(label)
            btn.setFixedHeight(34)
            btn.setFont(QFont(FONT, 12, QFont.Weight.Bold))
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setStyleSheet(_EXPORT_BTN_STYLE)
            btn.clicked.connect(slot)
            btn_row.addWidget(btn)
        root.addLayout(btn_row)
        root.addSpacing(16)

        # ── Table ─────────────────────────────────────────────────────────────
        tbl_card = QFrame()
        tbl_card.setStyleSheet(
            "QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }"
        )
        tbl_v = QVBoxLayout(tbl_card)
        tbl_v.setContentsMargins(0, 0, 0, 0)

        self._smart_hdr = _SmartHeader({0, 1})
        self._table = QTableWidget(0, len(self._COLS))
        self._table.setHorizontalHeader(self._smart_hdr)
        self._table.setHorizontalHeaderLabels(self._COLS)
        self._table.verticalHeader().setVisible(False)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)
        self._table.cellDoubleClicked.connect(self._on_cell_dblclick)
        self._smart_hdr.sortRequested.connect(self._on_sort)
        self._smart_hdr.visToggled.connect(self._vis_btn.on_vis_toggled)
        self._table.itemChanged.connect(
            lambda item: self._update_sel_count() if item.column() == 0 else None
        )
        self._smart_hdr.setSectionsClickable(True)
        self._smart_hdr.sectionClicked.connect(
            lambda col: self._toggle_all() if col == 0 else None
        )

        tbl_v.addWidget(self._table)
        root.addWidget(tbl_card, 1)

        self.refresh()

    # ── Unit selector ─────────────────────────────────────────────────────────

    def _rebuild_combo(self):
        self._unit_combo.blockSignals(True)
        self._unit_combo.clear()
        for r in self._wh_list:
            self._unit_combo.addItem(r["name"], r["id"])
        for i in range(self._unit_combo.count()):
            if self._unit_combo.itemData(i) == self._active_wh_id:
                self._unit_combo.setCurrentIndex(i)
                break
        self._unit_combo.blockSignals(False)

    def _on_unit_selected(self, index: int):
        if index < 0:
            return
        wh_id = self._unit_combo.itemData(index)
        if wh_id is None or wh_id == self._active_wh_id:
            return
        self._active_wh_id = wh_id
        self._search.clear()
        self._reload_data()

    def _configure_table(self, search_mode: bool):
        cols = self._COLS_SEARCH if search_mode else self._COLS
        self._table.setColumnCount(len(cols))
        self._table.setHorizontalHeaderLabels(cols)
        for i in range(len(cols)):
            self._table.setColumnHidden(i, False)
        h = self._table.horizontalHeader()
        specs = (
            [
                (QHeaderView.ResizeMode.Fixed, 36),
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Fixed, 90),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 56),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 80),
            ] if search_mode else [
                (QHeaderView.ResizeMode.Fixed, 36),
                (QHeaderView.ResizeMode.Fixed, 52),
                (QHeaderView.ResizeMode.Stretch, None),
                (QHeaderView.ResizeMode.Fixed, 56),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 72),
                (QHeaderView.ResizeMode.Fixed, 80),
            ]
        )
        for i, (mode, w) in enumerate(specs):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)
        skip = {0, 1}
        self._smart_hdr.set_skip(skip)
        self._vis_btn.bind(self._table, self._smart_hdr, cols, skip)
        if search_mode:
            self._sort_s_col = -1
            self._sort_s_asc = True
        else:
            self._sort_col = -1
            self._sort_asc = True
        self._smart_hdr.reset_sort()

    # ── Data ──────────────────────────────────────────────────────────────────

    def switch_to(self, wh_id: int):
        self._active_wh_id = wh_id
        self.refresh()

    def refresh(self):
        self._search_raw = []
        self._was_searching = False
        self._unit_row.setVisible(True)
        self._configure_table(search_mode=False)
        conn = database.get_conn()
        rows = conn.execute(
            "SELECT id, code, name, type FROM warehouses"
            " WHERE is_active=1 AND type='DON_VI' ORDER BY name"
        ).fetchall()
        self._wh_list = rows
        known_ids = {r["id"] for r in rows}
        if self._active_wh_id not in known_ids:
            self._active_wh_id = rows[0]["id"] if rows else None
        self._rebuild_combo()
        self._reload_data()

    def _reload_data(self):
        if self._active_wh_id is None:
            self._raw = []
            self._update_cards([])
            self._apply_local_filter()
            return
        conn = database.get_conn()
        rows = conn.execute("""
            SELECT t.id AS item_type_id,
                   t.code AS item_code, t.name AS item_name,
                   t.unit_of_measure,
                   t.total_lifespan_months,
                   SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                   SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                   SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                   SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                   SUM(i.quantity) AS total,
                   MAX(CASE
                       WHEN i.received_at_unit_date IS NOT NULL
                       THEN CAST(
                           (julianday(date('now','localtime'))
                            - julianday(i.received_at_unit_date)) / 30.44 AS INTEGER)
                       ELSE NULL
                   END) AS max_months
            FROM inventory i
            JOIN item_types t ON t.id = i.item_type_id
            WHERE i.warehouse_id = ? AND i.is_shared = 0 AND i.quantity > 0
            GROUP BY t.id
            ORDER BY t.name
        """, (self._active_wh_id,)).fetchall()
        self._raw = rows

        year = str(datetime.date.today().year)
        wh_id = self._active_wh_id
        yr_rows = conn.execute("""
            SELECT sub.item_type_id, sub.ql,
                   SUM(sub.nhap) AS nhap, SUM(sub.xuat) AS xuat
            FROM (
                SELECT tl.item_type_id, tl.quality_level_to AS ql,
                       tl.quantity AS nhap, 0 AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                WHERE tx.to_warehouse_id = ?
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_to IS NOT NULL
                UNION ALL
                SELECT tl.item_type_id, tl.quality_level_from AS ql,
                       0 AS nhap, tl.quantity AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                WHERE tx.from_warehouse_id = ?
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_from IS NOT NULL
            ) sub
            GROUP BY sub.item_type_id, sub.ql
        """, (wh_id, year, wh_id, year)).fetchall()
        self._year_map = {
            (r["item_type_id"], r["ql"]): (r["nhap"], r["xuat"])
            for r in yr_rows
        }

        self._update_cards(rows)
        self._apply_local_filter()

    def _load_search_data(self):
        conn = database.get_conn()
        year = str(datetime.date.today().year)
        self._search_raw = conn.execute("""
            SELECT w.id AS wh_id, w.code AS wh_code, w.name AS wh_name,
                   t.id AS item_type_id,
                   t.code AS item_code, t.name AS item_name,
                   t.unit_of_measure,
                   t.total_lifespan_months,
                   SUM(CASE WHEN i.quality_level='H1' THEN i.quantity ELSE 0 END) AS h1,
                   SUM(CASE WHEN i.quality_level='H2' THEN i.quantity ELSE 0 END) AS h2,
                   SUM(CASE WHEN i.quality_level='H3' THEN i.quantity ELSE 0 END) AS h3,
                   SUM(CASE WHEN i.quality_level='H4' THEN i.quantity ELSE 0 END) AS h4,
                   SUM(i.quantity) AS total,
                   MAX(CASE
                       WHEN i.received_at_unit_date IS NOT NULL
                       THEN CAST(
                           (julianday(date('now','localtime'))
                            - julianday(i.received_at_unit_date)) / 30.44 AS INTEGER)
                       ELSE NULL
                   END) AS max_months
            FROM inventory i
            JOIN warehouses w ON w.id = i.warehouse_id
            JOIN item_types t ON t.id = i.item_type_id
            WHERE i.is_shared = 0 AND i.quantity > 0
              AND w.type = 'DON_VI' AND w.is_active = 1
            GROUP BY w.id, t.id
            ORDER BY t.name, w.name
        """).fetchall()

        yr_rows = conn.execute("""
            SELECT sub.wh_id, sub.item_type_id, sub.ql,
                   SUM(sub.nhap) AS nhap, SUM(sub.xuat) AS xuat
            FROM (
                SELECT tx.to_warehouse_id AS wh_id,
                       tl.item_type_id, tl.quality_level_to AS ql,
                       tl.quantity AS nhap, 0 AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                JOIN warehouses w ON w.id = tx.to_warehouse_id
                WHERE w.type = 'DON_VI' AND w.is_active = 1
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_to IS NOT NULL
                UNION ALL
                SELECT tx.from_warehouse_id AS wh_id,
                       tl.item_type_id, tl.quality_level_from AS ql,
                       0 AS nhap, tl.quantity AS xuat
                FROM transaction_lines tl
                JOIN transactions tx ON tx.id = tl.transaction_id
                JOIN warehouses w ON w.id = tx.from_warehouse_id
                WHERE w.type = 'DON_VI' AND w.is_active = 1
                  AND strftime('%Y', tx.transaction_date) = ?
                  AND tl.quality_level_from IS NOT NULL
            ) sub
            GROUP BY sub.wh_id, sub.item_type_id, sub.ql
        """, (year, year)).fetchall()
        self._search_year_map = {
            (r["wh_id"], r["item_type_id"], r["ql"]): (r["nhap"], r["xuat"])
            for r in yr_rows
        }

    def _update_cards(self, rows):
        self._cards["total"].set_value(str(sum(r["total"] for r in rows)))
        for ql in ("h1", "h2", "h3", "h4"):
            self._cards[ql].set_value(str(sum(r[ql] for r in rows)))

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        is_searching = bool(q)

        if is_searching and not self._was_searching:
            self._load_search_data()
            self._configure_table(search_mode=True)
            self._unit_row.setVisible(False)
            self._local_search.setVisible(False)
            self._local_search.clear()
        elif not is_searching and self._was_searching:
            self._configure_table(search_mode=False)
            self._unit_row.setVisible(True)
            self._local_search.setVisible(True)

        self._was_searching = is_searching

        if is_searching:
            rows = [r for r in self._search_raw
                    if q in r["item_name"].lower()
                    or q in r["wh_name"].lower() or q in r["wh_code"].lower()]
            self._load_table(rows, search_mode=True)
        else:
            self._apply_local_filter()

    def _apply_local_filter(self):
        q = self._local_search.text().strip().lower()
        rows = self._raw if not q else [
            r for r in self._raw
            if q in r["item_name"].lower()
        ]
        self._load_table(rows, search_mode=False)

    @staticmethod
    def _fmt_months(mm: int) -> str:
        y, m = divmod(mm, 12)
        if y > 0 and m > 0:
            return f"{y} năm {m} tháng"
        return f"{y} năm" if y > 0 else f"{mm} tháng"

    def _load_table(self, rows, search_mode: bool = False):
        rows = self._sort_rows(rows, search_mode)
        self._table.blockSignals(True)
        self._table.setRowCount(0)
        _orange = QColor(204, 102, 0)

        for i, r in enumerate(rows):
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            self._table.setRowHeight(ri, 48)

            # Checkbox col 0
            chk = QTableWidgetItem()
            chk.setFlags(
                Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsUserCheckable
            )
            chk.setCheckState(Qt.CheckState.Unchecked)
            chk.setData(Qt.ItemDataRole.UserRole, r["item_type_id"])
            self._table.setItem(ri, 0, chk)

            tid = r["item_type_id"]

            def _tip(ql: str, current: int) -> str:
                if search_mode:
                    nhap, xuat = self._search_year_map.get((r["wh_id"], tid, ql), (0, 0))
                else:
                    nhap, xuat = self._year_map.get((tid, ql), (0, 0))
                so_du = current - nhap + xuat

                def _row(label, val, last=False):
                    sep = "" if last else "border-bottom: 1px solid #ebebeb;"
                    return (
                        f'<tr>'
                        f'<td style="padding: 9px 28px 9px 0; color: #111; {sep}">{label}</td>'
                        f'<td style="padding: 9px 0; color: #111; font-weight: 600;'
                        f' text-align: right; {sep}">{val}</td>'
                        f'</tr>'
                    )

                return (
                    '<table style="font-family: \'Segoe UI\'; font-size: 13px;'
                    ' min-width: 230px; border-spacing: 0;">'
                    '<tr><td colspan="2" style="padding: 0 0 10px 0;">'
                    '<b style="font-size: 14px; color: #111;">Thông tin Chi tiết</b>'
                    '</td></tr>'
                    + _row("Số dư đầu năm", so_du)
                    + _row("Tổng nhập", nhap)
                    + _row("Tổng xuất", xuat)
                    + _row("Hiện tại", current, last=True)
                    + '</table>'
                )

            if search_mode:
                cells = [
                    (str(i + 1),           True,  None,        None),
                    (r["wh_code"],         True,  None,        None),
                    (r["item_name"],       False, None,        None),
                    (r["unit_of_measure"], True,  None,        None),
                    (str(r["h1"]),         True,  None,        _tip("H1", r["h1"])),
                    (str(r["h2"]),         True,  None,        _tip("H2", r["h2"])),
                    (str(r["h3"]),         True,  None,        _tip("H3", r["h3"])),
                    (str(r["h4"]),         True,  "red" if r["h4"] > 0 else None,
                                                               _tip("H4", r["h4"])),
                    (str(r["total"]),      True,  None,        None),
                ]
            else:
                cells = [
                    (str(i + 1),           True,  None,        None),
                    (r["item_name"],       False, None,        None),
                    (r["unit_of_measure"], True,  None,        None),
                    (str(r["h1"]),         True,  None,        _tip("H1", r["h1"])),
                    (str(r["h2"]),         True,  None,        _tip("H2", r["h2"])),
                    (str(r["h3"]),         True,  None,        _tip("H3", r["h3"])),
                    (str(r["h4"]),         True,  "red" if r["h4"] > 0 else None,
                                                               _tip("H4", r["h4"])),
                    (str(r["total"]),      True,  None,        None),
                ]

            for c, (val, center, clr, tip) in enumerate(cells, start=1):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 12))
                if center:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if clr == "red":
                    cell.setForeground(Qt.GlobalColor.red)
                elif clr == "orange":
                    cell.setForeground(_orange)
                if tip:
                    cell.setToolTip(tip)
                self._table.setItem(ri, c, cell)

            # Store edit key on H1-H4 cells so double-click knows what to update
            wh_id = r["wh_id"] if search_mode else self._active_wh_id
            h_start = 5 if search_mode else 4
            for qi, ql in enumerate(("H1", "H2", "H3", "H4")):
                item = self._table.item(ri, h_start + qi)
                if item:
                    item.setData(Qt.ItemDataRole.UserRole,
                                 (wh_id, r["item_type_id"], ql, r["item_name"]))

        self._table.blockSignals(False)
        self._update_sel_count()

    def _on_cell_dblclick(self, row: int, col: int):
        item = self._table.item(row, col)
        if item is None:
            return
        key = item.data(Qt.ItemDataRole.UserRole)
        if key is None:
            return
        wh_id, item_type_id, ql, item_name = key
        current = int(item.text()) if item.text().isdigit() else 0
        dlg = _EditQtyDialog(item_name, ql, current, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        new_val = dlg.result_value()
        if new_val is None or new_val == current:
            return
        _apply_inv_qty(wh_id, item_type_id, ql, new_val)
        self.refresh()

    def _sort_rows(self, rows, search_mode: bool):
        col = self._sort_s_col if search_mode else self._sort_col
        asc = self._sort_s_asc if search_mode else self._sort_asc
        if col <= 0:
            return rows
        keys = self._SORT_KEYS_SEARCH if search_mode else self._SORT_KEYS_NORMAL
        fn = keys.get(col)
        if fn is None:
            return rows
        return sorted(rows, key=fn, reverse=not asc)

    # ── Multi-select ──────────────────────────────────────────────────────────

    def _get_checked_ids(self) -> list[int]:
        ids = []
        for row in range(self._table.rowCount()):
            item = self._table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                ids.append(item.data(Qt.ItemDataRole.UserRole))
        return ids

    def _update_sel_count(self):
        ids = self._get_checked_ids()
        n = len(ids)
        total = self._table.rowCount()
        all_checked = (total > 0 and n == total)
        hdr_label = "☑" if all_checked else "☐"
        self._table.horizontalHeaderItem(0).setText(hdr_label)
        if n > 0:
            self._btn_del_sel.setText(f"Xoá đã chọn ({n})")
            self._btn_del_sel.setVisible(True)
        else:
            self._btn_del_sel.setVisible(False)

    def _toggle_all(self):
        total = self._table.rowCount()
        if total == 0:
            return
        ids = self._get_checked_ids()
        new_state = Qt.CheckState.Unchecked if len(ids) == total else Qt.CheckState.Checked
        self._table.blockSignals(True)
        for row in range(total):
            item = self._table.item(row, 0)
            if item:
                item.setCheckState(new_state)
        self._table.blockSignals(False)
        self._update_sel_count()

    def _on_delete_selected(self):
        from PyQt6.QtWidgets import QMessageBox
        ids = self._get_checked_ids()
        if not ids:
            return
        reply = QMessageBox.question(
            self, "Xác nhận xoá",
            f"Xoá tồn kho của {len(ids)} mặt hàng đã chọn khỏi đơn vị này?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        conn = database.get_conn()
        for item_type_id in ids:
            conn.execute(
                "DELETE FROM inventory WHERE warehouse_id=? AND item_type_id=? AND is_shared=0",
                (self._active_wh_id, item_type_id),
            )
        conn.commit()
        self.refresh()

    def _download_template(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Mẫu", "mau_ton_kho_don_vi.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            conn = database.get_conn()
            items = sorted(
                conn.execute(
                    "SELECT name, unit_of_measure FROM item_types WHERE is_active=1"
                ).fetchall(),
                key=lambda r: _az_key(r["name"]),
            )
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tồn Kho Đơn Vị"
            _write_xlsx_header(ws, ["STT", "Tên hàng", "ĐVT", "H1", "H2", "H3", "H4", "Tổng", "Ngày Nhập ĐV (yyyy-mm-dd)"])
            for i, r in enumerate(items, 1):
                ws.append([i, r["name"], r["unit_of_measure"] or "", "", "", "", "", "", ""])
            _auto_width_xlsx(ws)
            wb.save(path)
            QMessageBox.information(self, "Đã lưu", f"File mẫu ({len(items)} mặt hàng):\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _import_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        if self._active_wh_id is None:
            QMessageBox.information(self, "Nhập Excel", "Vui lòng chọn đơn vị trước.")
            return
        path, _ = QFileDialog.getOpenFileName(self, "Chọn File Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from datetime import date as _date
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not all_rows:
                return

            hdr = [str(c or "").strip() for c in all_rows[0]]
            i_ten  = _find_col(hdr, "tên hàng", "ten hang") or 1
            i_h1   = _find_col(hdr, "h1")
            i_h2   = _find_col(hdr, "h2")
            i_h3   = _find_col(hdr, "h3")
            i_h4   = _find_col(hdr, "h4")
            i_ngay = _find_col(hdr, "ngày nhập", "ngay nhap")
            if i_h1 is None: i_h1 = 3
            if i_h2 is None: i_h2 = 4
            if i_h3 is None: i_h3 = 5
            if i_h4 is None: i_h4 = 6

            conn = database.get_conn()
            item_map = {_norm_name(r["name"]): r["id"] for r in conn.execute(
                "SELECT id, name FROM item_types WHERE is_active=1"
            ).fetchall()}
            inserted = updated = skipped = 0
            errors = []

            def _qty(row, idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() == "":
                    return 0
                try:
                    return int(float(str(v)))
                except (ValueError, TypeError):
                    return 0

            today = str(_date.today())

            for i, row in enumerate(all_rows[1:], start=2):
                if not row or not row[i_ten]:
                    continue
                ten_hang = _norm_name(row[i_ten])
                if not ten_hang:
                    continue
                if ten_hang not in item_map:
                    errors.append(f"Dòng {i}: không tìm thấy mặt hàng '{ten_hang}'")
                    skipped += 1
                    continue

                ngay_raw = row[i_ngay] if i_ngay is not None and len(row) > i_ngay else None
                ngay_nhap = str(ngay_raw).strip() if ngay_raw else today

                item_id = item_map[ten_hang]
                for ql, idx in [("H1", i_h1), ("H2", i_h2), ("H3", i_h3), ("H4", i_h4)]:
                    sl = _qty(row, idx)
                    try:
                        existing = conn.execute("""
                            SELECT id FROM inventory
                            WHERE warehouse_id=? AND item_type_id=?
                              AND quality_level=? AND is_shared=0
                            LIMIT 1
                        """, (self._active_wh_id, item_id, ql)).fetchone()
                        if sl <= 0:
                            if existing:
                                conn.execute(
                                    "UPDATE inventory SET quantity=0 WHERE id=?",
                                    (existing["id"],),
                                )
                                updated += 1
                        elif existing:
                            conn.execute(
                                "UPDATE inventory SET quantity=?, received_at_unit_date=? WHERE id=?",
                                (sl, ngay_nhap, existing["id"]),
                            )
                            updated += 1
                        else:
                            conn.execute(
                                "INSERT INTO inventory"
                                " (warehouse_id, item_type_id, quality_level, quantity, received_at_unit_date)"
                                " VALUES (?, ?, ?, ?, ?)",
                                (self._active_wh_id, item_id, ql, sl, ngay_nhap),
                            )
                            inserted += 1
                    except Exception as row_err:
                        errors.append(f"Dòng {i} {ql}: {row_err}")
                        skipped += 1

            conn.commit()
            msg = f"Thêm mới: {inserted}   Cộng dồn: {updated}   Bỏ qua: {skipped}"
            if errors:
                msg += "\n\nChi tiết lỗi:\n" + "\n".join(errors[:10])
            QMessageBox.information(self, "Nhập hoàn tất", msg)
            self.refresh()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        if not _check_openpyxl(self):
            return
        if self._active_wh_id is None or not self._raw:
            QMessageBox.information(self, "Xuất Excel", "Không có dữ liệu để xuất.")
            return
        wh_name = next((r["name"] for r in self._wh_list if r["id"] == self._active_wh_id), "Đơn Vị")
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Excel", f"ton_kho_{wh_name}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (wh_name or "Tồn Kho")[:31]

        rows = self._sort_rows(self._raw, search_mode=False)
        _write_xlsx_header(ws, self._COLS[1:])
        for i, r in enumerate(rows, start=1):
            ws.append([
                i, r["item_name"], r["unit_of_measure"],
                r["h1"], r["h2"], r["h3"], r["h4"], r["total"],
            ])
        _auto_width_xlsx(ws)
        wb.save(path)
        QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(rows)} dòng.\n{path}")

    def _on_sort(self, logical: int, ascending: bool):
        if self._was_searching:
            self._sort_s_col = logical
            self._sort_s_asc = ascending
            q = self._search.text().strip().lower()
            rows = [r for r in self._search_raw
                    if q in r["item_name"].lower()
                    or q in r["wh_name"].lower() or q in r["wh_code"].lower()]
            self._load_table(rows, search_mode=True)
        else:
            self._sort_col = logical
            self._sort_asc = ascending
            self._apply_local_filter()
