from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLineEdit, QComboBox, QMessageBox, QMenu, QGridLayout,
    QFileDialog,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QCursor, QAction
from database.luan_chuyen import (
    Transfer, TransferLine, get_all, get_lines, delete as transfer_delete,
)

FONT = "Segoe UI"

_TABS     = ["Luân Chuyển Giữa Các Kho", "Luân Chuyển Giữa Các Đơn Vị"]
_SUBTYPES = ["kho_kho", "dv_dv"]

_COLS = [
    "STT", "Số Phiếu", "Số M.Hàng", "Kho Nguồn",
    "Kho Đích", "Người Lập", "Ngày", "Nội Dung", "",
]

_TABLE_STYLE = """
    QTableWidget { border: none; background: white; outline: 0; gridline-color: transparent; }
    QHeaderView::section {
        background: white; color: #111; font-size: 12px;
        border: none; border-bottom: 1px solid #efefef;
        padding: 8px 12px; font-weight: normal;
    }
    QTableWidget::item {
        padding: 0 12px; border: none;
        color: #111; font-size: 13px;
    }
    QTableWidget::item:selected { background: #f0f4ff; color: #111; }
"""

_BTN_DARK = """
    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
        background: #111; color: white; font-size: 12px; font-weight: 600; }
    QPushButton:hover { background: #333; }
"""


# ── Detail panel ──────────────────────────────────────────────────────────────
class _DetailPanel(QWidget):
    def __init__(self, transfer: Transfer, lines: list[TransferLine], parent=None):
        super().__init__(parent)
        self._transfer = transfer
        self._lines = lines
        self.setAutoFillBackground(True)
        self.setStyleSheet("""
            _DetailPanel { background: #f0f0f0; }
            QLabel { background: transparent; border: none; }
        """)

        root = QVBoxLayout(self)
        root.setContentsMargins(28, 14, 28, 14)
        root.setSpacing(10)

        title = QLabel("Chi Tiết Phiếu Luân Chuyển")
        title.setFont(QFont(FONT, 13, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        root.addWidget(title)

        grid = QGridLayout()
        grid.setSpacing(4)
        grid.setHorizontalSpacing(24)

        def pair(label, value):
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            hl = QHBoxLayout(w)
            hl.setContentsMargins(0, 0, 0, 0)
            hl.setSpacing(4)
            l = QLabel(label + ":")
            l.setFont(QFont(FONT, 11))
            l.setStyleSheet("color: #888;")
            l.setFixedWidth(90)
            v = QLabel(value or "—")
            v.setFont(QFont(FONT, 11, QFont.Weight.Medium))
            v.setStyleSheet("color: #111;")
            v.setWordWrap(True)
            hl.addWidget(l)
            hl.addWidget(v, 1)
            return w

        is_kho = (transfer.subtype == "kho_kho")
        src_lbl = "Kho Nguồn" if is_kho else "ĐV Nguồn"
        dst_lbl = "Kho Đích"  if is_kho else "ĐV Đích"

        grid.addWidget(pair("Số Phiếu",  transfer.reference_number), 0, 0)
        grid.addWidget(pair(src_lbl,     transfer.from_warehouse_name), 0, 1)
        grid.addWidget(pair(dst_lbl,     transfer.to_warehouse_name), 0, 2)
        grid.addWidget(pair("Người Lập", transfer.created_by), 1, 0)
        grid.addWidget(pair("Ngày",      transfer.transaction_date), 1, 1)
        grid.addWidget(pair("Nội dung",  transfer.notes), 2, 0, 1, 3)
        root.addLayout(grid)

        sub_lbl = QLabel("Danh Sách Mặt Hàng")
        sub_lbl.setFont(QFont(FONT, 11, QFont.Weight.Bold))
        sub_lbl.setStyleSheet("color: #111;")
        root.addWidget(sub_lbl)

        show_ql = not is_kho
        sub_cols = ["STT", "Tên Hàng", "ĐVT"]
        if show_ql:
            sub_cols.append("Mức")
        sub_cols += ["Số Lượng", "Ghi Chú"]

        sub = QTableWidget(0, len(sub_cols))
        sub.setHorizontalHeaderLabels(sub_cols)
        sub.verticalHeader().setVisible(False)
        sub.setShowGrid(False)
        sub.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        sub.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        sub.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        sub.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sub.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        sub.setStyleSheet("""
            QTableWidget { border: none; background: #f8f8f8; outline: 0; border-radius: 6px; }
            QHeaderView::section { background: #f0f0f0; color: #777; font-size: 10px;
                border: none; padding: 4px 8px; }
            QTableWidget::item { padding: 4px 8px; color: #111; font-size: 12px; border: none; }
        """)
        sh = sub.horizontalHeader()
        modes = [
            (QHeaderView.ResizeMode.Fixed, 44),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 72),
        ]
        if show_ql:
            modes.append((QHeaderView.ResizeMode.Fixed, 56))
        modes += [
            (QHeaderView.ResizeMode.Fixed, 84),
            (QHeaderView.ResizeMode.Stretch, None),
        ]
        for i, (mode, w) in enumerate(modes):
            sh.setSectionResizeMode(i, mode)
            if w:
                sub.setColumnWidth(i, w)

        center_cols = {0, 2, 3, 4} if show_ql else {0, 2, 3}

        for i, line in enumerate(lines):
            r = sub.rowCount()
            sub.insertRow(r)
            sub.setRowHeight(r, 38)
            cells = [str(i + 1), line.item_name, line.unit_of_measure]
            if show_ql:
                cells.append(line.quality_level)
            cells += [str(line.quantity), line.notes]
            for c, val in enumerate(cells):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 11))
                if c in center_cols:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                sub.setItem(r, c, cell)

        sub.setFixedHeight(36 + max(1, len(lines)) * 38)
        root.addWidget(sub)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        _BTN = """
            QPushButton { border: 1px solid #ccc; border-radius: 6px; padding: 4px 16px;
                background: white; color: #333; font-size: 11px; }
            QPushButton:hover { background: #f5f5f5; }
        """
        btn_chi_tiet = QPushButton("Xuất Chi Tiết")
        btn_chi_tiet.setFont(QFont(FONT, 11))
        btn_chi_tiet.setStyleSheet(_BTN)
        btn_chi_tiet.clicked.connect(self._export_chi_tiet)
        btn_row.addWidget(btn_chi_tiet)

        btn_phieu_nhap = QPushButton("Xuất Phiếu Nhập")
        btn_phieu_nhap.setFont(QFont(FONT, 11))
        btn_phieu_nhap.setStyleSheet(_BTN)
        btn_phieu_nhap.clicked.connect(self._export_phieu_nhap)
        btn_row.addWidget(btn_phieu_nhap)
        root.addLayout(btn_row)

    def _export_chi_tiet(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Thiếu thư viện", "Vui lòng cài đặt: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Chi Tiết Luân Chuyển",
            f"chi_tiet_{self._transfer.reference_number}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chi Tiết"
            is_kho = (self._transfer.subtype == "kho_kho")
            src_lbl = "Kho Nguồn" if is_kho else "ĐV Nguồn"
            dst_lbl = "Kho Đích"  if is_kho else "ĐV Đích"
            ws.append([f"Phiếu luân chuyển: {self._transfer.reference_number}"])
            ws.append([f"{src_lbl}: {self._transfer.from_warehouse_name}   "
                       f"{dst_lbl}: {self._transfer.to_warehouse_name}   "
                       f"Ngày: {self._transfer.transaction_date}"])
            ws.append([])
            headers = ["STT", "Tên Hàng", "ĐVT"]
            if not is_kho:
                headers.append("Mức HH")
            headers += ["Số Lượng", "Ghi Chú"]
            ws.append(headers)
            for cell in ws[4]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="111111")
                cell.alignment = Alignment(horizontal="center")
            for i, line in enumerate(self._lines, 1):
                row = [i, line.item_name, line.unit_of_measure]
                if not is_kho:
                    row.append(line.quality_level)
                row += [line.quantity, line.notes]
                ws.append(row)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất {len(self._lines)} dòng.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _export_phieu_nhap(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Thiếu thư viện", "Vui lòng cài đặt: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Xuất Phiếu Nhập",
            f"phieu_nhap_{self._transfer.reference_number}.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Phiếu Nhập"
            is_kho = (self._transfer.subtype == "kho_kho")
            ws.merge_cells("A1:F1")
            ws["A1"] = "PHIẾU NHẬP KHO"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.append([])
            ws.append([f"Số phiếu: {self._transfer.reference_number}",
                       "", "", f"Ngày: {self._transfer.transaction_date}"])
            dst_lbl = "Kho Nhận" if is_kho else "Đơn Vị Nhận"
            ws.append([f"{dst_lbl}: {self._transfer.to_warehouse_name}"])
            ws.append([f"Người lập: {self._transfer.created_by}"])
            ws.append([])
            headers = ["STT", "Tên Hàng", "ĐVT"]
            if not is_kho:
                headers.append("Mức HH")
            headers += ["Số Lượng", "Ghi Chú"]
            ws.append(headers)
            hdr_row = ws.max_row
            for cell in ws[hdr_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="333333")
                cell.alignment = Alignment(horizontal="center")
            thin = Side(style="thin")
            for i, line in enumerate(self._lines, 1):
                row_data = [i, line.item_name, line.unit_of_measure]
                if not is_kho:
                    row_data.append(line.quality_level)
                row_data += [line.quantity, line.notes]
                ws.append(row_data)
                for cell in ws[ws.max_row]:
                    cell.border = Border(bottom=Side(style="hair"))
            ws.append([])
            ws.append(["Người giao", "", "Người nhận", "", "Thủ kho"])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            wb.save(path)
            QMessageBox.information(self, "Hoàn tất", f"Đã xuất phiếu nhập.\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))


# ── Main Page ─────────────────────────────────────────────────────────────────
class LuanChuyenPage(QWidget):

    def __init__(self):
        super().__init__()
        self.setStyleSheet("LuanChuyenPage { background: #fafafa; }")
        self._cache: list[Transfer] = []
        self._expanded_data_row: int | None = None
        self._active_subtype = "kho_kho"
        self._records_order: list[Transfer] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(0)

        top = QHBoxLayout()
        top.setSpacing(10)

        tab_container = QWidget()
        tab_container.setFixedHeight(36)
        tab_container.setStyleSheet("background: #f0f0f0; border-radius: 9px;")
        tab_h = QHBoxLayout(tab_container)
        tab_h.setContentsMargins(3, 3, 3, 3)
        tab_h.setSpacing(2)
        self._tab_btns: list[QPushButton] = []
        for i, txt in enumerate(_TABS):
            btn = QPushButton(txt)
            btn.setFont(QFont(FONT, 11))
            btn.setFixedHeight(30)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self._tab_btns.append(btn)
            tab_h.addWidget(btn)
            btn.clicked.connect(lambda _, i=i: self._on_tab(i))
        self._apply_tab_style(0)
        top.addWidget(tab_container)
        top.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kiếm phiếu...")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        top.addWidget(self._search)

        self._year_combo = QComboBox()
        self._year_combo.setFixedHeight(34)
        self._year_combo.setFont(QFont(FONT, 12))
        self._year_combo.setStyleSheet("""
            QComboBox { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white;
                color: #111; min-width: 120px; }
            QComboBox::drop-down { subcontrol-origin: padding;
                subcontrol-position: center right; width: 28px;
                border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px; border-bottom-right-radius: 8px;
                background: white; }
            QComboBox::down-arrow { width: 10px; height: 10px; image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent; }
            QComboBox QAbstractItemView { border: 1px solid #e0e0e0;
                border-radius: 6px; background: white; color: #111;
                selection-background-color: #f0f0f0; }
        """)
        from datetime import date as _date
        cur_year = _date.today().year
        self._year_combo.addItem("Tất cả", None)
        for y in range(cur_year, cur_year - 5, -1):
            self._year_combo.addItem(f"Năm {y}", y)
        self._year_combo.setCurrentIndex(1)
        self._year_combo.currentIndexChanged.connect(self._reload)
        top.addWidget(self._year_combo)

        root.addLayout(top)
        root.addSpacing(16)

        card = QFrame()
        card.setStyleSheet(
            "QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }"
        )
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(0, 0, 0, 0)
        card_v.setSpacing(0)

        self._table = QTableWidget(0, len(_COLS))
        self._table.setHorizontalHeaderLabels(_COLS)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setMinimumSectionSize(0)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)

        h = self._table.horizontalHeader()
        for i, (mode, w) in enumerate([
            (QHeaderView.ResizeMode.Fixed, 52),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 88),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 120),
            (QHeaderView.ResizeMode.Fixed, 100),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 52),
        ]):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)

        self._table.cellClicked.connect(self._on_cell_clicked)
        card_v.addWidget(self._table)
        root.addWidget(card, 1)
        root.addSpacing(16)

        bot = QHBoxLayout()
        bot.addStretch()
        self._btn_new = QPushButton("+ Tạo Phiếu")
        self._btn_new.setFixedHeight(38)
        self._btn_new.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        self._btn_new.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_new.setStyleSheet(_BTN_DARK)
        self._btn_new.clicked.connect(self._on_add)
        bot.addWidget(self._btn_new)
        root.addLayout(bot)

        self.refresh()

    # ── Tabs ──────────────────────────────────────────────────────────────────

    def _on_tab(self, idx: int):
        self._active_subtype = _SUBTYPES[idx]
        self._apply_tab_style(idx)
        self._search.clear()
        is_kho = (self._active_subtype == "kho_kho")
        self._table.setHorizontalHeaderItem(3, _hdr_item("Kho Nguồn" if is_kho else "ĐV Nguồn"))
        self._table.setHorizontalHeaderItem(4, _hdr_item("Kho Đích" if is_kho else "ĐV Đích"))
        self._reload()

    def _apply_tab_style(self, active_idx: int):
        for i, btn in enumerate(self._tab_btns):
            if i == active_idx:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: #111; color: white; font-size: 11px; font-weight: 600; }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: transparent; color: #111; font-size: 11px; }
                    QPushButton:hover { background: #e3e3e3; color: #111; }
                """)

    # ── Data ──────────────────────────────────────────────────────────────────

    def refresh(self):
        self._reload()

    def _reload(self):
        year = self._year_combo.currentData()
        self._cache = get_all(year=year, subtype=self._active_subtype)
        self._apply_filter()

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        data = [
            r for r in self._cache
            if not q or any(q in s for s in [
                r.reference_number.lower(),
                r.from_warehouse_name.lower(),
                r.to_warehouse_name.lower(),
                (r.created_by or "").lower(),
                (r.notes or "").lower(),
                r.transaction_date.lower(),
            ])
        ]
        self._load_table(data)

    def _load_table(self, records: list[Transfer]):
        self._expanded_data_row = None
        self._table.clearSpans()
        self._table.setRowCount(len(records) * 2)
        n = len(_COLS)
        for i, r in enumerate(records):
            dr  = i * 2
            det = i * 2 + 1
            self._table.setRowHeight(dr, 52)
            self._table.setRowHeight(det, 0)
            self._table.setSpan(det, 0, 1, n)

            cells = [
                str(i + 1), r.reference_number, str(r.line_count),
                r.from_warehouse_name, r.to_warehouse_name,
                r.created_by, r.transaction_date, r.notes,
            ]
            for c, val in enumerate(cells):
                item = QTableWidgetItem(val)
                item.setFont(QFont(FONT, 12))
                if c == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._table.setItem(dr, c, item)

            dots = QPushButton("···")
            dots.setFlat(True)
            dots.setFont(QFont(FONT, 14))
            dots.setFixedSize(40, 40)
            dots.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            dots.setStyleSheet("""
                QPushButton { color: #111; border: none; background: transparent;
                    letter-spacing: 2px; border-radius: 6px; }
                QPushButton:hover { background: #f0f0f0; color: #555; }
            """)
            rx = r
            dots.clicked.connect(lambda _, rx=rx: self._show_menu(rx))
            self._table.setCellWidget(dr, n - 1, dots)

        self._records_order = records

    # ── Interaction ───────────────────────────────────────────────────────────

    def _on_cell_clicked(self, row: int, col: int):
        if row % 2 == 1:
            return
        data_idx = row // 2
        det_row  = row + 1

        if self._expanded_data_row == row:
            self._table.setRowHeight(det_row, 0)
            self._table.removeCellWidget(det_row, 0)
            self._expanded_data_row = None
            return

        if self._expanded_data_row is not None:
            prev = self._expanded_data_row + 1
            self._table.setRowHeight(prev, 0)
            self._table.removeCellWidget(prev, 0)

        transfer = self._records_order[data_idx]
        lines = get_lines(transfer.id)
        panel = _DetailPanel(transfer, lines)
        h = 220 + max(1, len(lines)) * 38
        self._table.setCellWidget(det_row, 0, panel)
        self._table.setRowHeight(det_row, h)
        self._expanded_data_row = row

    def _show_menu(self, transfer: Transfer):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #e5e5e5;
                border-radius: 8px; padding: 4px; }
            QMenu::item { padding: 8px 20px; font-size: 13px;
                border-radius: 6px; color: #111; }
            QMenu::item:selected { background: #f5f5f5; }
            QMenu::separator { height: 1px; background: #efefef; margin: 2px 8px; }
        """)
        act_edit = QAction("Sửa", self)
        act_edit.triggered.connect(lambda: self._edit(transfer))
        act_del = QAction("Xóa", self)
        act_del.triggered.connect(lambda: self._delete(transfer))
        menu.addAction(act_edit)
        menu.addSeparator()
        menu.addAction(act_del)
        menu.exec(self._table.cursor().pos())

    def _on_add(self):
        from ui.dialogs.luan_chuyen_form import LuanChuyenFormDialog
        dlg = LuanChuyenFormDialog(self, subtype=self._active_subtype)
        if dlg.exec():
            self.refresh()

    def _edit(self, transfer: Transfer):
        from ui.dialogs.luan_chuyen_form import LuanChuyenFormDialog
        dlg = LuanChuyenFormDialog(self, transfer=transfer)
        if dlg.exec():
            self.refresh()

    def _delete(self, transfer: Transfer):
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Xóa Phiếu <b>{transfer.reference_number}</b>?<br>"
            "Tồn kho liên quan sẽ được hoàn lại.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            transfer_delete(transfer.id)
            self.refresh()


def _hdr_item(text: str) -> QTableWidgetItem:
    item = QTableWidgetItem(text)
    item.setFont(QFont(FONT, 12))
    return item
