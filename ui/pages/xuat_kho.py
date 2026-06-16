from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLineEdit, QComboBox, QMessageBox, QMenu, QGridLayout,
    QFileDialog,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QCursor, QAction
import database
from database.xuat_kho import Issue, IssueLine, get_all, get_lines, delete as issue_delete

def _get_dc_lines_for_issue(issue: Issue) -> list[IssueLine]:
    """Load hàng DC từ MUON transaction tự động tạo khi xuất đơn vị."""
    if not issue.to_warehouse_id:
        return []
    conn = database.get_conn()
    muon_rows = conn.execute("""
        SELECT id FROM transactions
        WHERE type = 'MUON'
          AND from_warehouse_id = ?
          AND to_warehouse_id = ?
          AND transaction_date = ?
          AND reference_number LIKE ?
        ORDER BY id
    """, (
        issue.from_warehouse_id,
        issue.to_warehouse_id,
        issue.transaction_date,
        issue.reference_number + "-DC%",
    )).fetchall()
    dc_lines: list[IssueLine] = []
    for row in muon_rows:
        for line in get_lines(row["id"]):
            line.is_shared = True
            dc_lines.append(line)
    return dc_lines


_STATUS_COLORS = {
    "Đã xuất":      ("#555",    "#f0f0f0"),
    "Chưa trả hết": ("#92400e", "#fef3c7"),
    "Hoàn thành":   ("#14532d", "#dcfce7"),
}
_ALL_STATUSES = ["Tất cả", "Đã xuất", "Chưa trả hết", "Hoàn thành"]

FONT = "Segoe UI"

_TABS     = ["Xuất Đi Đơn Vị", "Xuất Hàng Dùng Chung"]
_SUBTYPES = ["to_unit", "shared_loan"]

_COLS = [
    "STT", "Số Phiếu", "Số M.Hàng", "Kho Xuất",
    "Đơn Vị Nhận", "Người Lập", "Ngày Xuất", "Nội dung", "Vận Chuyển",
    "Tình Trạng", "",
]

_TABLE_STYLE = """
    QTableWidget { border: none; background: white; outline: 0; gridline-color: transparent; }
    QHeaderView::section {
        background: white; color: #111; font-size: 12px;
        border: none; border-bottom: 1px solid #efefef;
        padding: 8px 12px; font-weight: normal;
    }
    QTableWidget::item {
        padding: 0 12px; border: none;
        color: #111; font-size: 13px;
    }
    QTableWidget::item:selected { background: #f0f4ff; color: #111; }
"""

_BTN_GRAY = """
    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
        background: #555; color: white; font-size: 12px; font-weight: 600; }
    QPushButton:hover { background: #444; }
"""
_BTN_DARK = """
    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
        background: #111; color: white; font-size: 12px; font-weight: 600; }
    QPushButton:hover { background: #333; }
"""


# ── Detail panel ──────────────────────────────────────────────────────────────
class DetailPanel(QWidget):
    def __init__(self, issue: Issue, lines: list[IssueLine], parent=None):
        super().__init__(parent)
        self.setAutoFillBackground(True)
        self.setStyleSheet("""
            DetailPanel { background: #f0f0f0; }
            QLabel { background: transparent; border: none; }
        """)

        root = QVBoxLayout(self)
        root.setContentsMargins(28, 14, 28, 14)
        root.setSpacing(10)

        self._issue = issue

        show_price = (issue.subtype == "to_unit")

        # Fill in unit prices from catalog for lines that have none saved
        if show_price:
            zero_ids = list({l.item_type_id for l in lines if not l.unit_price})
            if zero_ids:
                conn = database.get_conn()
                placeholders = ",".join("?" * len(zero_ids))
                price_rows = conn.execute(
                    f"SELECT id, unit_price FROM item_types WHERE id IN ({placeholders})",
                    zero_ids,
                ).fetchall()
                price_map: dict[int, float] = {r["id"]: r["unit_price"] or 0.0 for r in price_rows}
            else:
                price_map = {}
            # Build effective prices (catalog fallback when stored price is 0)
            self._price_map = price_map
        else:
            self._price_map = {}

        self._lines = lines

        top = QHBoxLayout()
        title = QLabel("Chi Tiết Phiếu Xuất")
        title.setFont(QFont(FONT, 13, QFont.Weight.Bold))
        title.setStyleSheet("color: #111;")
        top.addWidget(title)
        top.addStretch()

        btn_excel = QPushButton("Xuất Chi Tiết")
        btn_excel.setFixedHeight(30)
        btn_excel.setFont(QFont(FONT, 11))
        btn_excel.setStyleSheet(_BTN_GRAY)
        btn_excel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_excel.clicked.connect(self._export_excel)
        top.addWidget(btn_excel)

        if issue.subtype in ("to_unit", "shared_loan"):
            top.addSpacing(6)
            btn_print = QPushButton("Xuất Phiếu Xuất")
            btn_print.setFixedHeight(30)
            btn_print.setFont(QFont(FONT, 11))
            btn_print.setStyleSheet(_BTN_DARK)
            btn_print.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_print.clicked.connect(self._export_word)
            top.addWidget(btn_print)

        root.addLayout(top)

        grid = QGridLayout()
        grid.setSpacing(4)
        grid.setHorizontalSpacing(24)

        def pair(label, value):
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            hl = QHBoxLayout(w)
            hl.setContentsMargins(0, 0, 0, 0)
            hl.setSpacing(4)
            l = QLabel(label + ":")
            l.setFont(QFont(FONT, 11))
            l.setStyleSheet("color: #111;")
            l.setFixedWidth(90)
            v = QLabel(value or "—")
            v.setFont(QFont(FONT, 11, QFont.Weight.Medium))
            v.setStyleSheet("color: #111;")
            v.setWordWrap(True)
            hl.addWidget(l)
            hl.addWidget(v, 1)
            return w

        dest       = issue.to_warehouse_name or issue.recipient
        dest_label = "Đơn Vị Nhận" if issue.subtype == "to_unit" else "Đơn Vị Mượn"

        grid.addWidget(pair("Số Phiếu",  issue.reference_number), 0, 0)
        grid.addWidget(pair("Kho Xuất",  issue.from_warehouse_name), 0, 1)
        grid.addWidget(pair(dest_label,  dest), 0, 2)
        grid.addWidget(pair("Người Lập", issue.created_by), 1, 0)
        grid.addWidget(pair("Ngày Xuất", issue.transaction_date), 1, 1)
        if issue.subtype == "to_unit":
            grid.addWidget(pair("Vận Chuyển", issue.transporter), 1, 2)
        grid.addWidget(pair("Nội dung", issue.notes), 2, 0, 1, 3)
        root.addLayout(grid)

        sub_lbl = QLabel("Danh Sách Mặt Hàng")
        sub_lbl.setFont(QFont(FONT, 11, QFont.Weight.Bold))
        sub_lbl.setStyleSheet("color: #111;")
        root.addWidget(sub_lbl)

        sub_cols = ["STT", "Tên Hàng", "ĐVT", "Số Lượng"]
        if show_price:
            sub_cols += ["Đơn Giá", "Thành Tiền"]
        sub_cols.append("Ghi Chú")

        self._sub = QTableWidget(0, len(sub_cols))
        self._sub.setHorizontalHeaderLabels(sub_cols)
        self._sub.verticalHeader().setVisible(False)
        self._sub.setShowGrid(False)
        self._sub.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sub.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._sub.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self._sub.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._sub.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._sub.setStyleSheet("""
            QTableWidget { border: none; background: #f8f8f8; outline: 0;
                border-radius: 6px; }
            QHeaderView::section {
                background: #f0f0f0; color: #777; font-size: 10px;
                border: none; padding: 4px 8px;
            }
            QTableWidget::item { padding: 4px 8px; color: #111; font-size: 12px; border: none; }
        """)
        sh = self._sub.horizontalHeader()
        modes = [
            (QHeaderView.ResizeMode.Fixed, 44),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 72),
            (QHeaderView.ResizeMode.Fixed, 84),
        ]
        if show_price:
            modes += [
                (QHeaderView.ResizeMode.Fixed, 110),
                (QHeaderView.ResizeMode.Fixed, 120),
            ]
        modes.append((QHeaderView.ResizeMode.Stretch, None))
        for i, (mode, w) in enumerate(modes):
            sh.setSectionResizeMode(i, mode)
            if w:
                self._sub.setColumnWidth(i, w)
        root.addWidget(self._sub)

        # Total row (below table) — chỉ tính hàng thường, bỏ qua hàng DC
        if show_price and lines:
            grand_total = sum(
                l.quantity * (l.unit_price or self._price_map.get(l.item_type_id, 0.0))
                for l in lines
                if not getattr(l, 'is_shared', False)
            )
            tw = QWidget()
            tw.setStyleSheet("background: transparent;")
            th = QHBoxLayout(tw)
            th.setContentsMargins(0, 4, 0, 0)
            th.addStretch()
            tl = QLabel(f"Tổng tiền: {grand_total:,.0f} đ" if grand_total else "Tổng tiền: —")
            tl.setFont(QFont(FONT, 12, QFont.Weight.Bold))
            tl.setStyleSheet("color: #111;")
            th.addWidget(tl)
            root.addWidget(tw)

        self._load_lines(lines, show_price)

    def _effective_price(self, line: IssueLine) -> float:
        return line.unit_price or self._price_map.get(line.item_type_id, 0.0)

    def _load_lines(self, lines: list[IssueLine], show_price: bool):
        self._sub.setRowCount(0)
        center_cols = {0, 3}
        if show_price:
            center_cols.update({4, 5})

        for i, line in enumerate(lines):
            r = self._sub.rowCount()
            self._sub.insertRow(r)
            self._sub.setRowHeight(r, 38)
            is_dc = getattr(line, 'is_shared', False)
            display_name = f"{line.item_name} (DC)" if is_dc else line.item_name
            cells = [str(i + 1), display_name, line.unit_of_measure, str(line.quantity)]
            if show_price:
                if is_dc:
                    cells += ["—", "—"]
                else:
                    price = self._effective_price(line)
                    total = line.quantity * price
                    cells += [
                        f"{price:,.0f}" if price else "—",
                        f"{total:,.0f}" if total else "—",
                    ]
            cells.append(line.notes or "")
            for c, val in enumerate(cells):
                cell = QTableWidgetItem(val)
                cell.setFont(QFont(FONT, 11))
                if c in center_cols:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._sub.setItem(r, c, cell)
        self._sub.setFixedHeight(36 + max(1, len(lines)) * 38)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Lỗi", "Thư viện openpyxl chưa được cài đặt.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu file Excel",
            f"PhieuXuat_{self._issue.reference_number}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phiếu Xuất"

        show_price = (self._issue.subtype == "to_unit")

        # ── Styles ───────────────────────────────────────────────────────────
        hdr_font  = Font(name="Calibri", bold=True, size=11)
        hdr_fill  = PatternFill("solid", fgColor="111111")
        hdr_font_w = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        body_font = Font(name="Calibri", size=10)
        info_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", bold=True, size=10)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")
        right  = Alignment(horizontal="right",  vertical="center")

        dest_label = "Đơn Vị Nhận" if self._issue.subtype == "to_unit" else "Đơn Vị Mượn"
        dest = self._issue.to_warehouse_name or self._issue.recipient or "—"

        # ── Info block ────────────────────────────────────────────────────────
        info_rows = [
            ("Số Phiếu:",   self._issue.reference_number),
            ("Kho Xuất:",   self._issue.from_warehouse_name),
            (dest_label + ":", dest),
            ("Người Lập:",  self._issue.created_by or "—"),
            ("Ngày Xuất:",  self._issue.transaction_date),
            ("Nội Dung:",   self._issue.notes or "—"),
        ]
        if self._issue.subtype == "to_unit":
            info_rows.insert(5, ("Vận Chuyển:", self._issue.transporter or "—"))

        for r_idx, (k, v) in enumerate(info_rows, start=1):
            ws.cell(r_idx, 1, k).font = bold_font
            ws.cell(r_idx, 2, v).font = info_font
            ws.cell(r_idx, 1).alignment = left
            ws.cell(r_idx, 2).alignment = left

        blank_row = len(info_rows) + 2

        # ── Table header ──────────────────────────────────────────────────────
        col_headers = ["STT", "Tên Hàng", "ĐVT", "Số Lượng"]
        col_widths   = [6,    36,         10,    10]
        if show_price:
            col_headers += ["Đơn Giá", "Thành Tiền"]
            col_widths   += [16, 16]
        col_headers.append("Ghi Chú")
        col_widths.append(28)

        for c_idx, (hdr, w) in enumerate(zip(col_headers, col_widths), start=1):
            cell = ws.cell(blank_row, c_idx, hdr)
            cell.font = hdr_font_w
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = w

        # ── Data rows ─────────────────────────────────────────────────────────
        grand_total = 0.0
        for i, line in enumerate(self._lines):
            row_n = blank_row + 1 + i
            is_dc = getattr(line, 'is_shared', False)
            display_name = f"{line.item_name} (DC)" if is_dc else line.item_name
            if is_dc:
                price = 0.0
                line_total = 0.0
            else:
                price = self._effective_price(line)
                line_total = line.quantity * price
                grand_total += line_total

            row_data = [i + 1, display_name, line.unit_of_measure, line.quantity]
            row_aligns = [center, left, center, center]
            if show_price:
                if is_dc:
                    row_data   += ["—", "—"]
                else:
                    row_data   += [price if price else "", line_total if line_total else ""]
                row_aligns += [right, right]
            row_data.append(line.notes or "")
            row_aligns.append(left)

            for c_idx, (val, aln) in enumerate(zip(row_data, row_aligns), start=1):
                cell = ws.cell(row_n, c_idx, val)
                cell.font = body_font
                cell.alignment = aln
                cell.border = border
                if show_price and c_idx in (5, 6) and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

        # ── Total row ─────────────────────────────────────────────────────────
        if show_price:
            total_row = blank_row + 1 + len(self._lines)
            label_col = len(col_headers) - 2
            total_col = len(col_headers) - 1
            lbl_cell = ws.cell(total_row, label_col, "Tổng tiền:")
            lbl_cell.font = bold_font
            lbl_cell.alignment = right
            tot_cell = ws.cell(total_row, total_col, grand_total if grand_total else "—")
            tot_cell.font = bold_font
            tot_cell.alignment = right
            if isinstance(grand_total, float) and grand_total:
                tot_cell.number_format = '#,##0'

        ws.row_dimensions[blank_row].height = 20
        for i in range(len(self._lines)):
            ws.row_dimensions[blank_row + 1 + i].height = 18

        try:
            wb.save(path)
            QMessageBox.information(self, "Thành công", f"Đã xuất file:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Không thể lưu file:\n{e}")

    def _export_word(self):
        from ui.word_export import export_xuat_kho_excel
        export_xuat_kho_excel(self, self._issue, self._lines)


# ── Main Page ─────────────────────────────────────────────────────────────────
class XuatKhoPage(QWidget):

    def __init__(self):
        super().__init__()
        self.setStyleSheet("XuatKhoPage { background: #fafafa; }")
        self._cache: list[Issue] = []
        self._expanded_data_row: int | None = None
        self._active_subtype = "to_unit"
        self._issues_order: list[Issue] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(36, 32, 36, 24)
        root.setSpacing(0)

        top = QHBoxLayout()
        top.setSpacing(10)

        tab_container = QWidget()
        tab_container.setFixedHeight(36)
        tab_container.setStyleSheet("background: #f0f0f0; border-radius: 9px;")
        tab_h = QHBoxLayout(tab_container)
        tab_h.setContentsMargins(3, 3, 3, 3)
        tab_h.setSpacing(2)
        self._tab_btns: list[QPushButton] = []
        for i, txt in enumerate(_TABS):
            btn = QPushButton(txt)
            btn.setFont(QFont(FONT, 11))
            btn.setFixedHeight(30)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self._tab_btns.append(btn)
            tab_h.addWidget(btn)
            btn.clicked.connect(lambda _, i=i: self._on_tab(i))
        self._apply_tab_style(0)
        top.addWidget(tab_container)
        top.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Tìm kiếm phiếu...")
        self._search.setFixedSize(260, 34)
        self._search.setFont(QFont(FONT, 12))
        self._search.setStyleSheet("""
            QLineEdit { border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 12px; background: white; color: #111; }
            QLineEdit:focus { border-color: #bbb; }
        """)
        self._search.textChanged.connect(self._apply_filter)
        top.addWidget(self._search)

        self._status_combo = QComboBox()
        self._status_combo.setFixedHeight(34)
        self._status_combo.setFont(QFont(FONT, 12))
        self._status_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white;
                color: #111; min-width: 150px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding; subcontrol-position: center right;
                width: 28px; border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px; border-bottom-right-radius: 8px;
                background: white;
            }
            QComboBox::down-arrow {
                width: 10px; height: 10px; image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #e0e0e0; border-radius: 6px;
                background: white; color: #111; selection-background-color: #f0f0f0;
            }
        """)
        for s in _ALL_STATUSES:
            self._status_combo.addItem(s)
        self._status_combo.currentIndexChanged.connect(self._apply_filter)
        self._status_combo.setVisible(False)
        top.addWidget(self._status_combo)

        self._year_combo = QComboBox()
        self._year_combo.setFixedHeight(34)
        self._year_combo.setFont(QFont(FONT, 12))
        self._year_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #e0e0e0; border-radius: 8px;
                padding: 0 32px 0 12px; background: white;
                color: #111; min-width: 120px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding; subcontrol-position: center right;
                width: 28px; border-left: 1px solid #e0e0e0;
                border-top-right-radius: 8px; border-bottom-right-radius: 8px;
                background: white;
            }
            QComboBox::down-arrow {
                width: 10px; height: 10px; image: none;
                border-top: 5px solid #111;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #e0e0e0; border-radius: 6px;
                background: white; color: #111; selection-background-color: #f0f0f0;
            }
        """)
        from datetime import date as _date
        cur_year = _date.today().year
        self._year_combo.addItem("Tất cả", None)
        for y in range(cur_year, cur_year - 5, -1):
            self._year_combo.addItem(f"Năm {y}", y)
        self._year_combo.setCurrentIndex(1)
        self._year_combo.currentIndexChanged.connect(self._reload)
        top.addWidget(self._year_combo)

        root.addLayout(top)
        root.addSpacing(16)

        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 12px; border: 1px solid #efefef; }")
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(0, 0, 0, 0)
        card_v.setSpacing(0)

        self._table = QTableWidget(0, len(_COLS))
        self._table.setHorizontalHeaderLabels(_COLS)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setMinimumSectionSize(0)
        self._table.setShowGrid(False)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self._table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._table.setStyleSheet(_TABLE_STYLE)

        h = self._table.horizontalHeader()
        for i, (mode, w) in enumerate([
            (QHeaderView.ResizeMode.Fixed, 52),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 88),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 100),
            (QHeaderView.ResizeMode.Stretch, None),
            (QHeaderView.ResizeMode.Fixed, 110),
            (QHeaderView.ResizeMode.Fixed, 120),
            (QHeaderView.ResizeMode.Fixed, 52),
        ]):
            h.setSectionResizeMode(i, mode)
            if w:
                self._table.setColumnWidth(i, w)

        self._table.cellClicked.connect(self._on_cell_clicked)
        card_v.addWidget(self._table)
        root.addWidget(card, 1)
        root.addSpacing(16)

        bot = QHBoxLayout()
        bot.addStretch()
        self._btn_new = QPushButton("+ Tạo Phiếu")
        self._btn_new.setFixedHeight(38)
        self._btn_new.setFont(QFont(FONT, 12, QFont.Weight.Bold))
        self._btn_new.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_new.setStyleSheet(_BTN_DARK)
        self._btn_new.clicked.connect(self._on_add)
        bot.addWidget(self._btn_new)
        root.addLayout(bot)

        self._update_col_visibility()
        self.refresh()

    def _on_tab(self, idx: int):
        self._active_subtype = _SUBTYPES[idx]
        self._apply_tab_style(idx)
        self._search.clear()
        self._status_combo.setCurrentIndex(0)
        self._update_col_visibility()
        dest_lbl = "Đơn Vị Mượn" if self._active_subtype == "shared_loan" else "Đơn Vị Nhận"
        self._table.setHorizontalHeaderItem(4, _hdr_item(dest_lbl))
        self._reload()

    def _apply_tab_style(self, active_idx: int):
        for i, btn in enumerate(self._tab_btns):
            if i == active_idx:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: #111; color: white; font-size: 11px; font-weight: 600; }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton { border: none; border-radius: 7px; padding: 0 14px;
                        background: transparent; color: #111; font-size: 11px; }
                    QPushButton:hover { background: #e3e3e3; color: #111; }
                """)

    def _update_col_visibility(self):
        is_shared = self._active_subtype == "shared_loan"
        self._table.setColumnHidden(8, is_shared)   # Vận Chuyển — hidden for shared
        self._table.setColumnHidden(9, not is_shared)  # Tình Trạng — only for shared
        self._status_combo.setVisible(is_shared)

    def refresh(self):
        self._reload()

    def _reload(self):
        year = self._year_combo.currentData()
        self._cache = get_all(year=year, subtype=self._active_subtype)
        self._apply_filter()

    def _apply_filter(self):
        q = self._search.text().strip().lower()
        selected_status = self._status_combo.currentText()
        data = []
        for r in self._cache:
            if q and not any(q in s for s in [
                r.reference_number.lower(),
                r.from_warehouse_name.lower(),
                r.to_warehouse_name.lower(),
                (r.recipient or "").lower(),
                (r.created_by or "").lower(),
                (r.transporter or "").lower(),
                (r.notes or "").lower(),
                r.transaction_date.lower(),
                r.status.lower(),
            ]):
                continue
            if selected_status != "Tất cả" and r.status != selected_status:
                continue
            data.append(r)
        self._load_table(data)

    def _load_table(self, issues: list[Issue]):
        self._expanded_data_row = None
        self._table.clearSpans()
        self._table.setRowCount(len(issues) * 2)
        n = len(_COLS)
        for i, r in enumerate(issues):
            dr  = i * 2
            det = i * 2 + 1
            self._table.setRowHeight(dr, 52)
            self._table.setRowHeight(det, 0)
            self._table.setSpan(det, 0, 1, n)

            dest  = r.to_warehouse_name or r.recipient
            cells = [
                str(i + 1), r.reference_number, str(r.line_count),
                r.from_warehouse_name, dest,
                r.created_by, r.transaction_date,
                r.notes, r.transporter,
            ]
            for c, val in enumerate(cells):
                item = QTableWidgetItem(val)
                item.setFont(QFont(FONT, 12))
                if c == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._table.setItem(dr, c, item)

            # Status badge (col 9) — only meaningful for shared_loan
            if r.status:
                self._table.setCellWidget(dr, 9, _make_status_badge(r.status))

            dots = QPushButton("···")
            dots.setFlat(True)
            dots.setFont(QFont(FONT, 14))
            dots.setFixedSize(40, 40)
            dots.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            dots.setStyleSheet("""
                QPushButton { color: #111; border: none; background: transparent;
                    letter-spacing: 2px; border-radius: 6px; }
                QPushButton:hover { background: #f0f0f0; color: #555; }
            """)
            rx = r
            dots.clicked.connect(lambda _, rx=rx: self._show_menu(rx))
            self._table.setCellWidget(dr, n - 1, dots)

        self._issues_order = issues

    def _on_cell_clicked(self, row: int, col: int):
        if row % 2 == 1:
            return
        data_idx = row // 2
        det_row  = row + 1

        if self._expanded_data_row == row:
            self._table.setRowHeight(det_row, 0)
            self._table.removeCellWidget(det_row, 0)
            self._expanded_data_row = None
            return

        if self._expanded_data_row is not None:
            prev = self._expanded_data_row + 1
            self._table.setRowHeight(prev, 0)
            self._table.removeCellWidget(prev, 0)

        issue = self._issues_order[data_idx]
        lines = get_lines(issue.id)
        if issue.subtype == "to_unit":
            lines = lines + _get_dc_lines_for_issue(issue)
        panel = DetailPanel(issue, lines)
        h = 200 + max(1, len(lines)) * 38 + 40
        self._table.setCellWidget(det_row, 0, panel)
        self._table.setRowHeight(det_row, h)
        self._expanded_data_row = row

    def _show_menu(self, issue: Issue):
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #e5e5e5;
                border-radius: 8px; padding: 4px; }
            QMenu::item { padding: 8px 20px; font-size: 13px;
                border-radius: 6px; color: #111; }
            QMenu::item:selected { background: #f5f5f5; }
            QMenu::separator { height: 1px; background: #efefef; margin: 2px 8px; }
        """)
        act_edit = QAction("Sửa", self)
        act_edit.triggered.connect(lambda: self._edit(issue))
        act_del = QAction("Xóa", self)
        act_del.triggered.connect(lambda: self._delete(issue))
        menu.addAction(act_edit)
        menu.addSeparator()
        menu.addAction(act_del)
        menu.exec(self._table.cursor().pos())

    def _on_add(self):
        from ui.dialogs.xuat_kho_form import XuatKhoFormDialog
        dlg = XuatKhoFormDialog(self, subtype=self._active_subtype)
        if dlg.exec():
            self.refresh()

    def _edit(self, issue: Issue):
        from ui.dialogs.xuat_kho_form import XuatKhoFormDialog
        dlg = XuatKhoFormDialog(self, issue=issue)
        if dlg.exec():
            self.refresh()

    def _delete(self, issue: Issue):
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Xóa Phiếu <b>{issue.reference_number}</b>?<br>"
            "Tồn kho liên quan sẽ được hoàn lại.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            issue_delete(issue.id)
            self.refresh()


def _hdr_item(text: str) -> QTableWidgetItem:
    item = QTableWidgetItem(text)
    item.setFont(QFont(FONT, 12))
    return item


def _make_status_badge(status: str) -> QWidget:
    fg, bg = _STATUS_COLORS.get(status, ("#555", "#f0f0f0"))
    container = QWidget()
    container.setStyleSheet("background: transparent;")
    h = QHBoxLayout(container)
    h.setContentsMargins(8, 0, 8, 0)
    h.setAlignment(Qt.AlignmentFlag.AlignCenter)
    badge = QLabel(status)
    badge.setFont(QFont(FONT, 10, QFont.Weight.Bold))
    badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
    badge.setStyleSheet(
        f"QLabel {{ background: {bg}; color: {fg}; border-radius: 8px;"
        f" padding: 3px 10px; border: 1px solid {bg}; }}"
    )
    h.addWidget(badge)
    return container
